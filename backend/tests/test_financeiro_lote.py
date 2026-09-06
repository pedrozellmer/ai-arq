# -*- coding: utf-8 -*-
"""Financeiro em LOTE: a planilha vai, o arquiteto preenche, e volta virando lançamento.

06/09/2026, Pedro: *"podemos baixar a planilha padrão para preencher os preços e subir depois né,
em lote"*. E, no mesmo pedido: *"tudo interligado"* — que é a regra nº7.

O que estes guardas cobram, cada um com o FATO na mão:
  • 💰 **a régua do dinheiro** — "1.234" vale MIL em pt-BR e um-vírgula-dois em inglês; errar isso é
    erro de 1000× EM DINHEIRO. Um parser só, no servidor, com os dois formatos e o teto;
  • 🔒 nº5 — a coluna de valor sai VAZIA no modelo; célula vazia volta como ausência, nunca 0;
  • 🔗 nº7 — linha do quantitativo grava `origem_ref_id` (o retrato vem do servidor, no mesmo
    insert); item que sumiu do quantitativo (o /add-file recria os ids) NÃO vira linha fantasma nem
    erro seco: entra como livre, com aviso em português;
  • 👀 nada é gravado sem conferência: `conferir_lote` diz o que VAI acontecer, e é isso que a tela
    mostra antes do OK;
  • 🪤 planilha torta é MENSAGEM, não exceção: arquivo ilegível, cabeçalho apagado e planilha
    gigante devolvem `erro_fatal` em português.
🧪 Controles: linha em branco não vira lançamento; linha sem mudança não vira PATCH; e o modelo
gerado é lido de volta pelo próprio leitor (ida e volta fecham).
"""
import os
import sys
from datetime import date

import pytest

_AQUI = os.path.dirname(os.path.abspath(__file__))
_BACKEND = os.path.dirname(_AQUI)
sys.path.insert(0, _BACKEND)
sys.path.insert(0, _AQUI)

import financeiro_lote as fl  # noqa: E402

ITEM_A = "11111111-1111-4111-8111-111111111111"
ITEM_B = "22222222-2222-4222-8222-222222222222"
LANC_1 = "33333333-3333-4333-8333-333333333333"

ITENS = [
    {"id": ITEM_A, "description": "Porcelanato 60x60", "quantity": 1062.0, "unit": "m2",
     "discipline": "Pisos"},
    {"id": ITEM_B, "description": "Luminária de emergência", "quantity": 14.0, "unit": "un",
     "discipline": "Elétrica"},
]
LANCAMENTOS = [
    {"id": LANC_1, "categoria": "Pisos", "descricao": "Rejunte epóxi", "origem": "livre",
     "origem_ref_id": None, "origem_quantidade": None, "origem_unidade": "",
     "fornecedor": "Cerâmica Boa", "valor": 1200.0, "forma_pagamento": "à vista",
     "venc_tipo": "data", "venc_data": "2026-10-01", "status": "contratado"},
]


# ══════════════════════════════════════════════════════════════════════════
#  💰 a régua do dinheiro — o guarda que mais importa
# ══════════════════════════════════════════════════════════════════════════
@pytest.mark.parametrize("entrada,esperado", [
    ("1.234", 1234.0),          # pt-BR: MIL (o erro de 1000× mora aqui)
    ("1.234,56", 1234.56),
    ("1.234.567", 1234567.0),
    ("1234,56", 1234.56),
    ("1234.5", 1234.5),         # sem vírgula e 1 casa: decimal
    ("1234.56", 1234.56),
    ("R$ 12.500,00", 12500.0),
    ("  48000 ", 48000.0),
    ("0", 0.0),                 # zero DIGITADO é valor
    (12500, 12500.0),
    (1234.5, 1234.5),
])
def test_valor_le_os_dois_formatos_e_nunca_erra_por_1000(entrada, esperado):
    valor, erro = fl.valor_do_texto(entrada)
    assert erro is None and valor == esperado, (entrada, valor, erro)


@pytest.mark.parametrize("entrada,pedaco", [
    ("abc", "não parece"), ("-5", "negativo"), ("1.000.000.000.000,00", "teto"),
    ("12,,3", "não parece"), (True, "inválido"), (float("nan"), "inválido"),
])
def test_CONTROLE_valor_torto_e_recusado_com_motivo(entrada, pedaco):
    valor, erro = fl.valor_do_texto(entrada)
    assert valor is None and erro and pedaco in erro


def test_celula_vazia_e_AUSENCIA_nunca_zero():
    for v in (None, "", "   ", "R$"):
        assert fl.valor_do_texto(v) == (None, None), v


# ══════════════════════════════════════════════════════════════════════════
#  ida: o modelo
# ══════════════════════════════════════════════════════════════════════════
def test_modelo_traz_o_que_ja_existe_e_os_itens_que_faltam_sem_repetir():
    linhas = fl.montar_linhas_do_modelo(ITENS, LANCAMENTOS)
    assert [l["ancora"] for l in linhas] == ["l:%s" % LANC_1, "i:%s" % ITEM_A, "i:%s" % ITEM_B]
    # item que JÁ virou lançamento não aparece duas vezes
    com_lanc = LANCAMENTOS + [{"id": "44444444-4444-4444-8444-444444444444", "categoria": "Pisos",
                               "descricao": "Porcelanato 60x60", "origem": "quantitativo",
                               "origem_ref_id": ITEM_A, "origem_quantidade": 1062.0,
                               "origem_unidade": "m2", "fornecedor": "", "valor": None,
                               "forma_pagamento": "", "venc_tipo": "fase", "status": "cotado"}]
    ancoras = [l["ancora"] for l in fl.montar_linhas_do_modelo(ITENS, com_lanc)]
    assert "i:%s" % ITEM_A not in ancoras and "i:%s" % ITEM_B in ancoras


def test_modelo_sai_com_a_coluna_de_valor_VAZIA_e_a_ressalva_dentro(tmp_path):
    from openpyxl import load_workbook
    p = str(tmp_path / "modelo.xlsx")
    fl.gerar_modelo_xlsx(fl.montar_linhas_do_modelo(ITENS, []), p,
                         {"project_name": "Casa Teste"}, hoje=date(2026, 9, 6))
    ws = load_workbook(p)[fl.ABA]
    texto = "\n".join(str(c.value) for r in ws.iter_rows() for c in r if c.value is not None)
    assert "O AI.arq não precifica obra" in texto and "Casa Teste" in texto
    assert "Deixe em branco o que ainda não sabe" in texto
    linhas = [[c.value for c in r] for r in ws.iter_rows()]
    dados = [r for r in linhas if r[2] in ("Porcelanato 60x60", "Luminária de emergência")]
    assert len(dados) == 2
    for r in dados:
        assert r[6] is None, "🔒 nº5: a coluna VALOR sai vazia — quem preenche é o arquiteto"
        assert r[3] is not None and r[4], "quantidade e unidade vêm preenchidas, pra consulta"
        assert str(r[10]).startswith("i:"), "a âncora liga a linha ao item"
    assert ws.column_dimensions["K"].hidden is True, "a âncora é maquinaria, fica escondida"


def test_ida_e_volta_fecham(tmp_path):
    """O modelo que a gente gera é lido de volta pelo nosso próprio leitor."""
    p = str(tmp_path / "m.xlsx")
    fl.gerar_modelo_xlsx(fl.montar_linhas_do_modelo(ITENS, LANCAMENTOS), p, {})
    lido = fl.ler_planilha_lote(p)
    assert lido["erro_fatal"] is None and len(lido["linhas"]) == 3
    assert [l["tipo"] for l in lido["linhas"]] == ["l", "i", "i"]
    assert lido["linhas"][0]["valor"] == 1200.0 and lido["linhas"][0]["status"] == "contratado"
    assert lido["linhas"][1]["valor"] is None


# ══════════════════════════════════════════════════════════════════════════
#  volta: ler o que veio preenchido
# ══════════════════════════════════════════════════════════════════════════
def _planilha(tmp_path, linhas_extra, nome="v.xlsx"):
    """Escreve uma planilha no formato do modelo com as linhas dadas (listas na ordem de COLS)."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = fl.ABA
    ws.append(["FINANCEIRO DA OBRA — PREENCHER OS PREÇOS"])
    ws.append(fl.COLS)
    for l in linhas_extra:
        ws.append(l)
    p = str(tmp_path / nome)
    wb.save(p)
    return p


def test_le_o_preenchido_e_recusa_planilha_torta(tmp_path):
    p = _planilha(tmp_path, [[1, "Pisos", "Porcelanato", 1062, "m2", "Cerâmica Boa", "48.000,00",
                              "3x", "2026-10-05", "Contratado", "i:%s" % ITEM_A]])
    lido = fl.ler_planilha_lote(p)
    l = lido["linhas"][0]
    assert l["valor"] == 48000.0 and l["status"] == "contratado" and l["venc"] == date(2026, 10, 5)
    assert l["tipo"] == "i" and l["ref"] == ITEM_A and not l["problemas"]

    from openpyxl import Workbook
    wb = Workbook(); wb.active.append(["nada a ver"]); q = str(tmp_path / "sem_cab.xlsx"); wb.save(q)
    assert "cabeçalho" in fl.ler_planilha_lote(q)["erro_fatal"]

    ruim = str(tmp_path / "nao_e_xlsx.xlsx")
    open(ruim, "wb").write(b"isto nao e uma planilha")
    assert "não consegui abrir" in fl.ler_planilha_lote(ruim)["erro_fatal"]

    grande = _planilha(tmp_path, [[i, "C", "x", 1, "un", "", "", "", "", "", ""] for i in range(60)], "g.xlsx")
    assert "mais de 10 linhas" in (fl.ler_planilha_lote(grande, max_linhas=10)["erro_fatal"] or "")


def test_data_em_varios_formatos_e_o_ilegivel_vira_problema(tmp_path):
    p = _planilha(tmp_path, [
        [1, "Pisos", "A", 1, "un", "", "10", "", "05/10/2026", "", "i:%s" % ITEM_A],
        [2, "Pisos", "B", 1, "un", "", "10", "", "amanhã", "", "i:%s" % ITEM_B],
    ])
    linhas = fl.ler_planilha_lote(p)["linhas"]
    assert linhas[0]["venc"] == date(2026, 10, 5) and not linhas[0]["problemas"]
    assert linhas[1]["venc"] is None and "data de vencimento ilegível" in linhas[1]["problemas"]


# ══════════════════════════════════════════════════════════════════════════
#  conferência: o que VAI acontecer, antes de gravar
# ══════════════════════════════════════════════════════════════════════════
_VIVOS = {i["id"]: i for i in ITENS}
_ATUAIS = {l["id"]: l for l in LANCAMENTOS}


def _linha(**k):
    base = {"n": 1, "ancora": "", "tipo": "novo", "ref": None, "categoria": "", "item": "",
            "fornecedor": "", "valor": None, "forma": "", "venc": None, "status": None,
            "problemas": []}
    base.update(k)
    return base


def test_cria_atualiza_ignora_e_a_frase_em_portugues():
    linhas = [
        _linha(n=1, tipo="i", ref=ITEM_A, item="Porcelanato", categoria="Pisos", valor=48000.0,
               fornecedor="Cerâmica Boa", status="contratado"),
        _linha(n=2, tipo="l", ref=LANC_1, item="Rejunte", valor=1500.0, fornecedor="Cerâmica Boa",
               forma="à vista", status="contratado", venc=date(2026, 10, 1)),
        _linha(n=3, tipo="i", ref=ITEM_B, item="Luminária"),                       # em branco
    ]
    c = fl.conferir_lote(linhas, _VIVOS, _ATUAIS)
    assert c["resumo"] == {"linhas": 3, "cria": 1, "atualiza": 1, "sem_mudanca": 0, "ignoradas": 1,
                           "sem_valor": 0, "sem_origem": 0, "com_problema": 0}
    assert c["frase"] == "1 lançamento novo, 1 atualizado e 1 em branco (ignorada)"
    cria = [a for a in c["acoes"] if a["acao"] == "cria"][0]
    assert cria["corpo"]["origem"] == "quantitativo" and cria["corpo"]["origem_ref_id"] == ITEM_A, (
        "🔗 nº7: a linha nasce amarrada ao item — é isso que faz o aviso 'item mudou' funcionar")
    assert cria["corpo"]["valor"] == 48000.0 and cria["corpo"]["status"] == "contratado"
    atualiza = [a for a in c["acoes"] if a["acao"] == "atualiza"][0]
    assert atualiza["id"] == LANC_1 and atualiza["campos"] == {"valor": 1500.0}, (
        "só o que MUDOU vira PATCH — fornecedor/forma/venc/status iguais ficam de fora")


def test_CONTROLE_linha_sem_mudanca_nao_vira_PATCH():
    igual = _linha(n=1, tipo="l", ref=LANC_1, item="Rejunte", valor=1200.0,
                   fornecedor="Cerâmica Boa", forma="à vista", status="contratado",
                   venc=date(2026, 10, 1))
    c = fl.conferir_lote([igual], _VIVOS, _ATUAIS)
    assert c["acoes"] == [] and c["resumo"]["sem_mudanca"] == 1


def test_item_que_sumiu_do_quantitativo_vira_LINHA_LIVRE_com_aviso():
    """🪤 O /add-file recria os ids: planilha baixada antes de re-subir a planta aponta pra item
    que não existe mais. Não pode virar erro seco nem lançamento fantasma."""
    sumiu = _linha(n=7, tipo="i", ref="99999999-9999-4999-8999-999999999999",
                   item="Porcelanato 60x60", categoria="Pisos", valor=48000.0)
    c = fl.conferir_lote([sumiu], _VIVOS, _ATUAIS)
    a = c["acoes"][0]
    assert a["acao"] == "cria" and a["corpo"]["origem"] == "livre"
    assert a["corpo"]["descricao"] == "Porcelanato 60x60" and a["corpo"]["valor"] == 48000.0
    assert "origem_ref_id" not in a["corpo"]
    assert c["resumo"]["sem_origem"] == 1
    assert any("não está mais no quantitativo" in x for x in c["avisos"])
    assert "sem vínculo com o quantitativo" in c["frase"]


def test_lancamento_que_sumiu_do_projeto_nao_quebra():
    fantasma = _linha(n=2, tipo="l", ref="88888888-8888-4888-8888-888888888888",
                      item="Sumiu", valor=10.0)
    c = fl.conferir_lote([fantasma], _VIVOS, _ATUAIS)
    assert c["acoes"][0]["acao"] == "cria" and c["acoes"][0]["corpo"]["origem"] == "livre"
    assert any("não existe mais no projeto" in x for x in c["avisos"])


def test_apagar_valor_de_linha_paga_manda_o_status_junto():
    """Mesma regra da tela: o servidor não rebaixa calado."""
    paga = {"id": LANC_1, "valor": 900.0, "status": "pago", "fornecedor": "", "forma_pagamento": "",
            "venc_tipo": "fase"}
    l = _linha(n=1, tipo="l", ref=LANC_1, item="x", valor=None)
    c = fl.conferir_lote([l], _VIVOS, {LANC_1: paga})
    assert c["acoes"][0]["campos"] == {"valor": None, "status": "contratado"}


def test_vencimento_amarra_na_fase_quando_a_pessoa_nao_pos_data():
    l = _linha(n=1, tipo="i", ref=ITEM_A, item="Porcelanato", categoria="Pisos", valor=10.0)
    corpo = fl.conferir_lote([l], _VIVOS, _ATUAIS)["acoes"][0]["corpo"]
    assert corpo["venc_tipo"] == "fase" and corpo["venc_fase"] == "Pisos"
    l2 = _linha(n=1, tipo="i", ref=ITEM_A, item="Porcelanato", valor=10.0, venc=date(2026, 11, 2))
    corpo2 = fl.conferir_lote([l2], _VIVOS, _ATUAIS)["acoes"][0]["corpo"]
    assert corpo2 == {**corpo2, "venc_tipo": "data", "venc_data": "2026-11-02"}


def test_total_do_lote_soma_so_o_que_tem_valor():
    acoes = [{"corpo": {"valor": 100.0}}, {"corpo": {"valor": None}}, {"campos": {"valor": 50.5}}]
    assert fl.total_das_acoes(acoes) == "R$ 150,50"
    assert fl.total_das_acoes([{"corpo": {"valor": None}}]) == "—"
