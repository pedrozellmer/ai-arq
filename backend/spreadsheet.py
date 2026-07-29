# -*- coding: utf-8 -*-
"""Gerador de planilha .xlsx de quantitativos (aba 'Orçamento' pronta pro orçamentista preencher preços)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import ProjectData, BudgetItem, Confidence


# Estilos — modernizado 16/07: fonte Calibri, paleta da marca (slate/indigo),
# amarelo suave, bordas cinza-claro. Daltonismo mantido (cor + selo de texto).
_FONT = 'Calibri'
F_TITLE = Font(name=_FONT, bold=True, size=15, color='0F172A')
F_SEC = Font(name=_FONT, bold=True, size=11, color='FFFFFF')
F_SUB = Font(name=_FONT, bold=True, size=10, color='0F172A')
F_HDR = Font(name=_FONT, bold=True, size=9, color='0F172A')
F_N = Font(name=_FONT, size=9)
F_BLUE = Font(name=_FONT, size=9, color='1D4ED8')
F_BOLD = Font(name=_FONT, bold=True, size=9)
F_TOT = Font(name=_FONT, bold=True, size=10, color='0F172A')
F_NOTE = Font(name=_FONT, size=8, italic=True, color='64748B')
F_SM = Font(name=_FONT, size=8)

P_SEC = PatternFill('solid', fgColor='0F172A')      # slate (cabeçalho de seção)
P_SUB = PatternFill('solid', fgColor='EEF2FF')      # indigo-50 (subseção)
P_HDR = PatternFill('solid', fgColor='E2E8F0')      # slate-200 (cabeçalho de coluna)
P_YEL = PatternFill('solid', fgColor='FFF0A6')      # amarelo suave (preencher preço)
P_TOT = PatternFill('solid', fgColor='E8ECF4')      # totais
P_LT = PatternFill('solid', fgColor='F8FAFC')       # linha alternada sutil
P_ORANGE = PatternFill('solid', fgColor='FFE1B3')   # estimado (laranja acessível)

AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALT = Alignment(horizontal='left', vertical='top', wrap_text=True)
AR = Alignment(horizontal='right', vertical='center')
_SD = Side(style='thin', color='D0D5DD')
BD = Border(left=_SD, right=_SD, top=_SD, bottom=_SD)


DISCIPLINE_ORDER = [
    "Serviços Preliminares",
    "Demolição e Remoção",
    "Estrutura",
    "Fechamentos Verticais",
    "Revestimentos",
    "Pisos e Rodapés",
    "Forros",
    "Portas e Ferragens",
    "Divisórias e Vidros",
    "Persianas e Cortinas",
    "Iluminação",
    "Instalações Elétricas e Dados",
    "Instalações Hidráulicas",
    "Instalações de Gás",
    "Ar-Condicionado",
    "Incêndio e Segurança",
    "Marcenaria",
    "Mobiliário",
    "Complementares",
]


def _style_row(ws, row, font, fill=None, align=None, cols=9):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        if fill: cell.fill = fill
        if align: cell.alignment = align
        cell.border = BD


# Confiança mínima pra um código SINAPI aparecer na coluna REF da planilha.
# Medido em 17/07 (ver comentário em _build_ref): com a nota honesta, match bom
# fica em 70-100% e match errado em 16-61%. Abaixo do corte o candidato não some —
# vai pra aba técnica com a descrição oficial, pra conferir em vez de copiar.
REF_MIN_CONFIDENCE = 0.70


def _unit_family(u: str) -> str:
    """Família da unidade, pra checar compatibilidade item × código SINAPI.
    Retorna '' quando a unidade é desconhecida (aí não flaga divergência)."""
    if not u:
        return ''
    u = u.strip().lower().replace('²', '2').replace('³', '3')
    if u in ('m2', 'm²'): return 'area'
    if u in ('m3',): return 'volume'
    if u in ('m', 'ml', 'metro', 'mlinear'): return 'linear'
    if u in ('un', 'und', 'unid', 'pç', 'pc', 'peca', 'peça', 'cj', 'conj', 'cjto'): return 'cont'
    if u in ('kg', 't', 'ton'): return 'peso'
    if u in ('vb', 'verba'): return 'verba'
    if u in ('l', 'lt', 'litro'): return 'liquido'
    return ''


def _build_ref_text(item) -> str:
    """Monta texto da coluna REF combinando código SINAPI + ref do projeto.

    Formato: 'SINAPI 86902 (conf. alta) · prancha original.dwg'
    Tudo opcional — só aparece o que existe.
    """
    parts = []

    # SINAPI (prioridade — referência oficial gov BR)
    sinapi_matches = getattr(item, 'sinapi_matches', None) or []
    if sinapi_matches:
        m = sinapi_matches[0]
        cod = m.get('codigo', '').strip()
        sim = m.get('similarity', 0) or 0
        # CORTE DE RUÍDO (17/07): a REF só recebe código que a gente banca.
        # O corte era 30%, mas em cima de uma nota inflada: o matcher pontuava
        # contra a palavra reduzida da busca, então "piso porcelanato 60x60" casava
        # com PISO DE BORRACHA a "100%" e passava folgado. Com a nota honesta
        # (sinapi_matcher._rescore_honest, 17/07) a separação medida ficou nítida:
        # match bom 70-100% (bacia 100, porta de vidro 81, luminária 80, split 70)
        # × match errado 16-61% (piso→borracha 16, pintura→parede 31, porta→balsa
        # 37, limpeza→revestimento 41, spot→cuba 61). Daí o corte em 70%.
        # Abaixo disso o código NÃO aparece aqui — segue na aba técnica, COM a
        # descrição oficial e o %, que é onde dá pra conferir em vez de copiar.
        # A IA olhou e disse que nenhum candidato é a mesma coisa: respeita, mesmo
        # que o texto pareça idêntico. É esse "não" que impede o SPOT → CUBA DE
        # EMBUTIR REDONDA (o texto casa; a coisa não é a mesma).
        if cod and not m.get('_llm_rejected') and sim >= REF_MIN_CONFIDENCE:
            level = m.get('_match_level', 'full')
            mark = '~' if level.startswith('simplified') else ''
            # Aviso de unidade incompatível (piso m² × código em metro linear).
            sinapi_unit = (m.get('unidade') or '').strip()
            fi, fs = _unit_family(item.unit), _unit_family(sinapi_unit)
            unit_warn = f' ⚠ unidade difere (item {item.unit} × SINAPI {sinapi_unit})' if (fi and fs and fi != fs) else ''
            parts.append(f'SINAPI {mark}{cod} (conf. alta){unit_warn}')
        elif cod:
            # Match fraco: sinaliza que há candidato SINAPI sem induzir com código.
            parts.append('SINAPI: sem correspondência forte (ver aba técnica)')

    # TCPO removido (16/07): catálogo é SINAPI, TCPO fora da fase atual.

    # Referência da prancha original (sempre por último)
    if item.ref_sheet:
        # Encurta nome se muito longo
        ref = item.ref_sheet[:35] + '...' if len(item.ref_sheet) > 35 else item.ref_sheet
        parts.append(ref)

    return ' · '.join(parts) if parts else ''


_TYPOLOGY_LABEL = {
    "office":      "ESCRITÓRIO / CORPORATIVO",
    "residential": "RESIDENCIAL",
    "retail":      "COMERCIAL / VAREJO",
    "hospital":    "HOSPITALAR / SAÚDE",
    "educational": "EDUCACIONAL",
}


def generate_spreadsheet(project: ProjectData, items: list[BudgetItem],
                         output_path: str, typology: str = "office"):
    """Gera a planilha .xlsx completa.

    `typology` define o cabeçalho da análise comparativa e regras de
    sugestões — evita que projeto residencial apareça como "reforma de
    escritório" no título, ou que itens corporativos (controle de acesso,
    iPad de agendamento) sejam sugeridos pra projetos que não fazem sentido.
    """
    wb = Workbook()

    # ================================================================
    # SHEET 1: RESUMO
    # ================================================================
    ws1 = wb.active
    ws1.title = 'Resumo Comparativo'
    ws1.sheet_properties.tabColor = '4F46E5'
    ws1.column_dimensions['A'].width = 4
    ws1.column_dimensions['B'].width = 90

    r = 1
    def add_title(text):
        nonlocal r
        ws1.merge_cells(f'A{r}:B{r}')
        ws1.cell(row=r, column=1, value=text).font = Font(name='Calibri', bold=True, size=12, color='4F46E5')
        r += 1

    def add_line(text, bold=False, fill=None):
        nonlocal r
        ws1.merge_cells(f'A{r}:B{r}')
        c = ws1.cell(row=r, column=1, value=text)
        c.font = Font(name='Calibri', bold=bold, size=10)
        c.alignment = ALT
        if fill: c.fill = fill
        r += 1

    def add_section(text):
        nonlocal r
        ws1.merge_cells(f'A{r}:B{r}')
        c = ws1.cell(row=r, column=1, value=text)
        c.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        c.fill = P_SEC
        c.alignment = AL
        r += 1

    # Disclaimer obrigatório na A1 (regras duras #1 e #5):
    # cliente vê SEMPRE no topo, mesmo se rolar pra baixo perde de vista.
    # Cor amarela forte pra não passar batido. Antes ficava só no email/site —
    # quando a planilha era reencaminhada pro orçamentista/cliente final, a
    # ressalva sumia.
    from openpyxl.styles import PatternFill as _PF
    _disclaimer_fill = _PF(start_color='FFF7CC', end_color='FFF7CC', fill_type='solid')
    ws1.merge_cells(f'A{r}:B{r}')
    _cell_disc = ws1.cell(row=r, column=1, value=(
        '⚠️ O AI.arq é uma IA e pode cometer erros — confira as informações importantes. '
        'Quantitativo gerado por IA, NÃO é orçamento. '
        'Revisão por arquiteto ou engenheiro habilitado é obrigatória antes do uso. '
        'Itens em LARANJA são sugestões da IA — confira contra o projeto antes de mandar pros fornecedores.'
    ))
    _cell_disc.font = Font(name='Calibri', bold=True, size=10, color='8B4A0F')
    _cell_disc.fill = _disclaimer_fill
    _cell_disc.alignment = ALT
    ws1.row_dimensions[r].height = 48
    r += 1
    r += 1  # respiro visual

    _titulo = _TYPOLOGY_LABEL.get(typology, "INTERIORES")
    add_title(f'QUANTITATIVO — {_titulo}')
    r += 1
    if project.name:
        add_line(f'Projeto: {project.name}', bold=True)
    if project.address:
        add_line(f'Endereço: {project.address}')
    if project.architect:
        add_line(f'Arquitetura: {project.architect}')
    add_line(f'Fase: {project.phase}')
    # Exibir áreas apenas se a IA conseguiu extrair. Se layout_area == 0 e
    # total_area > 0, fallback: assumir que toda a laje é área de intervenção
    # (caso típico de residência sem distinção de zonas).
    if project.total_area or project.layout_area:
        _total = project.total_area or 0
        _layout = project.layout_area or 0
        _nointer = project.no_intervention_area or 0
        if getattr(project, 'total_area_source', '') == 'informado':
            # Área veio do cliente (não medida) — rotula sem falar "laje bruta medida"
            add_line(f'Área total: {_total:,.1f} m² (informada por você — não medida pela planta)', bold=True)
        elif _total and not _layout:
            # layout=0 vira "a confirmar" visível em vez de número enganoso
            add_line(f'Área laje bruta: {_total:,.1f} m² | Área layout: a confirmar | Sem intervenção: {_nointer:,.1f} m²')
        else:
            add_line(f'Área laje bruta: {_total:,.1f} m² | Área layout: {_layout:,.1f} m² | Sem intervenção: {_nointer:,.1f} m²')
    # "Posições de trabalho" só faz sentido em escritório/escola.
    # Em residencial/hospital/varejo é campo sem significado.
    if project.workstations and typology in ("office", "educational"):
        add_line(f'Posições de trabalho: {project.workstations}')
    r += 1

    # Aviso "considera apenas o que MUDA" só faz sentido em REFORMA.
    # Detectamos reforma pela presença de notas de demolição. Se não tem
    # demolição, é provavelmente obra nova — tirar o aviso.
    if project.demolition_notes:
        if typology == "residential":
            add_line('ATENÇÃO: Reforma residencial. Quantitativos consideram apenas o que MUDA.', bold=True)
        elif typology == "office":
            add_line('ATENÇÃO: Reforma de andar existente. Quantitativos consideram apenas o que MUDA.', bold=True)
        else:
            add_line('ATENÇÃO: Esta é uma reforma. Quantitativos consideram apenas o que MUDA.', bold=True)
        r += 1

    # "DEPARTAMENTOS" só faz sentido em escritório/educacional; em residencial
    # o conceito é "AMBIENTES". Gate por tipologia pra evitar titular uma
    # residência como "Departamentos — Cozinha".
    if project.departments:
        _header_deps = 'DEPARTAMENTOS' if typology in ("office", "educational") else 'AMBIENTES'
        add_title(_header_deps)
        for dept in project.departments:
            name = dept.get('name', '')
            positions = dept.get('positions', 0)
            if typology in ("office", "educational") and positions:
                add_line(f'  {name}: {positions} posições')
            else:
                add_line(f'  {name}')
        r += 1

    if project.demolition_notes:
        add_title('DEMOLIÇÃO — O QUE SAI')
        add_section('Notas importantes das pranchas de demolição')
        for note in project.demolition_notes:
            add_line(f'  >> {note}', bold=True, fill=PatternFill('solid', fgColor='FEE2E2'))
        r += 1

    if project.new_rooms:
        add_title('LAYOUT NOVO — O QUE ENTRA')
        for room in project.new_rooms:
            if isinstance(room, dict):
                name = room.get('name', 'Ambiente')
                pd = room.get('ceiling_height', 'a definir')
                area = room.get('area', 'a definir')
                if pd and area and str(pd) != '' and str(area) != '':
                    add_line(f'  • {name} — PD={pd}, ~{area} m²', fill=PatternFill('solid', fgColor='DCFCE7'))
                else:
                    add_line(f'  • {name}', fill=PatternFill('solid', fgColor='DCFCE7'))
            else:
                add_line(f'  • {room}', fill=PatternFill('solid', fgColor='DCFCE7'))
        r += 1

    if project.kept_elements:
        add_title('O QUE PERMANECE')
        for elem in project.kept_elements:
            add_line(f'  • {elem}', fill=PatternFill('solid', fgColor='FFE0B2'))

    # Avisos técnicos do motor (prancha órfã, legenda ausente, etc).
    # Aparecem destacados em âmbar pra chamar atenção do orçamentista.
    if project.warnings:
        r += 1
        add_title('⚠ AVISOS DO MOTOR — REVISAR')
        for w in project.warnings:
            add_line(f'  ⚠ {w}', bold=True,
                     fill=PatternFill('solid', fgColor='FEF3C7'))

    # ================================================================
    # SHEET 2: ORÇAMENTO
    # ================================================================
    ws = wb.create_sheet('Orçamento')
    ws.sheet_properties.tabColor = '4F46E5'

    widths = [7, 62, 5, 8, 13, 13, 15, 35, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Cabeçalho
    ws.merge_cells('A1:I1')
    ws.cell(row=1, column=1, value='PLANILHA DE QUANTITATIVOS PARA CONCORRÊNCIA — AI.arq').font = F_TITLE
    ws.merge_cells('A2:I2')
    info_parts = []
    if project.name: info_parts.append(project.name)
    if project.architect: info_parts.append(project.architect)
    # Áreas aparecem na seção PREMISSAS (se extraídas) — não duplicar no subtítulo
    ws.cell(row=2, column=1, value=' | '.join(info_parts) if info_parts else 'Quantitativos de projeto').font = F_N
    ws.merge_cells('A3:I3')
    ws.cell(row=3, column=1, value='Cada item traz na coluna OBSERVAÇÕES um selo de status: "✓ MEDIDO do CAD" (confiável) ou "⚠ ESTIMADO — revisar". A cor é só reforço — fundo BRANCO = medido · LARANJA = estimado · CINZA = metadado · ROXO = custo indireto/gestão. Coluna AMARELA = preencher preço. Itens ⚠ ESTIMADO e os roxos exigem revisão antes de fechar o orçamento.').font = F_NOTE

    ro = 5
    hdrs = ['ITEM', 'DESCRIÇÃO DO SERVIÇO', 'UN', 'QTDE', 'MAT (R$)', 'M.O. (R$)', 'TOTAL (R$)', 'OBSERVAÇÕES', 'REF.']
    for c, h in enumerate(hdrs, 1):
        cl = ws.cell(row=ro, column=c, value=h)
        cl.font = F_HDR; cl.fill = P_HDR; cl.alignment = AC; cl.border = BD

    ro = 6

    # SEÇÃO 0: PREMISSAS
    ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=9)
    ws.cell(row=ro, column=1, value='0. PREMISSAS')
    _style_row(ws, ro, F_SEC, P_SEC, AL, 9)
    ro += 1

    premissas = []
    if project.total_area:
        if getattr(project, 'total_area_source', '') == 'informado':
            premissas.append(('0.1', 'Área total do projeto — INFORMADA POR VOCÊ (não medida pela planta)', 'm²',
                              project.total_area, 'Base para itens de área (piso/forro/pintura) — confira antes de orçar', ''))
        else:
            premissas.append(('0.1', 'Área construída — perímetro externo da laje', 'm²', project.total_area, '', ''))
    if project.no_intervention_area:
        # "core" é jargão de escritório (elevadores/escadas/banheiros comuns)
        label_noint = 'Área sem intervenção (core)' if typology == "office" else 'Área sem intervenção'
        premissas.append(('0.2', label_noint, 'm²', project.no_intervention_area, '', ''))
    if project.layout_area:
        premissas.append(('0.3', 'Área utilizada para layout / intervenção', 'm²', project.layout_area, '', ''))
    if project.workstations and typology in ("office", "educational"):
        premissas.append(('0.4', 'Posições de trabalho', 'un', project.workstations, 'Conforme quadro de departamentos', ''))

    # Premissas são metadados do projeto (não itens orçáveis) — fill cinza claro
    P_PREMISSA = PatternFill('solid', fgColor='F3F4F6')
    for num, desc, un, qtd, obs, ref in premissas:
        ws.cell(row=ro, column=1, value=num).font = F_N
        ws.cell(row=ro, column=2, value=desc).font = Font(name='Calibri', size=9, italic=True, color='6B7280')
        ws.cell(row=ro, column=3, value=un).font = F_N
        ws.cell(row=ro, column=4, value=qtd).font = F_N
        ws.cell(row=ro, column=8, value=obs or 'Metadado do projeto — revisar no arquivo original').font = Font(name='Calibri', size=8, italic=True, color='6B7280')
        ws.cell(row=ro, column=9, value=ref).font = Font(name='Calibri', size=7)
        for c in range(1, 10):
            ws.cell(row=ro, column=c).border = BD
            ws.cell(row=ro, column=c).alignment = AC if c in [1, 3, 4, 9] else AL
            ws.cell(row=ro, column=c).fill = P_PREMISSA
        ws.cell(row=ro, column=4).alignment = AR
        ro += 1

    ro += 1  # Linha vazia após premissas
    subtotal_rows = []

    # Agrupar itens por disciplina (deduplicar descrições similares)
    items_by_discipline = {}
    seen_descriptions = set()
    for item in items:
        disc = item.discipline or "Complementares"
        # Deduplicar por descrição normalizada
        desc_key = item.description.lower().strip()[:50]
        if desc_key in seen_descriptions:
            continue
        seen_descriptions.add(desc_key)

        if disc not in items_by_discipline:
            items_by_discipline[disc] = []
        items_by_discipline[disc].append(item)

    # Numerar disciplinas na ordem correta
    disc_num = 1
    for disc_name in DISCIPLINE_ORDER:
        disc_items = items_by_discipline.pop(disc_name, None)
        if not disc_items:
            continue

        # Cabeçalho da seção
        ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=9)
        ws.cell(row=ro, column=1, value=f'{disc_num}. {disc_name.upper()}')
        _style_row(ws, ro, F_SEC, P_SEC, AL, 9)
        ro += 1

        section_start = ro
        for idx, item in enumerate(disc_items, 1):
            item_num = f'{disc_num}.{idx}'
            ws.cell(row=ro, column=1, value=item_num).font = F_N
            ws.cell(row=ro, column=2, value=item.description).font = F_N
            ws.cell(row=ro, column=3, value=item.unit).font = F_N
            # qty=0 vira célula vazia (IA não tinha número e marcou como estimado)
            ws.cell(row=ro, column=4, value=(item.quantity if item.quantity else None)).font = F_BLUE
            ws.cell(row=ro, column=5).font = F_BLUE; ws.cell(row=ro, column=5).fill = P_YEL
            ws.cell(row=ro, column=6).font = F_BLUE; ws.cell(row=ro, column=6).fill = P_YEL
            ws.cell(row=ro, column=7, value=f'=D{ro}*(E{ro}+F{ro})').font = F_N
            # Selo de status TEXTUAL na observação. A cor (laranja/branco)
            # sozinha não basta — usuário daltônico não distingue. Cor + ícone
            # + texto (regra de acessibilidade 2026-05-21).
            # FAIL-SAFE: só CONFIRMADO vira branco/MEDIDO; todo o resto laranja.
            # E item de origem 'vision_pdf' NUNCA é "medido do CAD" — Vision lê
            # número numa imagem, não mede geometria.
            _medido = (item.confidence == Confidence.CONFIRMADO
                       and getattr(item, 'origem', '') != 'vision_pdf')
            _selo = ('✓ MEDIDO do CAD' if _medido else '⚠ ESTIMADO — revisar')
            _obs = f'{_selo}. {item.observations}' if item.observations else _selo
            ws.cell(row=ro, column=8, value=_obs).font = F_N
            # Enriquecer REF com código SINAPI (se houver match)
            ref_text = _build_ref_text(item)
            ws.cell(row=ro, column=9, value=ref_text).font = Font(name='Calibri', size=7)

            for c in range(1, 10):
                ws.cell(row=ro, column=c).border = BD
                ws.cell(row=ro, column=c).alignment = AC if c in [1, 3, 4, 9] else AL
            for c in [4, 5, 6, 7]:
                ws.cell(row=ro, column=c).alignment = AR
            for c in [5, 6, 7]:
                ws.cell(row=ro, column=c).number_format = '#,##0.00'

            # Marcar itens não-medidos (estimados) em laranja
            if not _medido:
                for c in [1, 2, 3, 4]:
                    ws.cell(row=ro, column=c).fill = P_ORANGE

            ro += 1

        # Subtotal
        section_end = ro - 1
        ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=6)
        ws.cell(row=ro, column=1, value=f'SUBTOTAL {disc_num} — {disc_name.upper()}')
        ws.cell(row=ro, column=7, value=f'=SUM(G{section_start}:G{section_end})')
        ws.cell(row=ro, column=7).number_format = '#,##0.00'
        _style_row(ws, ro, F_BOLD, P_LT, AR, 9)
        ws.cell(row=ro, column=1).alignment = AL
        subtotal_rows.append(ro)
        ro += 1

        disc_num += 1

    # Itens de disciplinas não mapeadas
    for disc_name, disc_items in items_by_discipline.items():
        ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=9)
        ws.cell(row=ro, column=1, value=f'{disc_num}. {disc_name.upper()}')
        _style_row(ws, ro, F_SEC, P_SEC, AL, 9)
        ro += 1

        section_start = ro
        for idx, item in enumerate(disc_items, 1):
            item_num = f'{disc_num}.{idx}'
            ws.cell(row=ro, column=1, value=item_num).font = F_N
            ws.cell(row=ro, column=2, value=item.description).font = F_N
            ws.cell(row=ro, column=3, value=item.unit).font = F_N
            # qty=0 vira célula vazia (IA não tinha número e marcou como estimado)
            ws.cell(row=ro, column=4, value=(item.quantity if item.quantity else None)).font = F_BLUE
            ws.cell(row=ro, column=5).font = F_BLUE; ws.cell(row=ro, column=5).fill = P_YEL
            ws.cell(row=ro, column=6).font = F_BLUE; ws.cell(row=ro, column=6).fill = P_YEL
            ws.cell(row=ro, column=7, value=f'=D{ro}*(E{ro}+F{ro})').font = F_N
            # Selo de status TEXTUAL na observação (cor + ícone + texto).
            # Fail-safe: só CONFIRMADO+geometria vira branco; resto laranja.
            _medido = (item.confidence == Confidence.CONFIRMADO
                       and getattr(item, 'origem', '') != 'vision_pdf')
            _selo = ('✓ MEDIDO do CAD' if _medido else '⚠ ESTIMADO — revisar')
            _obs = f'{_selo}. {item.observations}' if item.observations else _selo
            ws.cell(row=ro, column=8, value=_obs).font = F_N
            # Enriquecer REF com código SINAPI (se houver match)
            ref_text = _build_ref_text(item)
            ws.cell(row=ro, column=9, value=ref_text).font = Font(name='Calibri', size=7)
            for c in range(1, 10):
                ws.cell(row=ro, column=c).border = BD
                ws.cell(row=ro, column=c).alignment = AC if c in [1, 3, 4, 9] else AL
            for c in [4, 5, 6, 7]: ws.cell(row=ro, column=c).alignment = AR
            for c in [5, 6, 7]: ws.cell(row=ro, column=c).number_format = '#,##0.00'
            if item.confidence in [Confidence.ESTIMADO, Confidence.VERIFICAR]:
                for c in [1, 2, 3, 4]: ws.cell(row=ro, column=c).fill = P_ORANGE
            ro += 1

        section_end = ro - 1
        ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=6)
        ws.cell(row=ro, column=1, value=f'SUBTOTAL {disc_num} — {disc_name.upper()}')
        ws.cell(row=ro, column=7, value=f'=SUM(G{section_start}:G{section_end})')
        ws.cell(row=ro, column=7).number_format = '#,##0.00'
        _style_row(ws, ro, F_BOLD, P_LT, AR, 9)
        ws.cell(row=ro, column=1).alignment = AL
        subtotal_rows.append(ro)
        ro += 1
        disc_num += 1

    # ================================================================
    # SEÇÃO: SUGESTÕES POR TIPO DE PROJETO
    # ================================================================
    ro += 1
    ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=9)
    ws.cell(row=ro, column=1, value='SUGESTÕES POR TIPO DE PROJETO (itens que NÃO aparecem nas pranchas)')
    _style_row(ws, ro, Font(name='Calibri', bold=True, size=11, color='FFFFFF'), PatternFill('solid', fgColor='7B2D8E'), AL, 9)
    ro += 1

    ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=9)
    ws.cell(row=ro, column=1, value='Itens que não constam nas pranchas — são custos de gestão e execução típicos de obras.').font = Font(name='Calibri', size=8, italic=True, color='7B2D8E')
    ro += 1

    P_PURPLE = PatternFill('solid', fgColor='F3E8FF')

    # Checklist de itens típicos de obra — sem quantidades hardcoded.
    # O orçamentista preenche conforme o projeto específico.
    # Sugestões BASE — aplicáveis a qualquer obra (residencial, office etc)
    suggestions = [
        ('S.1', 'Equipe técnica — Gerente de contrato / PMO', 'mês', None, 'Preencher conforme prazo da obra'),
        ('S.2', 'Equipe técnica — Engenheiro de campo residente', 'mês', None, 'Preencher conforme prazo da obra'),
        ('S.3', 'Equipe técnica — Engenheiro de instalações', 'mês', None, 'Se houver instalações complexas'),
        ('S.4', 'Equipe técnica — Mestre de obras residente', 'mês', None, 'Preencher conforme prazo da obra'),
        ('S.5', 'Equipe técnica — Técnico de Segurança do Trabalho', 'mês', None, 'Visita semanal típica'),
        ('S.6', 'Equipe técnica — Auxiliar administrativo', 'mês', None, 'Apoio administrativo de obra'),
        ('S.7', 'Serventia — ajudante geral de obra (seg-sex)', 'dia', None, 'Preencher conforme prazo da obra'),
        ('S.8', 'Caçambas de entulho (classe A + classe C)', 'un', None, 'Conforme volume de resíduos do projeto'),
        ('S.9', 'Limpeza permanente de obra', 'dia', None, 'Preencher conforme prazo da obra'),
        ('S.10', 'Limpeza fina pré-entrega', 'm²', None, 'Área total de intervenção'),
        ('S.11', 'Seguro de obra e responsabilidade civil', 'vb', None, 'Valor conforme porte da obra'),
        ('S.12', 'As-built (elétrica, hidráulica, AC quando houver)', 'vb', None, 'Conforme escopo do projeto'),
        ('S.13', 'Fee / Administração de obra', '%', None, 'Percentual conforme contrato'),
        ('S.14', 'Impostos sobre faturamento', '%', None, 'Conforme regime tributário'),
        ('S.15', 'Gerenciamento de terceiros (marcenaria, divisórias, acabamentos)', 'vb', None, 'Quando houver terceiros no escopo'),
    ]

    # Sugestões ESPECÍFICAS por tipologia (só entram se fizerem sentido pra
    # essa obra). Não contaminam residencial com "iPad de reuniões" nem
    # corporativo com "móveis de área gourmet".
    if typology == "office":
        suggestions += [
            ('S.16', 'Certificação de todos os pontos elétricos', 'vb', None, 'Verificar exigência do condomínio corporativo'),
            ('S.17', 'Termografia de quadros elétricos (OPCIONAL)', 'vb', None, 'Opcional — verificar necessidade'),
            ('S.18', 'FM-200 gás inerte para CPD (OPCIONAL)', 'vb', None, 'Depende do projeto de PPCI'),
            ('S.19', 'Transporte vertical de mobiliário (entre andares)', 'vb', None, 'Se mobiliário armazenado em outro andar'),
        ]
    elif typology == "residential":
        suggestions += [
            ('S.16', 'Taxa de obra do condomínio', 'vb', None, 'Conforme regulamento do condomínio'),
            ('S.17', 'Instalação de ar-condicionado split (se não especificado)', 'un', None, 'Somente se não constar no projeto elétrico'),
            ('S.18', 'Projeto executivo de iluminação residencial', 'vb', None, 'Se não contratado separadamente'),
        ]
    elif typology == "hospital":
        suggestions += [
            ('S.16', 'Certificação sanitária (Anvisa)', 'vb', None, 'Conforme regulamento local'),
            ('S.17', 'Gases medicinais (instalação e teste)', 'vb', None, 'Se aplicável ao projeto'),
        ]
    elif typology == "retail":
        suggestions += [
            ('S.16', 'Comunicação visual / fachada', 'vb', None, 'Se não contratado separadamente'),
            ('S.17', 'Sistema de alarme e CFTV', 'vb', None, 'Conforme padrão da operação'),
        ]

    section_start_sug = ro
    for num, desc, un, qtd, obs in suggestions:
        ws.cell(row=ro, column=1, value=num).font = F_N
        ws.cell(row=ro, column=2, value=desc).font = F_N
        ws.cell(row=ro, column=3, value=un).font = F_N
        ws.cell(row=ro, column=4, value=qtd).font = F_BLUE
        ws.cell(row=ro, column=5).font = F_BLUE; ws.cell(row=ro, column=5).fill = P_YEL
        ws.cell(row=ro, column=6).font = F_BLUE; ws.cell(row=ro, column=6).fill = P_YEL
        if un == '%':
            ws.cell(row=ro, column=7, value='Calcular sobre o total').font = F_N
        else:
            ws.cell(row=ro, column=7, value=f'=D{ro}*(E{ro}+F{ro})').font = F_N
        ws.cell(row=ro, column=8, value=obs).font = F_N
        ws.cell(row=ro, column=9, value='Experiência').font = Font(name='Calibri', size=7)
        for c in range(1, 10):
            ws.cell(row=ro, column=c).border = BD
            ws.cell(row=ro, column=c).alignment = AC if c in [1, 3, 4, 9] else AL
            ws.cell(row=ro, column=c).fill = P_PURPLE
        ws.cell(row=ro, column=4).alignment = AR
        for c in [5, 6, 7]:
            ws.cell(row=ro, column=c).alignment = AR
            ws.cell(row=ro, column=c).number_format = '#,##0.00'
        ro += 1

    section_end_sug = ro - 1
    ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=6)
    ws.cell(row=ro, column=1, value='SUBTOTAL SUGESTÕES (custos indiretos e gestão)')
    ws.cell(row=ro, column=7, value=f'=SUM(G{section_start_sug}:G{section_end_sug})')
    ws.cell(row=ro, column=7).number_format = '#,##0.00'
    _style_row(ws, ro, F_BOLD, PatternFill('solid', fgColor='E9D5FF'), AR, 9)
    ws.cell(row=ro, column=1).alignment = AL
    subtotal_rows.append(ro)
    ro += 1

    # Resumo
    ro += 1
    ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=9)
    ws.cell(row=ro, column=1, value='RESUMO GERAL')
    _style_row(ws, ro, F_SEC, P_SEC, AL, 9)
    ro += 1

    resumo_start = ro
    for st_row in subtotal_rows:
        ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=6)
        # Copiar label do subtotal
        label = ws.cell(row=st_row, column=1).value or ""
        ws.cell(row=ro, column=1, value=label)
        ws.cell(row=ro, column=7, value=f'=G{st_row}')
        ws.cell(row=ro, column=7).number_format = '#,##0.00'
        _style_row(ws, ro, F_N, None, None, 9)
        ws.cell(row=ro, column=1).alignment = AL
        ws.cell(row=ro, column=7).alignment = AR
        ro += 1
    resumo_end = ro - 1

    # Total direto
    ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=6)
    ws.cell(row=ro, column=1, value='TOTAL CUSTO DIRETO (sem BDI)')
    ws.cell(row=ro, column=7, value=f'=SUM(G{resumo_start}:G{resumo_end})')
    ws.cell(row=ro, column=7).number_format = '#,##0.00'
    _style_row(ws, ro, F_TOT, P_TOT, AR, 9)
    ws.cell(row=ro, column=1).alignment = AL
    td = ro; ro += 1

    # Contingência
    ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=5)
    ws.cell(row=ro, column=1, value='CONTINGÊNCIA (%)')
    ws.cell(row=ro, column=6, value=0.10)
    ws.cell(row=ro, column=6).font = F_BLUE
    ws.cell(row=ro, column=6).number_format = '0.00%'
    ws.cell(row=ro, column=6).fill = P_YEL
    ws.cell(row=ro, column=7, value=f'=G{td}*F{ro}')
    ws.cell(row=ro, column=7).number_format = '#,##0.00'
    _style_row(ws, ro, F_BOLD, None, None, 9)
    ws.cell(row=ro, column=1).alignment = AL
    ws.cell(row=ro, column=6).alignment = AR
    ws.cell(row=ro, column=7).alignment = AR
    ws.cell(row=ro, column=8, value='Reserva técnica para imprevistos (ajustável 5-15%)').font = F_SM
    cont = ro; ro += 1

    # BDI — ponto de partida de 27,5%, EDITÁVEL pelo orçamentista.
    # 🪤 28/07/2026: não carimbar "Ref. TCU" neste valor. O Acórdão 2622/2013
    # traz faixa de 20,34% a 25% para CONSTRUÇÃO DE EDIFÍCIOS (1º ao 3º quartil);
    # 27,5% está acima do teto, então atribuir o número ao TCU era incorreto.
    # A FÓRMULA abaixo é a do acórdão; o percentual é que não vem dele.
    ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=5)
    ws.cell(row=ro, column=1, value='BDI (%) — ponto de partida, ajuste pro seu caso')
    ws.cell(row=ro, column=6, value=0.275)
    ws.cell(row=ro, column=6).font = F_BLUE
    ws.cell(row=ro, column=6).number_format = '0.00%'
    ws.cell(row=ro, column=6).fill = P_YEL
    ws.cell(row=ro, column=7, value=f'=(G{td}+G{cont})*F{ro}')
    ws.cell(row=ro, column=7).number_format = '#,##0.00'
    _style_row(ws, ro, F_BOLD, None, None, 9)
    ws.cell(row=ro, column=1).alignment = AL
    ws.cell(row=ro, column=6).alignment = AR
    ws.cell(row=ro, column=7).alignment = AR
    ws.cell(row=ro, column=8, value='AC 4% + CF 1,5% + S 0,8% + R 0,5% + G 0,5% + L 6% + T 11%').font = F_SM
    bdi = ro; ro += 1

    # Total com BDI
    ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=6)
    ws.cell(row=ro, column=1, value='TOTAL GERAL COM CONTINGÊNCIA + BDI')
    ws.cell(row=ro, column=7, value=f'=G{td}+G{cont}+G{bdi}')
    ws.cell(row=ro, column=7).number_format = '#,##0.00'
    _style_row(ws, ro, Font(name='Calibri', bold=True, size=12, color='FFFFFF'), P_SEC, AR, 9)
    ws.cell(row=ro, column=1).alignment = AL

    # ================================================================
    # SEÇÃO: OMISSOS (itens não incluídos que podem ser necessários)
    # ================================================================
    ro += 2
    ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=9)
    ws.cell(row=ro, column=1, value='OMISSOS — Itens não incluídos que provavelmente serão necessários')
    _style_row(ws, ro, Font(name='Calibri', bold=True, size=10, color='FFFFFF'), PatternFill('solid', fgColor='B45309'), AL, 9)
    ro += 1
    # OMISSOS por tipologia — itens que COSTUMAM ser esquecidos em cada tipo de obra.
    # Base comum + específicos.
    _omissos_base = [
        'Projeto executivo de instalações (elétrica, hidráulica, AC) — se não contratado separadamente',
        'Reforço estrutural — se necessário para novas cargas (marcenaria pesada, bancadas de pedra)',
        'Impermeabilização — se houver alteração em áreas úmidas (cozinha, banheiros, lavanderia)',
    ]
    if typology == "residential":
        omissos = _omissos_base + [
            'Adequação de prumadas do condomínio (elétrica, hidráulica, gás) — se o projeto exigir',
            'Aprovação em prefeitura / taxa de habite-se, se ampliação',
            'Automação residencial (persianas, iluminação cênica, áudio) — se desejado',
            'Paisagismo / jardins / área verde — se prevê plantas',
            'Tratamento acústico adicional (pisos, portas) — apartamentos de condomínio exigente',
        ]
    elif typology == "office":
        omissos = _omissos_base + [
            'Aprovação no Corpo de Bombeiros (PPCI) — taxas e honorários do projetista',
            'Compatibilização de projetos (elétrica × forro × sprinkler × AC)',
            'Adequação de infraestrutura do condomínio (elétrica, hidráulica, incêndio)',
            'Paisagismo interno — se o projeto prever jardineiras ou verde',
            'Automação e integração de sistemas (BMS, controle de iluminação)',
        ]
    elif typology == "hospital":
        omissos = _omissos_base + [
            'Projeto de gases medicinais e aprovação Anvisa',
            'Licença sanitária municipal',
            'Piso hospitalar condutivo em áreas específicas',
        ]
    elif typology == "retail":
        omissos = _omissos_base + [
            'Comunicação visual / fachada / letreiro',
            'Sistema de som e sonorização ambiente',
            'Adequação a exigências do shopping / locador',
        ]
    else:
        omissos = _omissos_base + [
            'Aprovação em órgãos (Bombeiros, Vigilância Sanitária, Prefeitura) — conforme tipologia',
            'Compatibilização de projetos entre disciplinas',
        ]
    for om in omissos:
        ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=9)
        ws.cell(row=ro, column=1, value=f'  • {om}').font = Font(name='Calibri', size=8, color='92400E')
        ws.cell(row=ro, column=1).fill = PatternFill('solid', fgColor='FEF3C7')
        ro += 1

    # ================================================================
    # SEÇÃO: EXCLUSOS (itens explicitamente fora do escopo)
    # ================================================================
    ro += 1
    ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=9)
    ws.cell(row=ro, column=1, value='EXCLUSOS — Itens explicitamente fora deste escopo (padrão de mercado)')
    _style_row(ws, ro, Font(name='Calibri', bold=True, size=10, color='FFFFFF'), PatternFill('solid', fgColor='6B7280'), AL, 9)
    ro += 1
    # EXCLUSOS por tipologia — itens que PADRÃO DE MERCADO não entram neste tipo de obra.
    # Contaminação residencial era colocar CFTV e "móveis de escritório" pra casa.
    _exclusos_base = [
        'Contas de água, luz e telefone durante a obra — cargo do condomínio/contratante',
    ]
    if typology == "residential":
        exclusos = _exclusos_base + [
            'Eletrodomésticos (geladeira, fogão, cooktop, coifa, microondas, lava-louças) — fornecimento do cliente',
            'Louças sanitárias e metais (vasos, cubas, torneiras, chuveiros) — costumam ser fornecidos separadamente',
            'Cortinas, persianas e tapeçaria — fornecimento do cliente/decorador',
            'Mobiliário solto e decoração — fornecimento do cliente',
            'Sistema de som, automação e TV — projetos especializados separados',
            'Paisagismo e vasos — projeto de paisagismo separado',
        ]
    elif typology == "office":
        exclusos = _exclusos_base + [
            'Divisórias industriais piso-teto (vidro liso, polarizado) — cargo do contratante',
            'Carpete — fornecimento pelo cliente; instalação pode estar inclusa',
            'Marcenaria sob medida (bancadas, armários, painéis) — cargo do contratante',
            'Mobiliário decorativo e de escritório — cargo do contratante',
            'Persianas e cortinas — cargo do contratante',
            'Equipamentos de TI (switches, servidores, APs, nobreaks) — cargo do contratante',
            'Sistema de CFTV e controle de acesso — quando fornecido por empresa especializada',
        ]
    elif typology == "hospital":
        exclusos = _exclusos_base + [
            'Equipamentos médicos (raio-x, ultrassom, macas, etc) — fornecimento separado',
            'Mobiliário clínico (armários de medicamentos, carrinhos) — fornecimento separado',
            'Sistemas de TI médico (PACS, prontuário) — cargo do contratante',
        ]
    elif typology == "retail":
        exclusos = _exclusos_base + [
            'Mercadoria / estoque inicial — cargo do contratante',
            'Expositores e vitrines específicas da marca — design ou fornecimento separado',
            'Sistema PDV e TI da loja — cargo do contratante',
        ]
    else:
        exclusos = _exclusos_base + [
            'Mobiliário solto e decoração — cargo do contratante',
            'Equipamentos específicos da operação — cargo do contratante',
        ]
    for ex in exclusos:
        ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=9)
        ws.cell(row=ro, column=1, value=f'  • {ex}').font = Font(name='Calibri', size=8, color='374151')
        ws.cell(row=ro, column=1).fill = PatternFill('solid', fgColor='F3F4F6')
        ro += 1

    # Notas profissionais
    ro += 2
    ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=9)
    ws.cell(row=ro, column=1, value='NOTAS:').font = F_BOLD; ro += 1
    notas = [
        '1. REFORMA: quantitativos consideram apenas o que MUDA. Conferir in loco e em projeto executivo.',
        '2. Colunas MAT e M.O. (amarelo): preencher pelo orçamentista/fornecedor.',
        '3. BDI: 27,5% é só um ponto de partida — ajuste pro seu caso. Fórmula do Acórdão TCU '
        '2622/2013: ((1+AC)(1+CF)(1+S)(1+R)(1+G)(1+L)/(1-T))-1. O mesmo acórdão traz faixa de '
        '20,34% a 25% para construção de edifícios em obra pública.',
        '4. Itens em BRANCO: quantidade medida/contada diretamente do arquivo (bloco, hachura, linha). Confiável pra aprovar direto.',
        '5. Itens em LARANJA: quantidade sugerida pela IA sem medição direta — SEMPRE confirmar antes de orçar.',
        '6. Itens em CINZA (Premissas): metadados do projeto extraídos do arquivo — revisar no original.',
        '7. Itens em ROXO (Sugestões): checklist de custos indiretos típicos — preencher quantidade conforme o projeto.',
        '8. Contingência 10% — reserva técnica para imprevistos. Ajustar conforme risco do projeto.',
        '9. Perdas de material (5-10% típico) NÃO aplicadas automaticamente — adicionar ao preencher a coluna de custo se pertinente.',
        '10. OMISSOS: itens que podem ser necessários mas não foram incluídos — avaliar com equipe de projeto.',
        '11. EXCLUSOS: itens padrão de mercado excluídos do escopo de empreiteiras.',
        '12. Planilha gerada por AI.arq (ai.arq.br) — validar com engenheiro de custos.',
    ]
    for n in notas:
        ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=9)
        ws.cell(row=ro, column=1, value=n).font = F_SM; ro += 1

    # Configurações
    ws.freeze_panes = 'A6'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1

    # ================================================================
    # SHEET 3: REFERÊNCIAS SINAPI (códigos oficiais por item)
    # ================================================================
    # TCPO removido (16/07): catálogo é SINAPI, TCPO fora da fase atual.
    items_with_sinapi = [it for it in items if getattr(it, 'sinapi_matches', None)]
    if items_with_sinapi:
        wsm = wb.create_sheet('Referências SINAPI')
        wsm.sheet_properties.tabColor = '059669'  # verde

        widths_m = [7, 45, 5, 10, 12, 12, 50, 6, 8]
        for i, w in enumerate(widths_m, 1):
            wsm.column_dimensions[get_column_letter(i)].width = w

        # Cabeçalho geral
        wsm.merge_cells('A1:I1')
        wsm.cell(row=1, column=1,
                 value='REFERÊNCIAS SINAPI — códigos oficiais por item').font = F_TITLE
        wsm.merge_cells('A2:I2')
        wsm.cell(row=2, column=1, value=(
            'SINAPI (Caixa) = referência oficial de preço/quantitativo no Brasil, '
            'atualizada mensalmente. Cada item do quantitativo recebe os códigos '
            'SINAPI mais próximos pra você buscar o preço/composição.')).font = F_NOTE
        wsm.merge_cells('A3:I3')
        wsm.cell(row=3, column=1, value=(
            'AI.arq NÃO entrega preço. Use o código pra consultar o preço atualizado '
            'no SINAPI oficial (https://www.caixa.gov.br). '
            'A busca por texto levanta os candidatos e a IA escolhe qual é o mesmo '
            'serviço: "✓ IA conferiu" é o escolhido (vai pra coluna REF do orçamento); '
            'os demais são alternativas com o % de semelhança do TEXTO, que erra — '
            'confira sempre a composição oficial ao lado antes de usar.')
        ).font = F_NOTE

        rm = 5
        hdrs_m = ['ITEM', 'DESCRIÇÃO DO QUANTITATIVO', 'UN', 'QTDE',
                  'BASE', 'CÓDIGO',
                  'COMPOSIÇÃO OFICIAL (referência)',
                  'UN.', 'MATCH %']
        for c, h in enumerate(hdrs_m, 1):
            cl = wsm.cell(row=rm, column=c, value=h)
            cl.font = F_HDR; cl.fill = P_HDR; cl.alignment = AC; cl.border = BD
        rm += 1

        P_SINAPI = PatternFill('solid', fgColor='DBEAFE')   # azul claro pra SINAPI (gov)
        P_NOMATCH = PatternFill('solid', fgColor='FEE2E2')  # vermelho claro pra sem match

        for item in items:
            sinapi_matches = getattr(item, 'sinapi_matches', []) or []
            # A IA olhou os candidatos e disse que nenhum é o mesmo serviço: não
            # adianta listá-los aqui "pra conferir". Era assim que "Pendente
            # decorativo" oferecia SPRINKLER TIPO PENDENTE — o texto casa, a coisa
            # não. Oferecer o candidato reprovado é convidar a copiar.
            _reprovados = bool(sinapi_matches) and all(m.get('_llm_rejected') for m in sinapi_matches)
            if _reprovados:
                sinapi_matches = []

            if not sinapi_matches:
                # Item sem match nenhum — linha informativa
                wsm.cell(row=rm, column=1, value=item.item_num).font = F_N
                wsm.cell(row=rm, column=2, value=item.description).font = F_N
                wsm.cell(row=rm, column=3, value=item.unit).font = F_N
                wsm.cell(row=rm, column=4, value=item.quantity).font = F_N
                wsm.merge_cells(start_row=rm, start_column=5, end_row=rm, end_column=9)
                wsm.cell(row=rm, column=5,
                         value=('Sem correspondência no SINAPI — a IA conferiu os candidatos e '
                                'nenhum é o mesmo serviço. Buscar manualmente em '
                                'https://www.caixa.gov.br/sinapi')
                         if _reprovados else
                         ('Sem match SINAPI (descrição muito específica) — '
                          'buscar manualmente em https://www.caixa.gov.br/sinapi')).font = F_NOTE
                for c in range(1, 10):
                    wsm.cell(row=rm, column=c).border = BD
                    wsm.cell(row=rm, column=c).fill = P_NOMATCH
                    wsm.cell(row=rm, column=c).alignment = AC if c in (1, 3, 4) else AL
                rm += 1
                continue

            # Linha principal: descrição do item (uma vez por item)
            wsm.cell(row=rm, column=1, value=item.item_num).font = F_BOLD
            wsm.cell(row=rm, column=2, value=item.description).font = F_BOLD
            wsm.cell(row=rm, column=3, value=item.unit).font = F_N
            wsm.cell(row=rm, column=4, value=item.quantity).font = F_N
            wsm.merge_cells(start_row=rm, start_column=5, end_row=rm, end_column=9)
            wsm.cell(row=rm, column=5, value='↓ matches encontrados ↓').font = F_NOTE
            for c in range(1, 10):
                wsm.cell(row=rm, column=c).border = BD
                wsm.cell(row=rm, column=c).alignment = AC if c in (1, 3, 4, 5) else AL
            rm += 1

            # SINAPI primeiro (gov, atualizado mensal) — até 2 matches
            for idx, sm in enumerate(sinapi_matches[:2]):
                level = sm.get('_match_level', 'full')
                mark = '~' if level.startswith('simplified') else ''
                wsm.cell(row=rm, column=1, value='').font = F_SM
                wsm.cell(row=rm, column=2,
                         value='' if idx == 0 else '').font = F_SM
                wsm.cell(row=rm, column=5, value='SINAPI').font = F_BOLD
                wsm.cell(row=rm, column=6,
                         value=f'{mark}{sm.get("codigo", "")}').font = F_BOLD
                wsm.cell(row=rm, column=7,
                         value=sm.get('descricao', '')[:95]).font = F_N
                wsm.cell(row=rm, column=8, value=sm.get('unidade', '')).font = F_N
                # Quem foi escolhido pela IA não mostra %: a nota é de TEXTO e não
                # foi o critério da escolha (no "spot", a cuba pontua mais que a
                # luminária certa). Cravar um número aqui seria a mesma precisão
                # fingida do "100%" que este trabalho veio consertar.
                if sm.get('_llm_picked'):
                    sim_pct = '✓ IA conferiu'
                else:
                    sim_pct = f"{min(100, max(0, int((sm.get('similarity', 0) or 0) * 100)))}% (texto)"
                wsm.cell(row=rm, column=9, value=sim_pct).font = F_BOLD
                for c in range(1, 10):
                    wsm.cell(row=rm, column=c).border = BD
                    wsm.cell(row=rm, column=c).fill = P_SINAPI
                    wsm.cell(row=rm, column=c).alignment = AC if c in (1, 3, 4, 5, 6, 8, 9) else AL
                rm += 1

            rm += 1  # linha em branco entre itens

        # Rodapé
        wsm.merge_cells(start_row=rm + 1, start_column=1, end_row=rm + 1, end_column=9)
        wsm.cell(row=rm + 1, column=1, value=(
            'Como ler: MATCH % indica a similaridade entre a descrição do item '
            'e a composição SINAPI. Acima de 70% = referência forte. 40-70% = '
            'revisar. Abaixo de 30% = candidato fraco, buscar código manualmente '
            'em https://www.caixa.gov.br/sinapi.')).font = F_NOTE

        wsm.freeze_panes = 'A6'
        wsm.page_setup.orientation = 'landscape'
        wsm.page_setup.fitToWidth = 1

    wb.save(output_path)
    return output_path
