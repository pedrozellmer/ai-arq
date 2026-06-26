# -*- coding: utf-8 -*-
"""Auto-calibration system for AI.arq budget platform.

Compares AI-generated spreadsheets with user-revised versions,
stores correction factors, and applies them to future outputs.
"""
import os
import json
import time
import unicodedata
import re
import difflib
import urllib.request
import urllib.parse

from openpyxl import load_workbook

# ── Supabase config (same as main.py) ──
# `apikey` continua sendo a anon (PostgREST exige). O `Authorization: Bearer`
# usa a service_role pra escrever em density_benchmarks/catalog_* depois que a
# RLS fechar pra anon. service_role roda só server-side, nunca exposta ao frontend.
SUPABASE_URL = "https://kqjabzwgbfuivzlcfvvu.supabase.co"
SUPABASE_KEY = os.getenv("SUPABASE_KEY") or os.getenv("SUPABASE_ANON_KEY") or "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImtxamFiendnYmZ1aXZ6bGNmdnZ1Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzYwMDg5NzcsImV4cCI6MjA5MTU4NDk3N30.48xSenZlDV0LfD94ZxwGvX41Kf9Je2n-ouZpJrrCSKI"
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or SUPABASE_KEY

# ── In-memory cache for correction factors ──
_factors_cache = {"data": None, "timestamp": 0}
_CACHE_TTL = 300  # 5 minutes


# ═══════════════════════════════════════════════
#  HELPER: normalize_item_type
# ═══════════════════════════════════════════════

def normalize_item_type(description: str) -> str:
    """Normalize a description into a stable key for calibration lookup.

    Converts the description to lowercase, strips accents, replaces
    non-alphanumeric characters with underscores, and truncates to 50 chars.
    Used as a stable lookup key regardless of how the description is written.
    """
    if not description:
        return ""
    # Lowercase
    text = description.lower().strip()
    # Remove accents
    text = unicodedata.normalize("NFD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    # Replace non-alphanumeric with underscore
    text = re.sub(r"[^a-z0-9]+", "_", text)
    # Collapse multiple underscores and strip edges
    text = re.sub(r"_+", "_", text).strip("_")
    # Limit to 50 chars
    return text[:50]


# ═══════════════════════════════════════════════
#  1. compare_spreadsheets
# ═══════════════════════════════════════════════

def _parse_float(val) -> float:
    """Try to extract a float from a cell value."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", ".").strip()
    # Remove units like m², un etc.
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _extract_items_from_sheet(ws) -> list[dict]:
    """Extract budget items from an 'Orcamento' worksheet.

    Returns list of dicts with keys: item_num, description, unit, quantity, discipline.
    The spreadsheet layout has columns:
      A=item_num, B=description, C=unit, D=quantity, ...
    Section headers (discipline) are merged rows like "1. DEMOLICAO E REMOCAO".
    """
    items = []
    current_discipline = ""

    for row in ws.iter_rows(min_row=1, max_col=9, values_only=False):
        # Check for section header (merged cell spanning columns)
        cell_a = row[0]
        cell_b = row[1]
        val_a = cell_a.value
        val_b = cell_b.value

        # Section headers: value in column A like "1. SERVICOS PRELIMINARES"
        # They have the section fill and are merged across all columns
        if val_a and isinstance(val_a, str):
            stripped = val_a.strip()
            # Match section header pattern: "N. DISCIPLINE NAME"
            m = re.match(r"^\d+\.\s+(.+)$", stripped)
            if m and val_b is None:
                # This is a section header (discipline)
                current_discipline = m.group(1).strip()
                # Normalize: "DEMOLICAO E REMOCAO" -> titlecase
                current_discipline = current_discipline.title()
                continue

        # Item rows have item_num like "1.1", "2.3" in column A
        if val_a and isinstance(val_a, str) and re.match(r"^\d+\.\d+", str(val_a).strip()):
            item_num = str(val_a).strip()
            description = str(val_b).strip() if val_b else ""
            unit = str(row[2].value).strip() if row[2].value else ""
            quantity = _parse_float(row[3].value)

            if description and len(description) >= 3:
                items.append({
                    "item_num": item_num,
                    "description": description,
                    "unit": unit,
                    "quantity": quantity,
                    "discipline": current_discipline,
                })
        # Also handle numeric item_num (openpyxl may return float)
        elif val_a and isinstance(val_a, (int, float)):
            num_str = str(val_a)
            if "." in num_str and not num_str.startswith("0."):
                item_num = num_str
                description = str(val_b).strip() if val_b else ""
                unit = str(row[2].value).strip() if row[2].value else ""
                quantity = _parse_float(row[3].value)

                if description and len(description) >= 3:
                    items.append({
                        "item_num": item_num,
                        "description": description,
                        "unit": unit,
                        "quantity": quantity,
                        "discipline": current_discipline,
                    })

    return items


def compare_spreadsheets(original_path: str, revised_path: str) -> list[dict]:
    """Compare an AI-generated XLSX with a user-revised version.

    Opens both files, finds the 'Orcamento' sheet, matches items by
    item_num (exact) then by fuzzy description match, and returns
    a list of calibration entries for items whose quantities differ.
    """
    wb_orig = load_workbook(original_path, read_only=True, data_only=True)
    wb_rev = load_workbook(revised_path, read_only=True, data_only=True)

    # Find the Orcamento sheet
    def find_sheet(wb):
        for name in wb.sheetnames:
            if "orcamento" in name.lower() or "orçamento" in name.lower():
                return wb[name]
        # Fallback: second sheet (first is usually Resumo)
        if len(wb.sheetnames) >= 2:
            return wb[wb.sheetnames[1]]
        return wb[wb.sheetnames[0]]

    ws_orig = find_sheet(wb_orig)
    ws_rev = find_sheet(wb_rev)

    orig_items = _extract_items_from_sheet(ws_orig)
    rev_items = _extract_items_from_sheet(ws_rev)

    wb_orig.close()
    wb_rev.close()

    if not orig_items or not rev_items:
        return []

    # Build lookup by item_num for revised items
    rev_by_num = {}
    for item in rev_items:
        rev_by_num[item["item_num"]] = item

    # Build lookup by normalized description for fuzzy matching
    rev_by_desc = {}
    for item in rev_items:
        key = normalize_item_type(item["description"])
        if key:
            rev_by_desc[key] = item

    comparisons = []
    matched_rev_nums = set()

    for orig in orig_items:
        matched = None

        # Strategy 1: exact item_num match
        if orig["item_num"] in rev_by_num:
            matched = rev_by_num[orig["item_num"]]
            matched_rev_nums.add(matched["item_num"])
        else:
            # Strategy 2: fuzzy description match
            orig_key = normalize_item_type(orig["description"])
            if orig_key and orig_key in rev_by_desc:
                matched = rev_by_desc[orig_key]
                matched_rev_nums.add(matched["item_num"])
            else:
                # Strategy 3: SequenceMatcher on description
                best_ratio = 0
                best_match = None
                orig_desc_lower = orig["description"].lower()
                for rev in rev_items:
                    if rev["item_num"] in matched_rev_nums:
                        continue
                    ratio = difflib.SequenceMatcher(
                        None, orig_desc_lower, rev["description"].lower()
                    ).ratio()
                    if ratio > best_ratio:
                        best_ratio = ratio
                        best_match = rev
                if best_ratio > 0.7 and best_match:
                    matched = best_match
                    matched_rev_nums.add(matched["item_num"])

        if not matched:
            continue

        ai_qty = orig["quantity"]
        real_qty = matched["quantity"]

        # Skip if quantities are the same or zero
        if ai_qty == real_qty:
            continue
        if ai_qty <= 0 or real_qty <= 0:
            continue

        deviation_pct = (ai_qty - real_qty) / real_qty * 100
        correction_factor = real_qty / ai_qty

        item_type = normalize_item_type(orig["description"])
        discipline = orig.get("discipline", "") or matched.get("discipline", "")

        comparisons.append({
            "item_type": item_type,
            "discipline": discipline,
            "unit": orig.get("unit", "") or matched.get("unit", ""),
            "ai_quantity": round(ai_qty, 4),
            "real_quantity": round(real_qty, 4),
            "deviation_pct": round(deviation_pct, 2),
            "correction_factor": round(correction_factor, 4),
        })

    return comparisons


# ═══════════════════════════════════════════════
#  2. save_calibration_data
# ═══════════════════════════════════════════════

def _supabase_insert(table: str, data: dict) -> bool:
    """Insert a record into Supabase via REST API (urllib, no deps)."""
    try:
        url = f"{SUPABASE_URL}/rest/v1/{table}"
        body = json.dumps(data).encode("utf-8")
        req = urllib.request.Request(url, data=body, method="POST")
        req.add_header("apikey", SUPABASE_KEY)
        req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
        req.add_header("Content-Type", "application/json")
        req.add_header("Prefer", "return=minimal")
        urllib.request.urlopen(req, timeout=5)
        return True
    except Exception as e:
        print(f"[calibrator] Supabase insert error: {e}")
        return False


def save_calibration_data(
    comparisons: list[dict],
    source: str = "user",
    project_id: str = "",
    user_id: str = "",
) -> int:
    """Insert each comparison into the Supabase `calibration` table.

    Returns the number of successfully inserted rows.
    """
    inserted = 0
    for comp in comparisons:
        record = {
            "item_type": comp["item_type"],
            "discipline": comp.get("discipline", ""),
            "unit": comp.get("unit", ""),
            "ai_quantity": comp["ai_quantity"],
            "real_quantity": comp["real_quantity"],
            "deviation_pct": comp["deviation_pct"],
            "correction_factor": comp["correction_factor"],
            "source": source,
            "project_id": project_id,
            "user_id": user_id,
        }
        if _supabase_insert("calibration", record):
            inserted += 1
    # Invalidate cache after new data
    _factors_cache["data"] = None
    _factors_cache["timestamp"] = 0
    return inserted


# ═══════════════════════════════════════════════
#  3. get_correction_factors
# ═══════════════════════════════════════════════

def get_correction_factors() -> dict:
    """Fetch the calibration_factors view from Supabase.

    Returns dict: {item_type: {factor, data_points, deviation}, ...}
    Caches result for 5 minutes.
    """
    now = time.time()
    if _factors_cache["data"] is not None and (now - _factors_cache["timestamp"]) < _CACHE_TTL:
        return _factors_cache["data"]

    factors = {}
    try:
        url = f"{SUPABASE_URL}/rest/v1/calibration_factors?select=*"
        req = urllib.request.Request(url, method="GET")
        req.add_header("apikey", SUPABASE_KEY)
        req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
        req.add_header("Accept", "application/json")
        resp = urllib.request.urlopen(req, timeout=10)
        rows = json.loads(resp.read().decode("utf-8"))

        for row in rows:
            item_type = row.get("item_type", "")
            if not item_type:
                continue
            factors[item_type] = {
                "factor": row.get("avg_factor", 1.0),
                "data_points": row.get("data_points", 0),
                "deviation": row.get("avg_deviation", 0),
                "discipline": row.get("discipline", ""),
                "unit": row.get("unit", ""),
                "stddev": row.get("stddev_factor"),
                "min_factor": row.get("min_factor"),
                "max_factor": row.get("max_factor"),
            }
    except Exception as e:
        print(f"[calibrator] Error fetching factors: {e}")

    _factors_cache["data"] = factors
    _factors_cache["timestamp"] = now
    return factors


# ═══════════════════════════════════════════════
#  4. apply_corrections
# ═══════════════════════════════════════════════

def apply_corrections(items: list, factors: dict) -> list:
    """Sinaliza (NÃO altera) anomalias de quantidade com base na calibração.

    Para cada item, tenta casar o item_type normalizando a descrição. Se casar
    E data_points >= 2, grava um ALERTA na observação comparando a quantidade
    medida com a ordem de grandeza que o histórico sugere. NUNCA multiplica a
    quantidade — a quantidade medida do CAD é intocável (regra dura de
    isolamento entre projetos). O ratio do histórico só serve pra alertar.

    Returns the items list (só observações podem mudar, nunca quantity).
    """
    if not factors:
        return items

    corrections_applied = 0

    for item in items:
        item_type = normalize_item_type(item.description)
        if not item_type or item_type not in factors:
            continue

        f = factors[item_type]
        if f["data_points"] < 2:
            continue

        factor_val = f["factor"]
        # Skip if factor is essentially 1.0 (no meaningful correction)
        if 0.98 <= factor_val <= 1.02:
            continue

        # ⚠ ISOLAMENTO (regra dura): NUNCA multiplicar a quantidade medida por
        # fator de OUTROS projetos — isso contamina o projeto atual. A calibração
        # por ratio serve SÓ pra ALERTAR anomalia, não pra copiar/ajustar valor.
        # Mantemos a quantidade medida INTACTA e gravamos um alerta pra revisão.
        original_qty = item.quantity
        sugerido = round(item.quantity * factor_val, 2)

        cal_note = (
            f"⚠ Calibracao: o historico ({f['data_points']} projetos) sugere ordem "
            f"de ~{sugerido} pra este tipo (fator {factor_val:.2f}); confira se a "
            f"quantidade medida ({original_qty}) faz sentido. Quantidade NAO alterada."
        )
        if item.observations:
            item.observations = f"{item.observations} | {cal_note}"
        else:
            item.observations = cal_note

        # NÃO promover confidence: calibração é uma estimativa baseada em histórico,
        # não uma medição. Itens corrigidos por calibração permanecem como "estimado"
        # (laranja) pro usuário revisar manualmente. Regra do produto: nunca marcar
        # como "confirmado" algo que não tenha origem em medição objetiva do CAD.

        corrections_applied += 1

    if corrections_applied > 0:
        print(f"[calibrator] Flagged {corrections_applied} possible anomalies (quantities NOT changed).")

    return items
