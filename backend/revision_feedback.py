# -*- coding: utf-8 -*-
"""Feedback de revisão — mede ONDE A IA MAIS ERRA usando a planilha REVISADA
que o cliente devolve (endpoint /api/projects/{job_id}/revised-sheet/upload).

Fluxo:
  1. parse_planilha_revisada() — lê a NOSSA planilha (formato spreadsheet.py),
     possivelmente EDITADA pelo cliente: quantidades alteradas, linhas
     removidas e linhas adicionadas. Robusto a célula vazia/fórmula/mesclada.
  2. comparar() — pareia com os itens ORIGINAIS (tabela project_items) por
     descrição (exato → fuzzy conservador por tokens) e calcula, por item,
     qty_original/qty_revisada/delta_pct/acao + agregados por disciplina e o
     recorte por confidence (a pergunta de ouro: o cliente corrige mais os
     ESTIMADOS? os MEDIDOS sobrevivem?).
  3. salvar_feedback() — INSERT via REST na tabela `revision_feedback`
     (service_role; RLS ligada sem policy = só o backend escreve/lê).
  4. resumo_para_admin() — agrega as rows da tabela pro painel
     "Onde a IA mais erra" no admin.

REGRA DE CASA (isolamento entre projetos): isto aqui NUNCA vira correção
automática de quantidade em outros projetos. É medição de erro, ponto.
"""
import os
import re
import json
import difflib
import statistics
import unicodedata
import urllib.parse
import urllib.request

from openpyxl import load_workbook

# ── Supabase config (mesmo padrão do calibrator.py) ──
# `apikey` continua sendo a anon (PostgREST exige). O `Authorization: Bearer`
# usa a service_role — roda só server-side, nunca exposta ao frontend.
SUPABASE_URL = "https://kqjabzwgbfuivzlcfvvu.supabase.co"
SUPABASE_KEY = os.getenv("SUPABASE_KEY") or os.getenv("SUPABASE_ANON_KEY") or "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImtxamFiendnYmZ1aXZ6bGNmdnZ1Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzYwMDg5NzcsImV4cCI6MjA5MTU4NDk3N30.48xSenZlDV0LfD94ZxwGvX41Kf9Je2n-ouZpJrrCSKI"
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or SUPABASE_KEY

# Quantos itens detalhados guardar por planilha na tabela (jsonb não é lugar
# de planilha inteira de shopping center; 400 cobre 99% dos projetos).
MAX_ITENS_DETALHE = 400

# Tolerância relativa pra considerar quantidade "igual" (arredondamento do
# Excel: 88.5 salvo como 88.4999… não é correção do cliente).
_TOL_REL = 0.005

# Score mínimo do pareamento fuzzy. Conservador de propósito: pareamento
# errado gera delta fantasma (pior que contar removido+adicionado).
FUZZY_MIN_SCORE = 0.5


# ═══════════════════════════════════════════════
#  Normalização
# ═══════════════════════════════════════════════

_STOPWORDS = {
    "de", "da", "do", "das", "dos", "e", "em", "com", "para", "por",
    "a", "o", "as", "os", "no", "na", "nos", "nas", "um", "uma",
    "sob", "sobre", "ate", "the",
}

_CONECTIVOS = {"e", "de", "da", "do", "das", "dos", "em", "com", "para", "por", "a", "o"}

_RE_ITEM_NUM = re.compile(r"^\d+\.\d+$")
_RE_SECTION = re.compile(r"^(\d+)\.\s+(\S.*)$")


def _strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s))
    return s.encode("ascii", "ignore").decode("ascii")


def _norm_desc(s) -> str:
    """Descrição normalizada pra pareamento: minúscula, sem acento,
    só alfanumérico, espaços únicos."""
    if s is None:
        return ""
    text = _strip_accents(str(s)).lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()[:200]


def _tokens(desc_norm: str) -> set:
    """Tokens significativos (>=3 chars, fora stopwords) pro fuzzy."""
    return {w for w in desc_norm.split() if len(w) >= 3 and w not in _STOPWORDS}


def _parse_qty(val):
    """Float de uma célula de quantidade. None se vazia/fórmula/não-numérica.

    Com data_only=True fórmula vem como valor em cache; se o arquivo foi salvo
    sem cache (raro), vem string '=...' — tratamos como None (sem número)."""
    if val is None:
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s.startswith("="):
        return None
    s = s.replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")   # 1.234,56 → 1234.56
    elif "," in s:
        s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _unit_family(u: str) -> str:
    """Família da unidade (mesma lógica do spreadsheet._unit_family) — usada
    como guarda do fuzzy: item em m² não pareia com item em metro linear."""
    if not u:
        return ""
    u = str(u).strip().lower().replace("²", "2").replace("³", "3")
    if u in ("m2",):
        return "area"
    if u in ("m3",):
        return "volume"
    if u in ("m", "ml", "metro", "mlinear"):
        return "linear"
    if u in ("un", "und", "unid", "pç", "pc", "peca", "peça", "cj", "conj", "cjto"):
        return "cont"
    if u in ("kg", "t", "ton"):
        return "peso"
    if u in ("vb", "verba"):
        return "verba"
    if u in ("l", "lt", "litro"):
        return "liquido"
    return ""


def _canon_discipline(s) -> str:
    """'3. PISOS E RODAPÉS' → 'Pisos e Rodapés' (mesmo formato da tabela
    project_items, pra agregados casarem entre original e revisada)."""
    s = re.sub(r"^\d+\.\s*", "", str(s or "")).strip()
    if not s:
        return ""
    words = s.title().split()
    out = [words[0]] + [w.lower() if w.lower() in _CONECTIVOS else w
                        for w in words[1:]]
    return " ".join(out)


def _norm_confidence(c) -> str:
    """Fail-safe do produto: só 'confirmado' é medido; resto (estimado,
    verificar, vazio, lixo) cai em 'estimado'."""
    return "confirmado" if str(c or "").strip().lower() == "confirmado" else "estimado"


# ═══════════════════════════════════════════════
#  1. parse_planilha_revisada
# ═══════════════════════════════════════════════

def _find_orcamento_sheet(wb):
    for name in wb.sheetnames:
        if "orcamento" in _strip_accents(name).lower():
            return wb[name]
    # Fallback: 2ª aba (1ª é o Resumo Comparativo)
    if len(wb.sheetnames) >= 2:
        return wb[wb.sheetnames[1]]
    return wb[wb.sheetnames[0]] if wb.sheetnames else None


def parse_planilha_revisada(xlsx_path: str) -> list[dict]:
    """Lê a NOSSA planilha (aba 'Orçamento'), possivelmente editada.

    Layout esperado (spreadsheet.py): A=item_num ('1.1'), B=descrição, C=un,
    D=qtde, H=observações (selo '✓ MEDIDO do CAD' / '⚠ ESTIMADO — revisar').
    Cabeçalhos de seção são linhas mescladas 'N. DISCIPLINA' (só coluna A).

    Robustez:
      - célula vazia → quantity None (nossa planilha grava qty=0 como vazia);
      - fórmula sem cache ('=...') → None;
      - mesclada → openpyxl devolve valor só na âncora, o resto None (ok);
      - PREMISSAS (seção 0) e SUGESTÕES ('S.x' em diante) ficam FORA —
        são metadados/checklist, não itens medidos;
      - linha ADICIONADA pelo cliente sem numeração (col A vazia, descrição
        na B, dentro de uma disciplina) entra com item_num ''.

    Retorna [] em qualquer falha (arquivo corrompido, aba ausente…) — este
    módulo é best-effort, nunca pode derrubar o upload.
    """
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"[revision_feedback] não abriu '{xlsx_path}': {type(e).__name__}: {e}")
        return []

    try:
        ws = _find_orcamento_sheet(wb)
        if ws is None:
            return []

        itens = []
        disciplina = ""
        for row in ws.iter_rows(min_row=1, max_col=9):
            val_a = row[0].value
            val_b = row[1].value
            a_str = str(val_a).strip() if val_a is not None else ""

            # Fim da área de itens: SUGESTÕES / RESUMO GERAL / OMISSOS.
            a_flat = _strip_accents(a_str).upper()
            if a_flat.startswith(("SUGESTOES", "RESUMO GERAL", "OMISSOS")):
                break

            # Cabeçalho de seção "N. DISCIPLINA" (mesclado → B vazio).
            m = _RE_SECTION.match(a_str)
            if m and (val_b is None or str(val_b).strip() == ""):
                if m.group(1) == "0":
                    disciplina = ""  # 0. PREMISSAS = metadados, pula os 0.x
                else:
                    disciplina = _canon_discipline(m.group(2))
                continue

            # Linha de item numerada ("1.1"; float 1.1 do openpyxl vira "1.1").
            if _RE_ITEM_NUM.match(a_str) and not a_str.startswith("0."):
                desc = str(val_b).strip() if val_b is not None else ""
                if len(desc) < 3:
                    continue
                itens.append(_montar_item(a_str, desc, row, disciplina))
                continue

            # Linha ADICIONADA pelo cliente: sem numeração, com descrição,
            # dentro de uma disciplina (evita pegar título/cabeçalho/nota).
            if not a_str and disciplina and val_b is not None:
                desc = str(val_b).strip()
                if len(desc) >= 3:
                    itens.append(_montar_item("", desc, row, disciplina))

        return itens
    except Exception as e:
        print(f"[revision_feedback] erro lendo '{xlsx_path}': {type(e).__name__}: {e}")
        return []
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _montar_item(item_num: str, desc: str, row, disciplina: str) -> dict:
    unit = str(row[2].value).strip() if row[2].value is not None else ""
    quantity = _parse_qty(row[3].value)
    obs = str(row[7].value).strip() if row[7].value is not None else ""
    obs_flat = _strip_accents(obs).upper()
    if "MEDIDO" in obs_flat:
        selo_medido = True
    elif "ESTIMADO" in obs_flat:
        selo_medido = False
    else:
        selo_medido = None  # cliente pode ter apagado a observação
    return {
        "item_num": item_num,
        "description": desc,
        "unit": unit,
        "quantity": quantity,
        "discipline": disciplina,
        "selo_medido": selo_medido,
        "observations": obs[:300],
    }


# ═══════════════════════════════════════════════
#  2. comparar
# ═══════════════════════════════════════════════

def _prep_orig(o):
    desc = str(o.get("description") or "").strip()
    if len(desc) < 3:
        return None
    qty = _parse_qty(o.get("quantity"))
    return {
        "descricao": desc,
        "desc_norm": _norm_desc(desc),
        "tokens": _tokens(_norm_desc(desc)),
        "unit": str(o.get("unit") or "").strip(),
        "qty": qty if qty is not None else 0.0,
        "confidence": _norm_confidence(o.get("confidence")),
        "disciplina": str(o.get("discipline") or "").strip() or "Complementares",
    }


def _prep_rev(r):
    desc = str(r.get("description") or "").strip()
    if len(desc) < 3:
        return None
    qty = _parse_qty(r.get("quantity"))
    return {
        "descricao": desc,
        "desc_norm": _norm_desc(desc),
        "tokens": _tokens(_norm_desc(desc)),
        "unit": str(r.get("unit") or "").strip(),
        # None = célula vazia; nossa planilha grava qty=0 como vazia, então
        # vazio ≡ 0. Cliente que LIMPOU o número aparece como delta pra 0.
        "qty": qty if qty is not None else 0.0,
        "disciplina": str(r.get("discipline") or "").strip() or "Complementares",
    }


def _match_score(o: dict, r: dict) -> float:
    """Score fuzzy CONSERVADOR entre item original e revisado.

    - unidade de família diferente (m² × m) → 0 (nunca pareia);
    - caminho principal: Jaccard de tokens >= 0.5 com >=2 tokens em comum
      (>=1 quando a descrição só tem 1-2 tokens);
    - fallback: SequenceMatcher >= 0.85 (renome de leve que troca tokens,
      ex. singular/plural) — ainda conservador.
    """
    fo, fr = _unit_family(o["unit"]), _unit_family(r["unit"])
    if fo and fr and fo != fr:
        return 0.0
    ta, tb = o["tokens"], r["tokens"]
    if ta and tb:
        inter = ta & tb
        union = ta | tb
        jac = len(inter) / len(union)
        min_comum = 1 if min(len(ta), len(tb)) <= 2 else 2
        if len(inter) >= min_comum and jac >= FUZZY_MIN_SCORE:
            return jac
    ratio = difflib.SequenceMatcher(None, o["desc_norm"], r["desc_norm"]).ratio()
    if ratio >= 0.85:
        return ratio * 0.99  # sinal secundário: nunca ganha de Jaccard perfeito
    return 0.0


def _bucket(regs: list, n_adicionados: int = 0) -> dict:
    """Agregado padrão sobre registros do lado ORIGINAL (mantido/alterado/
    removido). Adicionados entram só como contagem (não têm original)."""
    n_itens = len(regs)
    n_alt = sum(1 for x in regs if x["acao"] == "alterado")
    n_rem = sum(1 for x in regs if x["acao"] == "removido")
    n_man = sum(1 for x in regs if x["acao"] == "mantido")
    deltas = [abs(x["delta_pct"]) for x in regs
              if x["acao"] == "alterado" and isinstance(x["delta_pct"], (int, float))]
    return {
        "n_itens": n_itens,
        "n_mantidos": n_man,
        "n_alterados": n_alt,
        "n_removidos": n_rem,
        "n_adicionados": n_adicionados,
        "pct_alterados": round(100.0 * n_alt / n_itens, 1) if n_itens else None,
        "mediana_abs_delta_pct": round(statistics.median(deltas), 1) if deltas else None,
    }


def comparar(itens_originais: list[dict], itens_revisados: list[dict]) -> dict:
    """Pareia original × revisada e devolve o mapa do erro.

    Pareamento: (1) descrição normalizada EXATA; (2) fuzzy conservador por
    tokens (Jaccard >= 0.5, unidade compatível, cada revisado usado 1×,
    melhores scores primeiro). item_num NÃO é chave: linhas removidas/
    adicionadas renumeram tudo.

    Ações por item: mantido / alterado / removido / adicionado.
    delta_pct = (revisada - original) / original * 100 (None se original = 0).
    """
    orig = [p for p in (_prep_orig(o) for o in (itens_originais or [])) if p]
    rev = [p for p in (_prep_rev(r) for r in (itens_revisados or [])) if p]

    pares = {}      # idx original -> idx revisado
    usados = set()  # idx revisados já pareados

    # 1) exato por descrição normalizada
    rev_por_desc = {}
    for j, r in enumerate(rev):
        rev_por_desc.setdefault(r["desc_norm"], []).append(j)
    for i, o in enumerate(orig):
        for j in rev_por_desc.get(o["desc_norm"], []):
            if j not in usados:
                pares[i] = j
                usados.add(j)
                break

    # 2) fuzzy conservador — melhores scores primeiro, sem reuso
    candidatos = []
    for i, o in enumerate(orig):
        if i in pares:
            continue
        for j, r in enumerate(rev):
            if j in usados:
                continue
            s = _match_score(o, r)
            if s >= FUZZY_MIN_SCORE:
                candidatos.append((s, i, j))
    for s, i, j in sorted(candidatos, key=lambda t: (-t[0], t[1], t[2])):
        if i in pares or j in usados:
            continue
        pares[i] = j
        usados.add(j)

    # ── registros por item (lado original) ──
    registros = []
    for i, o in enumerate(orig):
        j = pares.get(i)
        base = {
            "descricao": o["descricao"][:200],
            "disciplina": o["disciplina"],
            "confidence": o["confidence"],
            "unit": o["unit"],
            "qty_original": round(o["qty"], 4),
        }
        if j is None:
            registros.append({**base, "qty_revisada": None,
                              "delta_pct": None, "acao": "removido"})
            continue
        r = rev[j]
        qo, qr = o["qty"], r["qty"]
        if abs(qr - qo) <= max(1e-9, _TOL_REL * abs(qo)):
            registros.append({**base, "qty_revisada": round(qr, 4),
                              "delta_pct": 0.0, "acao": "mantido"})
        else:
            delta = round((qr - qo) / qo * 100.0, 2) if qo > 0 else None
            registros.append({**base, "qty_revisada": round(qr, 4),
                              "delta_pct": delta, "acao": "alterado"})

    # ── adicionados (lado revisado sem par) ──
    adicionados = []
    for j, r in enumerate(rev):
        if j in usados:
            continue
        adicionados.append({
            "descricao": r["descricao"][:200],
            "disciplina": r["disciplina"],
            "confidence": None,   # não existia → não tem selo original
            "unit": r["unit"],
            "qty_original": None,
            "qty_revisada": round(r["qty"], 4),
            "delta_pct": None,
            "acao": "adicionado",
        })

    # ── agregados ──
    por_disciplina = {}
    discs = {x["disciplina"] for x in registros} | {x["disciplina"] for x in adicionados}
    for d in sorted(discs):
        regs_d = [x for x in registros if x["disciplina"] == d]
        n_add = sum(1 for x in adicionados if x["disciplina"] == d)
        por_disciplina[d] = _bucket(regs_d, n_add)

    # A pergunta de ouro: o recorte por confidence
    por_confidence = {}
    for conf in ("confirmado", "estimado"):
        por_confidence[conf] = _bucket(
            [x for x in registros if x["confidence"] == conf])

    totais = _bucket(registros, len(adicionados))
    totais["n_originais"] = len(orig)
    totais["n_revisados"] = len(rev)

    return {
        "totais": totais,
        "por_disciplina": por_disciplina,
        "por_confidence": por_confidence,
        "itens": registros + adicionados,
    }


# ═══════════════════════════════════════════════
#  3. salvar_feedback (+ busca dos originais)
# ═══════════════════════════════════════════════

def _service_headers() -> dict:
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "Content-Type": "application/json",
    }


def _insert_supabase(table: str, record: dict) -> bool:
    """POST /rest/v1/{table} com service_role (padrão do calibrator)."""
    try:
        url = f"{SUPABASE_URL}/rest/v1/{table}"
        body = json.dumps(record).encode("utf-8")
        req = urllib.request.Request(url, data=body, method="POST")
        for k, v in _service_headers().items():
            req.add_header(k, v)
        req.add_header("Prefer", "return=minimal")
        urllib.request.urlopen(req, timeout=10)
        return True
    except Exception as e:
        print(f"[revision_feedback] insert {table} falhou: {type(e).__name__}: {e}")
        return False


def buscar_itens_originais(job_id: str) -> list[dict]:
    """Itens originais do job direto da tabela project_items (service_role).

    É a MESMA fonte da planilha gerada — não dependemos de reparsear o nosso
    próprio XLSX (que pode nem existir mais no disco do Render)."""
    try:
        jid = urllib.parse.quote(str(job_id), safe="")
        url = (f"{SUPABASE_URL}/rest/v1/project_items?job_id=eq.{jid}"
               f"&select=item_num,description,unit,quantity,confidence,discipline"
               f"&order=sort_order.asc&limit=3000")
        req = urllib.request.Request(url, method="GET")
        for k, v in _service_headers().items():
            req.add_header(k, v)
        resp = urllib.request.urlopen(req, timeout=10)
        rows = json.loads(resp.read().decode("utf-8"))
        return rows if isinstance(rows, list) else []
    except Exception as e:
        print(f"[revision_feedback] buscar originais job={job_id}: {type(e).__name__}: {e}")
        return []


def salvar_feedback(job_id: str, resultado: dict, arquivo: str = "") -> bool:
    """Grava o resultado de comparar() na tabela revision_feedback."""
    if not resultado or not isinstance(resultado, dict):
        return False
    t = resultado.get("totais") or {}
    record = {
        "job_id": str(job_id),
        "arquivo": str(arquivo or "")[:200],
        "n_originais": int(t.get("n_originais") or 0),
        "n_revisados": int(t.get("n_revisados") or 0),
        "n_mantidos": int(t.get("n_mantidos") or 0),
        "n_alterados": int(t.get("n_alterados") or 0),
        "n_removidos": int(t.get("n_removidos") or 0),
        "n_adicionados": int(t.get("n_adicionados") or 0),
        "mediana_abs_delta_pct": t.get("mediana_abs_delta_pct"),
        "por_disciplina": resultado.get("por_disciplina") or {},
        "por_confidence": resultado.get("por_confidence") or {},
        "itens": (resultado.get("itens") or [])[:MAX_ITENS_DETALHE],
    }
    return _insert_supabase("revision_feedback", record)


def processar_revisao(job_id: str, xlsx_path: str) -> bool:
    """Pipeline completo best-effort: busca originais → parse → compara →
    salva. Pensado pra rodar em thread no fim do upload da planilha revisada.
    Nunca levanta exceção."""
    try:
        originais = buscar_itens_originais(job_id)
        revisados = parse_planilha_revisada(xlsx_path)
        if not originais or not revisados:
            print(f"[revision_feedback] job={job_id} sem base pra comparar "
                  f"(originais={len(originais)}, revisados={len(revisados)})")
            return False
        resultado = comparar(originais, revisados)
        ok = salvar_feedback(job_id, resultado,
                             arquivo=os.path.basename(xlsx_path))
        t = resultado["totais"]
        print(f"[revision_feedback] job={job_id} salvo={ok} "
              f"alterados={t['n_alterados']}/{t['n_itens']} "
              f"removidos={t['n_removidos']} adicionados={t['n_adicionados']}")
        return ok
    except Exception as e:
        print(f"[revision_feedback] processar_revisao job={job_id}: "
              f"{type(e).__name__}: {e}")
        return False


# ═══════════════════════════════════════════════
#  4. resumo_para_admin
# ═══════════════════════════════════════════════

def _novo_acc() -> dict:
    return {"n_itens": 0, "n_alterados": 0, "n_removidos": 0,
            "n_adicionados": 0, "deltas": []}


def _fechar_acc(acc: dict) -> dict:
    deltas = acc.pop("deltas", [])
    acc["pct_alterados"] = (round(100.0 * acc["n_alterados"] / acc["n_itens"], 1)
                            if acc["n_itens"] else None)
    acc["mediana_abs_delta_pct"] = (round(statistics.median(deltas), 1)
                                    if deltas else None)
    return acc


def resumo_para_admin(rows: list[dict]) -> dict:
    """Agrega as rows da tabela revision_feedback pro painel do admin.

    Recalcula do detalhe `itens` (medianas EXATAS agregando todos os deltas,
    não mediana de medianas). Row antiga sem `itens` entra só nas contagens.
    """
    rows = rows or []
    disc_acc = {}
    conf_acc = {"confirmado": _novo_acc(), "estimado": _novo_acc()}
    tot = _novo_acc()
    ultimas = []

    for row in rows:
        if not isinstance(row, dict):
            continue
        itens = row.get("itens") or []
        if itens:
            for it in itens:
                if not isinstance(it, dict):
                    continue
                acao = it.get("acao")
                disc = str(it.get("disciplina") or "Complementares")
                d = disc_acc.setdefault(disc, _novo_acc())
                if acao == "adicionado":
                    d["n_adicionados"] += 1
                    tot["n_adicionados"] += 1
                    continue
                c = conf_acc[_norm_confidence(it.get("confidence"))]
                for acc in (d, c, tot):
                    acc["n_itens"] += 1
                if acao == "alterado":
                    dp = it.get("delta_pct")
                    for acc in (d, c, tot):
                        acc["n_alterados"] += 1
                        if isinstance(dp, (int, float)):
                            acc["deltas"].append(abs(dp))
                elif acao == "removido":
                    for acc in (d, c, tot):
                        acc["n_removidos"] += 1
        else:
            # Row sem detalhe (legado/planilha gigante): só contagens globais.
            tot["n_itens"] += int(row.get("n_originais") or 0)
            tot["n_alterados"] += int(row.get("n_alterados") or 0)
            tot["n_removidos"] += int(row.get("n_removidos") or 0)
            tot["n_adicionados"] += int(row.get("n_adicionados") or 0)

        ultimas.append({
            "job_id": row.get("job_id", ""),
            "created_at": row.get("created_at", ""),
            "arquivo": row.get("arquivo", ""),
            "n_originais": row.get("n_originais", 0),
            "n_alterados": row.get("n_alterados", 0),
            "n_removidos": row.get("n_removidos", 0),
            "n_adicionados": row.get("n_adicionados", 0),
        })

    por_disciplina = []
    for disc, acc in disc_acc.items():
        fechado = _fechar_acc(acc)
        fechado["disciplina"] = disc
        por_disciplina.append(fechado)
    # Onde a IA mais erra primeiro: % alterados desc, depois volume.
    por_disciplina.sort(key=lambda x: (-(x["pct_alterados"] if x["pct_alterados"]
                                         is not None else -1.0), -x["n_itens"]))

    return {
        "n_planilhas": len(ultimas),
        "totais": _fechar_acc(tot),
        "por_disciplina": por_disciplina,
        "por_confidence": {k: _fechar_acc(v) for k, v in conf_acc.items()},
        "ultimas": ultimas[:10],
    }
