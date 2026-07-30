# -*- coding: utf-8 -*-
"""Agente "tira-dúvidas" da planilha — Claude + ferramentas pra investigar
um projeto específico (job_id) e responder perguntas do cliente.

Loop padrão de agente:
  user → Claude → (opcional) chamada de tool → resultado da tool → Claude → resposta final

Tools disponíveis (todas operam no escopo de UM job_id):
  - list_items: lista resumida de itens da planilha
  - get_item_details: dados completos + observação de UM item específico
  - search_items: busca por palavra-chave na descrição
  - read_dxf_summary: estatísticas de um DXF (layers, blocos, walls)
  - check_density: roda check_density_anomaly num item específico
"""
import json
import os
import re
import urllib.request
from typing import Any, Optional

from openpyxl import load_workbook


SUPABASE_URL = os.getenv("SUPABASE_URL", "https://kqjabzwgbfuivzlcfvvu.supabase.co")
SUPABASE_KEY = os.getenv("SUPABASE_KEY") or os.getenv("SUPABASE_ANON_KEY") or "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImtxamFiendnYmZ1aXZ6bGNmdnZ1Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzYwMDg5NzcsImV4cCI6MjA5MTU4NDk3N30.48xSenZlDV0LfD94ZxwGvX41Kf9Je2n-ouZpJrrCSKI"
# `apikey` continua sendo a anon (PostgREST exige). O `Authorization: Bearer`
# usa a service_role pra ler project_items/agent_conversations etc. depois que a
# RLS fechar pra anon. service_role roda só server-side, nunca exposta ao frontend.
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or SUPABASE_KEY

WORK_DIR = os.path.join(os.environ.get("TMPDIR", "/tmp"), "aiarq_jobs")


# ════════════════════════════════════════════════════════════════
#  Tools — funções Python que o agente pode chamar
# ════════════════════════════════════════════════════════════════

def _planilha_path(job_id: str) -> Optional[str]:
    """Path local da planilha. Se sumiu (Render restart), tenta baixar
    do Supabase Storage via helper do main.py."""
    local = os.path.join(WORK_DIR, job_id, f"orcamento_{job_id}.xlsx")
    if os.path.exists(local):
        return local
    try:
        from main import get_planilha_path as _get
        return _get(job_id)
    except Exception:
        return None


def _open_planilha(job_id: str):
    path = _planilha_path(job_id)
    if not path or not os.path.exists(path):
        return None
    return load_workbook(path, read_only=True, data_only=True)


def _iter_orcamento_rows(wb):
    """Itera linhas da aba Orçamento, devolvendo dicts."""
    if not wb or "Orçamento" not in wb.sheetnames:
        return
    ws = wb["Orçamento"]
    for row in ws.iter_rows(min_row=1, max_col=9, values_only=True):
        if not row or len(row) < 4:
            continue
        item_num = row[0]
        if not isinstance(item_num, str):
            continue
        if not re.match(r"^\d+\.\d+", item_num.strip()):
            continue
        yield {
            "item_num": item_num.strip(),
            "description": str(row[1] or "").strip(),
            "unit": str(row[2] or "").strip(),
            "quantity": row[3],
            "observations": str(row[7] or "").strip() if len(row) > 7 else "",
            "ref_sheet": str(row[8] or "").strip() if len(row) > 8 else "",
        }


def tool_list_items(job_id: str, max_items: int = 200) -> dict:
    """Lista itens da planilha — número, descrição (40 chars), unit, qty."""
    wb = _open_planilha(job_id)
    if wb is None:
        return {"error": f"planilha do job {job_id} não encontrada"}
    items = []
    for r in _iter_orcamento_rows(wb):
        items.append({
            "item_num": r["item_num"],
            "description": r["description"][:80],
            "unit": r["unit"],
            "quantity": r["quantity"],
        })
        if len(items) >= max_items:
            break
    wb.close()
    return {"count": len(items), "items": items}


def tool_get_item_details(job_id: str, item_num: str) -> dict:
    """Dados completos de UM item + classificação automática (família,
    grupo, capítulo) e atributos folha extraídos (cor, PD, marca, etc.)."""
    wb = _open_planilha(job_id)
    if wb is None:
        return {"error": f"planilha do job {job_id} não encontrada"}
    found = None
    for r in _iter_orcamento_rows(wb):
        if r["item_num"] == item_num.strip():
            found = r
            break
    wb.close()
    if not found:
        return {"error": f"item {item_num} não encontrado"}

    # Enriquece com classificação (LLM Haiku, ~3s)
    try:
        from classifier import classify_item
        cls = classify_item(found["description"], found["unit"])
        found["categoria"] = {
            "capitulo": cls.get("capitulo_code"),
            "grupo": cls.get("grupo_code"),
            "familia": cls.get("familia_code"),
            "confidence": round(cls.get("confidence", 0), 2),
        }
        found["atributos_folha"] = cls.get("attributes") or {}
    except Exception as e:
        found["categoria"] = {"error": str(e)[:100]}
    return found


def tool_search_items(job_id: str, query: str, max_hits: int = 20) -> dict:
    """Busca itens cuja descrição contenha o termo (case-insensitive)."""
    wb = _open_planilha(job_id)
    if wb is None:
        return {"error": f"planilha do job {job_id} não encontrada"}
    q = query.lower().strip()
    hits = []
    for r in _iter_orcamento_rows(wb):
        if q in r["description"].lower():
            hits.append({
                "item_num": r["item_num"],
                "description": r["description"][:120],
                "unit": r["unit"],
                "quantity": r["quantity"],
                "observation_preview": r["observations"][:120],
            })
            if len(hits) >= max_hits:
                break
    wb.close()
    return {"query": query, "count": len(hits), "items": hits}


def tool_read_dxf_summary(job_id: str, dxf_filename: str = "") -> dict:
    """Estatísticas dos DXFs de um job — layers, blocos, walls.
    Se dxf_filename vazio, lista os arquivos disponíveis."""
    work = os.path.join(WORK_DIR, job_id)
    if not os.path.isdir(work):
        return {"error": f"work dir do job {job_id} não encontrado"}
    dxfs = [f for f in os.listdir(work) if f.lower().endswith(".dxf")]
    if not dxf_filename:
        return {"available_dxfs": dxfs}
    target = next((f for f in dxfs if dxf_filename.lower() in f.lower()), None)
    if not target:
        return {"error": f"dxf '{dxf_filename}' não encontrado entre {dxfs}"}
    try:
        from dwg_extractor import extract_from_file
        result = extract_from_file(os.path.join(work, target))
        # Resumo compacto pra não estourar contexto
        return {
            "filename": target,
            "blocks_count": len(result.get("blocks", [])),
            "walls_count": len(result.get("walls", [])),
            "layers": result.get("layers", [])[:50],
            "areas_m2": result.get("areas_m2", []),
        }
    except Exception as e:
        return {"error": f"erro ao ler DXF: {type(e).__name__}: {e}"}


def tool_get_sinapi_codes(familia_code: str, query: str = "",
                           top_k: int = 5) -> dict:
    """Lista códigos SINAPI mapeados pra uma família. Se `query` informada,
    filtra por similaridade simples na descrição (palavras-chave)."""
    if not familia_code:
        return {"error": "familia_code obrigatório"}
    try:
        # Pega o id da família
        url = (f"{SUPABASE_URL}/rest/v1/catalog_familia"
               f"?select=id,name&code=eq.{familia_code}")
        req = urllib.request.Request(url, method="GET")
        req.add_header("apikey", SUPABASE_KEY)
        req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
        req.add_header("Accept", "application/json")
        resp = urllib.request.urlopen(req, timeout=10)
        fams = json.loads(resp.read().decode("utf-8"))
        if not fams:
            return {"error": f"família '{familia_code}' não encontrada"}
        familia_id = fams[0]["id"]
        familia_nome = fams[0]["name"]

        # Pega SINAPIs da família
        url = (f"{SUPABASE_URL}/rest/v1/sinapi_composicao"
               f"?select=codigo,descricao,unidade&familia_id=eq.{familia_id}"
               f"&limit=50")
        req = urllib.request.Request(url, method="GET")
        req.add_header("apikey", SUPABASE_KEY)
        req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
        req.add_header("Accept", "application/json")
        resp = urllib.request.urlopen(req, timeout=10)
        rows = json.loads(resp.read().decode("utf-8"))

        # Se query informada, ranqueia por keyword overlap
        if query and rows:
            q_tokens = set(re.findall(r"\w{4,}", query.lower()))
            for r in rows:
                d_tokens = set(re.findall(r"\w{4,}", (r.get("descricao") or "").lower()))
                r["_score"] = len(q_tokens & d_tokens)
            rows.sort(key=lambda r: -r.get("_score", 0))
            for r in rows:
                r.pop("_score", None)

        return {
            "familia_code": familia_code,
            "familia_nome": familia_nome,
            "total_sinapis_mapeados": len(rows),
            "top": rows[:top_k],
        }
    except Exception as e:
        return {"error": f"erro: {type(e).__name__}: {e}"}


def tool_check_density_for_item(job_id: str, item_num: str,
                                 typology: str = "office") -> dict:
    """Roda check_density_anomaly num item específico da planilha."""
    details = tool_get_item_details(job_id, item_num)
    if "error" in details:
        return details
    # Busca area do projeto no Supabase
    project = _supabase_select_project(job_id)
    ref_area = (project.get("layout_area") or project.get("total_area") or 0) if project else 0
    typ = (project.get("typology") if project else None) or typology

    # cria objeto-like pra check
    class _Item:
        pass
    it = _Item()
    it.description = details["description"]
    it.unit = details["unit"]
    try:
        it.quantity = float(details["quantity"])
    except Exception:
        it.quantity = 0

    try:
        from density_calibration import check_density_anomaly
        is_anom, msg = check_density_anomaly(it, float(ref_area or 0), typology=typ)
    except Exception as e:
        return {"error": f"check error: {e}"}
    return {
        "item_num": item_num,
        "description": details["description"],
        "qty": details["quantity"],
        "unit": details["unit"],
        "ref_area_m2": ref_area,
        "typology": typ,
        "is_anomaly": is_anom,
        "alert": msg or "sem alerta — densidade dentro do padrão (ou sem benchmark)",
    }


def tool_list_supplier_quotes(job_id: str) -> dict:
    """Lista cotações de fornecedores submetidas neste projeto.
    Retorna: [{supplier_name, n_items_quoted, total_bruto, total_material,
    total_mao_obra}].
    """
    try:
        url = (f"{SUPABASE_URL}/rest/v1/project_supplier_quotes"
               f"?job_id=eq.{job_id}"
               f"&select=id,supplier_name,n_items_quoted,total_bruto,"
               f"total_material,total_mao_obra,status"
               f"&order=uploaded_at.asc")
        req = urllib.request.Request(url, method="GET")
        req.add_header("apikey", SUPABASE_KEY)
        req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
        resp = urllib.request.urlopen(req, timeout=8)
        quotes = json.loads(resp.read().decode("utf-8"))
        return {
            "job_id": job_id,
            "n_quotes": len(quotes),
            "quotes": quotes,
        }
    except Exception as e:
        return {"error": str(e)}


def tool_compare_supplier_quotes(job_id: str) -> dict:
    """Compara as cotações de fornecedores deste projeto entre si e contra
    o quantitativo original do AI.arq.

    Retorna análise completa: ranking por total pareado, cobertura de cada
    fornecedor, maiores discrepâncias item-a-item, e (se aplicável) itens
    do quantitativo que fornecedores esqueceram.
    """
    try:
        from supplier_quote_compare import compare_quotes
    except ImportError:
        return {"error": "supplier_quote_compare indisponível"}

    # Busca quotes
    try:
        url = (f"{SUPABASE_URL}/rest/v1/project_supplier_quotes"
               f"?job_id=eq.{job_id}&select=*&order=uploaded_at.asc")
        req = urllib.request.Request(url, method="GET")
        req.add_header("apikey", SUPABASE_KEY)
        req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
        resp = urllib.request.urlopen(req, timeout=8)
        quotes_raw = json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        return {"error": str(e)}

    if len(quotes_raw) < 2:
        return {"error": "precisa de pelo menos 2 cotações"}

    quotes = [{
        "supplier_name": q["supplier_name"],
        "n_items_quoted": q.get("n_items_quoted", 0),
        "total_bruto": float(q.get("total_bruto") or 0),
        "total_material": float(q.get("total_material") or 0),
        "total_mao_obra": float(q.get("total_mao_obra") or 0),
        "items": q.get("items") or [],
    } for q in quotes_raw]

    # Busca reference items
    reference_items = None
    try:
        ref_url = (f"{SUPABASE_URL}/rest/v1/project_items"
                   f"?job_id=eq.{job_id}&select=description,unit,quantity")
        ref_req = urllib.request.Request(ref_url, method="GET")
        ref_req.add_header("apikey", SUPABASE_KEY)
        ref_req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
        ref_resp = urllib.request.urlopen(ref_req, timeout=8)
        reference_items = json.loads(ref_resp.read().decode("utf-8"))
    except Exception:
        pass

    analysis = compare_quotes(quotes, reference_items=reference_items)

    # Retorna versão compacta (o agente não precisa de tudo)
    return {
        "suppliers": analysis["suppliers"],
        "n_items_comparados": analysis["paired_count"],
        "ranking": [{"fornecedor": s, "total_pareado": t}
                    for s, t in analysis["ranking"]],
        "cobertura_por_fornecedor": {
            sup: f"{cov['pct']:.0f}% ({cov['n_priced']}/{cov['n_total_unique']})"
            for sup, cov in analysis["coverage"].items()
        },
        "top_10_discrepancias": [
            {
                "descricao": d["desc"],
                "diferenca_pct": f"{d['pct_diff']:.0f}%",
                "mais_barato": d["cheapest"],
                "mais_caro": d["most_expensive"],
                "valores": {k: f"R$ {v:,.2f}" for k, v in d["totals"].items()},
            }
            for d in analysis["biggest_discrepancies"][:10]
        ],
        "vs_quantitativo_aiarq": analysis.get("reference_check", {}).get("summary")
            if analysis.get("reference_check") else None,
    }


def tool_check_market_heuristics(description: str, typology: str = "office") -> dict:
    """Checa um item contra as heurísticas de mercado (dispersão, cobertura,
    share MAT/MO) extraídas de orçamentos reais anonimizados.

    Retorna alertas curtos + métricas agregadas. NUNCA valor absoluto.
    """
    try:
        from market_heuristics import (
            categorize_item, check_item_anomaly,
            get_dispersion_for_category, get_mat_mo_share_for_category,
            get_coverage_pattern_for_category,
        )
    except ImportError:
        return {"error": "market_heuristics indisponível"}

    if not description or len(description.strip()) < 3:
        return {"error": "descrição muito curta"}

    cat = categorize_item(description)
    alertas = check_item_anomaly({"description": description, "unit": ""},
                                   typology=typology)

    return {
        "categoria": cat,
        "tipologia": typology,
        "alertas": alertas,
        "dispersao_mercado": get_dispersion_for_category(cat, typology),
        "share_mat_mo_tipico": get_mat_mo_share_for_category(cat, typology),
        "cobertura_tipica": get_coverage_pattern_for_category(cat, typology),
        "obs": "Heurísticas de mercado anônimas — agregado de orçamentos reais. "
                "Use pra orientar o cliente sobre variação esperada, NÃO "
                "como valor de referência pra esse projeto específico.",
    }


def _supabase_select_project(job_id: str) -> Optional[dict]:
    try:
        url = f"{SUPABASE_URL}/rest/v1/projects?job_id=eq.{job_id}&select=*"
        req = urllib.request.Request(url, method="GET")
        req.add_header("apikey", SUPABASE_KEY)
        req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
        req.add_header("Accept", "application/json")
        resp = urllib.request.urlopen(req, timeout=8)
        rows = json.loads(resp.read().decode("utf-8"))
        return rows[0] if rows else None
    except Exception:
        return None


# ════════════════════════════════════════════════════════════════
#  Schema das tools (formato Anthropic Tool Use)
# ════════════════════════════════════════════════════════════════

TOOLS = [
    {
        "name": "list_items",
        "description": "Lista itens da planilha de quantitativos (número, descrição, unidade, quantidade). Use pra ter visão geral do que existe.",
        "input_schema": {
            "type": "object",
            "properties": {
                "max_items": {"type": "integer", "description": "Limite de itens (default 200)"},
            },
        },
    },
    {
        "name": "get_item_details",
        "description": "Retorna dados completos de UM item, incluindo observação inteira (cita fonte/layer CAD/consolidador) E classificação automática (capítulo > grupo > família) e atributos folha extraídos (cor, PD, marca, dimensão, código produto). Use SEMPRE que o usuário perguntar 'por que tem X' ou pedir detalhes de um item específico — a categoria + atributos enriquecem a explicação.",
        "input_schema": {
            "type": "object",
            "properties": {
                "item_num": {"type": "string", "description": "Número do item, ex: '6.3' ou '8.1'"},
            },
            "required": ["item_num"],
        },
    },
    {
        "name": "search_items",
        "description": "Busca itens da planilha por palavra-chave na descrição. Use quando o usuário menciona um termo (LED, alvenaria, forro etc).",
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Termo de busca"},
            },
            "required": ["query"],
        },
    },
    {
        "name": "read_dxf_summary",
        "description": "Estatísticas de um DXF (layers, qtd de blocos, walls). Sem dxf_filename, lista os DXFs do projeto.",
        "input_schema": {
            "type": "object",
            "properties": {
                "dxf_filename": {"type": "string", "description": "Nome (ou parte) do DXF. Vazio pra listar."},
            },
        },
    },
    {
        "name": "check_density_for_item",
        "description": "Compara a densidade (qty/área) de um item contra o padrão histórico da mesma tipologia. Usa pra justificar quantidades.",
        "input_schema": {
            "type": "object",
            "properties": {
                "item_num": {"type": "string", "description": "Número do item"},
            },
            "required": ["item_num"],
        },
    },
    {
        "name": "get_sinapi_codes",
        "description": "Lista códigos SINAPI mapeados pra uma família do catálogo. Use quando o cliente perguntar 'qual código SINAPI desse item?' ou pedir referência oficial pra compra. Fornece familia_code (você obtém de get_item_details.categoria.familia) e opcionalmente query (descrição original do item) pra ranqueamento por similaridade.",
        "input_schema": {
            "type": "object",
            "properties": {
                "familia_code": {"type": "string", "description": "Código da família, ex: 'fam_pint_acrilica' (vem de get_item_details.categoria.familia)"},
                "query": {"type": "string", "description": "Descrição do item original pra ranquear melhores matches (opcional)"},
                "top_k": {"type": "integer", "description": "Quantos códigos retornar (default 5)"},
            },
            "required": ["familia_code"],
        },
    },
    {
        "name": "list_supplier_quotes",
        "description": "Lista as cotações de fornecedores que o cliente enviou pra este projeto (planilhas de orçamento recebidas dos fornecedores). Use quando o cliente perguntar 'quais fornecedores já mandaram proposta?', 'quantas cotações eu tenho?', 'qual fornecedor é mais barato?' (pra este último, use depois compare_supplier_quotes).",
        "input_schema": {
            "type": "object",
            "properties": {},
        },
    },
    {
        "name": "compare_supplier_quotes",
        "description": "Compara todas as cotações de fornecedores submetidas no projeto. Retorna ranking por total pareado (só itens que todos orçaram), cobertura de cada fornecedor, top 10 maiores discrepâncias item-a-item, e verificação contra o quantitativo original (itens esquecidos, divergências de quantidade). Use quando o cliente perguntar 'compara os orçamentos', 'qual o mais caro em elétrica?', 'quem esqueceu ar-condicionado?'.",
        "input_schema": {
            "type": "object",
            "properties": {},
        },
    },
    {
        "name": "check_market_heuristics",
        "description": "Checa heurísticas agregadas de mercado (dispersão de preço entre fornecedores, share típico de material vs mão de obra, padrão de cobertura) pra uma categoria de item. Use quando o cliente perguntar 'quanto varia o preço disso?', 'é normal esse item ser 80% material?', 'esse serviço costuma ser esquecido?'. Responde SEM valores absolutos — só ratios e percentuais. Úteis pra orientar a revisão e pra explicar intervalo esperado de cotações. A base é anônima (agregada de orçamentos reais, sem identificar projetos).",
        "input_schema": {
            "type": "object",
            "properties": {
                "description": {"type": "string", "description": "Descrição do item (ex: 'demolição de drywall', 'luminária LED 60x60')"},
                "typology": {"type": "string", "description": "office|residential|retail|hospital|educational (default office)"},
            },
            "required": ["description"],
        },
    },
]


def _dispatch_tool(name: str, job_id: str, tool_input: dict) -> Any:
    if name == "list_items":
        return tool_list_items(job_id, tool_input.get("max_items", 200))
    if name == "get_item_details":
        return tool_get_item_details(job_id, tool_input["item_num"])
    if name == "search_items":
        return tool_search_items(job_id, tool_input["query"])
    if name == "read_dxf_summary":
        return tool_read_dxf_summary(job_id, tool_input.get("dxf_filename", ""))
    if name == "check_density_for_item":
        return tool_check_density_for_item(job_id, tool_input["item_num"])
    if name == "get_sinapi_codes":
        return tool_get_sinapi_codes(
            tool_input["familia_code"],
            tool_input.get("query", ""),
            tool_input.get("top_k", 5),
        )
    if name == "list_supplier_quotes":
        return tool_list_supplier_quotes(job_id)
    if name == "compare_supplier_quotes":
        return tool_compare_supplier_quotes(job_id)
    if name == "check_market_heuristics":
        return tool_check_market_heuristics(
            tool_input["description"],
            tool_input.get("typology", "office"),
        )
    return {"error": f"tool '{name}' desconhecida"}


# ════════════════════════════════════════════════════════════════
#  Loop do agente
# ════════════════════════════════════════════════════════════════

SYSTEM_PROMPT = """Você é o assistente do AI.arq — atende clientes que usam o site pra gerar orçamento de obra a partir de plantas (DWG/PDF).

CONTEXTO DESTA CONVERSA:
- O cliente está perguntando sobre UM projeto específico (job_id={job_id}).
- Você tem ferramentas pra ler a planilha gerada, buscar itens, ler DXFs e checar calibração.

REGRAS:
- Use as ferramentas pra investigar antes de responder. NUNCA invente número.
- Se uma ferramenta retornar erro (ex.: {{"error": ...}} — a planilha pode ter sumido após um reinício do servidor), NÃO chute um valor: diga que não conseguiu acessar agora e que é só recarregar/reprocessar. Nunca preencha um número que você não leu de uma ferramenta.
- MEDIDO vs ESTIMADO: ao citar uma quantidade, diga se o item está "✓ medido do CAD" (confirmado) ou "⚠ estimativa pra revisar" (estimado) — isso está na observação/selo do item. Nunca apresente uma estimativa como certeza; deixe claro o que é chão firme e o que o usuário precisa conferir.
- O AI.arq NÃO precifica. Se pedirem preço, valor ou custo, explique que o quantitativo sai sem preço de propósito — quem precifica é o orçamentista, com o BDI e os fornecedores dele. Você ajuda com as quantidades, não com R$.
- Quando citar um item, mencione o item_num e cite a observação que justifica a quantidade.
- Se o usuário perguntar "por que essa quantidade?", busca o item, leia a observação (que cita layer CAD ou processo de consolidação) e explique.
- Respostas curtas (3-5 frases). Use linguagem comum, sem jargão técnico de IA.
- Se a pergunta sair do escopo do quantitativo, redirecione: "Não tenho acesso a isso, posso ajudar com itens da sua planilha?"
"""


def _log_conversation(job_id: str, question: str, answer: str,
                      tool_calls: list, iterations: int, duration_ms: int,
                      error: str = "") -> None:
    """Persiste a conversa em agent_conversations pra auditoria/admin."""
    try:
        record = {
            "job_id": job_id,
            "question": question[:2000],
            "answer": (answer or "")[:5000],
            "tool_calls": tool_calls[:30],  # limita pra não estourar
            "iterations": iterations,
            "duration_ms": duration_ms,
            "error": (error or "")[:500] or None,
        }
        url = f"{SUPABASE_URL}/rest/v1/agent_conversations"
        body = json.dumps(record, default=str, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(url, data=body, method="POST")
        req.add_header("apikey", SUPABASE_KEY)
        req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_ROLE_KEY}")
        req.add_header("Content-Type", "application/json")
        req.add_header("Prefer", "return=minimal")
        urllib.request.urlopen(req, timeout=8)
    except Exception as e:
        print(f"[agent] log error: {e}")
    # 🚨 O chat é o melhor detector de buraco do motor que a gente tem — e era
    # CEGO. Em 21/07 um cliente reclamou de eletroduto faltando e a resposta
    # explicou a causa raiz ("a planta não tem polilinha fechada"); ficou 9 dias
    # numa tabela sem ninguém ler. Das 11 conversas de sempre, 3 eram falha real
    # de medição, com job_id anexado. Agora isso vira alerta. (30/07/2026)
    try:
        _alerta_lacuna(job_id, question)
    except Exception as _e:
        print(f"[agent] alerta de lacuna falhou (nao-fatal): {_e}")


# Sinais de que o cliente está APONTANDO UMA FALHA, não tirando dúvida.
_LACUNA_RE = re.compile(
    r"(n[ãa]o\s+(tem|veio|apareceu|consta|calculou|mediu|bateu)|"
    r"falt(a|ou|aram|ando)|sem\s+(quantitativo|medi[çc][ãa]o|quantidade)|"
    r"(t[áa]|est[áa])\s+(em\s+branco|zerad|vazi|errad)|"
    r"n[ãa]o\s+apresentou|incomplet|"
    r"(pq|por\s*que|porqu[êe])\s+(n[ãa]o|nao))",
    re.IGNORECASE,
)


def _alerta_lacuna(job_id: str, question: str) -> None:
    """Avisa o dono quando o cliente aponta falta de medição no chat.

    Best-effort e silencioso: nunca pode atrapalhar a resposta ao cliente.
    Deduplica por job (1 alerta por projeto) pra não virar spam em conversa
    longa — o objetivo é sinalizar o projeto, não cada frase.
    """
    if not question or not _LACUNA_RE.search(question):
        return
    try:
        from main import _notify_admin, _email_auto_registrar, _email_auto_ja_enviado, NOTIFY_EMAIL
    except Exception:
        return
    _ref = f"chat:{job_id}"
    if _email_auto_ja_enviado(NOTIFY_EMAIL, "alerta_chat_lacuna", ref=_ref):
        return
    import html as _h
    _corpo = (
        f"<b>Um cliente apontou falta de medição pelo chat.</b><br><br>"
        f"<b>Pergunta:</b> {_h.escape(question[:400])}<br>"
        f"<b>Projeto:</b> {_h.escape(job_id)}<br><br>"
        f"Isso costuma ser buraco do motor, não dúvida do cliente — vale olhar "
        f"o que a planilha deixou de medir nesse projeto.<br><br>"
        f'<a href="https://ai.arq.br/admin.html#projetos">Abrir no painel</a>'
    )
    if _notify_admin(f"Chat: cliente diz que faltou medição — {job_id}", _corpo):
        _email_auto_registrar(NOTIFY_EMAIL, "alerta_chat_lacuna", ref=_ref)


def ask(job_id: str, question: str, max_iterations: int = 8,
        history: Optional[list] = None) -> dict:
    """Roda o loop do agente até ele dar resposta final.

    Args:
        job_id: identificador do projeto (escopo das tools)
        question: pergunta atual do cliente
        history: opcional, lista de {role, content} de turnos anteriores
                 da MESMA conversa. Permite o cliente perguntar "e o item 3.4?"
                 como continuação. Se None, conversa inicia do zero.

    Retorna {answer, tool_calls: [(name, input, result)...], iterations}.
    Loga conversa em agent_conversations no Supabase.
    """
    import time as _t
    t0 = _t.time()

    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        result = {"answer": "API key não configurada.", "tool_calls": [], "iterations": 0}
        _log_conversation(job_id, question, result["answer"], [], 0,
                          int((_t.time()-t0)*1000), "no api key")
        return result

    try:
        import anthropic
    except ImportError:
        result = {"answer": "SDK anthropic não instalado.", "tool_calls": [], "iterations": 0}
        _log_conversation(job_id, question, result["answer"], [], 0,
                          int((_t.time()-t0)*1000), "no anthropic SDK")
        return result

    client = anthropic.Anthropic(api_key=api_key)

    # Monta histórico — turnos anteriores + pergunta atual.
    # Filtra cada entrada do history pra ficar só com {role, content} string
    # (sem tool_use blocks intermediários, que não fazem sentido fora da
    # iteração original e podem confundir o modelo se não pareados com tool_result).
    messages: list = []
    if history:
        for turn in history[-20:]:  # limita pra não estourar contexto
            role = turn.get("role")
            content = turn.get("content")
            if role in ("user", "assistant") and isinstance(content, str) and content.strip():
                messages.append({"role": role, "content": content.strip()[:4000]})
    messages.append({"role": "user", "content": question})
    tool_calls_log = []
    final_answer = ""

    from llm_retry import call_with_retry
    for it in range(max_iterations):
        try:
            resp = call_with_retry(
                client,
                tag=f"agent:job={job_id}",
                max_retries=3,
                model="claude-sonnet-4-6",
                max_tokens=2000,
                system=SYSTEM_PROMPT.format(job_id=job_id),
                tools=TOOLS,
                messages=messages,
            )
        except Exception as e:
            err_msg = f"Erro na chamada Claude: {type(e).__name__}: {e}"
            _log_conversation(job_id, question, err_msg, tool_calls_log, it,
                              int((_t.time()-t0)*1000), str(e)[:300])
            return {"answer": err_msg, "tool_calls": tool_calls_log,
                    "iterations": it}

        # Coleta texto + tool_use blocks
        text_chunks = []
        tool_uses = []
        for block in resp.content:
            if block.type == "text":
                text_chunks.append(block.text)
            elif block.type == "tool_use":
                tool_uses.append(block)

        # Se não houve tool_use, é a resposta final
        if not tool_uses:
            final_answer = "\n".join(text_chunks).strip()
            break

        # Adiciona a resposta do assistant ao histórico
        messages.append({"role": "assistant", "content": resp.content})

        # Executa cada tool e devolve o resultado
        tool_results = []
        for tu in tool_uses:
            try:
                result = _dispatch_tool(tu.name, job_id, tu.input)
            except Exception as e:
                result = {"error": f"tool exception: {e}"}
            tool_calls_log.append({
                "name": tu.name, "input": tu.input,
                "result_preview": json.dumps(result, default=str, ensure_ascii=False)[:300],
            })
            tool_results.append({
                "type": "tool_result",
                "tool_use_id": tu.id,
                "content": json.dumps(result, default=str, ensure_ascii=False)[:8000],
            })
        messages.append({"role": "user", "content": tool_results})

    if not final_answer:
        final_answer = "Não consegui formular uma resposta após várias iterações."

    duration_ms = int((_t.time() - t0) * 1000)
    _log_conversation(job_id, question, final_answer, tool_calls_log,
                      it + 1, duration_ms)

    return {
        "answer": final_answer,
        "tool_calls": tool_calls_log,
        "iterations": it + 1,
        "duration_ms": duration_ms,
    }
