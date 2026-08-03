# -*- coding: utf-8 -*-
"""Cronograma físico-financeiro em .xlsx — mesmo padrão visual do quantitativo.

Pedido do Pedro (03/08/2026): *"vamos deixar a planilha no mesmo padrão de
formatação do quantitativo"*. Por isso os estilos são **importados** de
`spreadsheet.py`, não copiados: cópia diverge no primeiro ajuste de paleta e
aí ficam duas identidades visuais no mesmo produto.

🔒 Regra dura nº5 — o AI.arq NÃO precifica. Todo valor aqui foi digitado pelo
cliente etapa por etapa; a planilha só distribui pelos meses no mesmo ritmo do
cronograma físico e soma. Não há tabela de preço, sugestão de valor nem BDI.

🪤 Diferença proposital em relação à planilha de mercado (Sienge/Prevision, que
o Pedro mandou como referência): lá o usuário digita o desembolso **mês a mês**
(N etapas × M meses de digitação). Aqui ele digita **um valor por etapa** e o
rateio sai do próprio cronograma — e continua batendo quando ele arrasta uma
fase no Gantt. É a mesma razão da regra nº7: número derivado não pode
envelhecer em silêncio.
"""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Mesma paleta/tipografia do quantitativo — importada, nunca duplicada.
from spreadsheet import (
    F_TITLE, F_SEC, F_SUB, F_HDR, F_N, F_BOLD, F_TOT, F_NOTE,
    P_SEC, P_SUB, P_HDR, P_TOT, P_LT,
    AC, AL, AR, BD,
)

FMT_MOEDA = 'R$ #,##0.00'
FMT_PCT = '0.0%'


def _cabecalho(ws, branding: dict, tem_financeiro: bool, n_cols: int) -> int:
    """Bloco de identificação. Devolve a próxima linha livre."""
    titulo = ('CRONOGRAMA FÍSICO-FINANCEIRO DA OBRA' if tem_financeiro
              else 'CRONOGRAMA FÍSICO DA OBRA')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(6, min(n_cols, 12)))
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = F_TITLE
    c.alignment = AL

    ro = 2
    for rot, val in (
        ('Projeto', branding.get('project_name') or 'Projeto sem nome'),
        ('Escritório', branding.get('architect_name') or branding.get('company') or ''),
        ('Cliente', branding.get('client_name') or ''),
        ('Emitido em', branding.get('emitido_em') or ''),
    ):
        if not val:
            continue
        ws.cell(row=ro, column=1, value=rot).font = F_HDR
        ws.cell(row=ro, column=2, value=val).font = F_N
        ro += 1
    return ro + 1


def _aviso_origem(ws, ro: int, fin: dict, n_cols: int) -> int:
    """A frase que mantém a regra nº1 e a nº5 visíveis DENTRO do arquivo.

    Vale mais aqui do que na tela: o .xlsx sai do site, vai por e-mail e chega
    no banco sem nenhum contexto em volta."""
    if not fin:
        return ro
    ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=max(6, min(n_cols, 12)))
    c = ws.cell(row=ro, column=1, value=(
        'Os valores desta planilha foram INFORMADOS POR VOCÊ, etapa por etapa. '
        'O AI.arq não precifica obra: ele apenas distribui o seu valor pelos meses, '
        'seguindo o mesmo ritmo do cronograma físico. Confira antes de enviar a terceiros.'
    ))
    c.font = F_NOTE
    c.alignment = AL
    ro += 1
    if fin.get('n_fases_com_valor', 0) < fin.get('n_fases', 0):
        ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=max(6, min(n_cols, 12)))
        c = ws.cell(row=ro, column=1, value=(
            f"ATENÇÃO — preenchimento PARCIAL: {fin['n_fases_com_valor']} de "
            f"{fin['n_fases']} etapas têm valor. O total abaixo é só do que foi "
            f"preenchido, NÃO é o custo da obra inteira."
        ))
        c.font = F_SUB
        c.fill = P_SUB
        c.alignment = AL
        ro += 1
    return ro + 1


def gerar_cronograma_xlsx(cronograma: dict, output_path: str,
                          branding: dict = None) -> str:
    """Escreve o .xlsx e devolve o caminho."""
    branding = branding or {}
    fases = cronograma.get('fases') or []
    meses = cronograma.get('meses') or []
    matriz = cronograma.get('matriz_pct') or []
    fin = cronograma.get('financeiro') or None
    tem_fin = bool(fin and fin.get('total_informado'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Cronograma'

    # 🪤 Vocabulário PRÓPRIO, de propósito (Pedro, 03/08: "vamos fazer parecido,
    # não igual"). A planilha de referência é produto de um concorrente — os
    # rótulos dela ("Descrição da Etapa", "Total do mês", "Total acumulado")
    # não entram aqui. Usamos as palavras que o resto do AI.arq já usa:
    # disciplina, fase, desembolso. Ver feedback_nao_falar_mal_concorrente:
    # nem copiar, nem cutucar.
    FIXAS = ['Nº', 'FASE / DISCIPLINA', 'INÍCIO', 'FIM', 'DURAÇÃO (DIAS)', '% EXECUTADO']
    if tem_fin:
        FIXAS.append('VALOR DA FASE (R$)')
    n_cols = len(FIXAS) + len(meses)

    ro = _cabecalho(ws, branding, tem_fin, n_cols)
    ro = _aviso_origem(ws, ro, fin if tem_fin else None, n_cols)

    # ── faixa de agrupamento das colunas de mês
    if meses:
        ws.merge_cells(start_row=ro, start_column=len(FIXAS) + 1,
                       end_row=ro, end_column=len(FIXAS) + len(meses))
        c = ws.cell(row=ro, column=len(FIXAS) + 1,
                    value=('DESEMBOLSO POR MÊS (R$) — distribuição do seu valor'
                           if tem_fin else '% EXECUTADO POR MÊS'))
        c.font = F_SEC
        c.fill = P_SEC
        c.alignment = AC
    ro += 1

    # ── cabeçalho de colunas
    linha_hdr = ro
    for i, h in enumerate(FIXAS, start=1):
        c = ws.cell(row=ro, column=i, value=h)
        c.font = F_HDR
        c.fill = P_HDR
        c.alignment = AC
        c.border = BD
    for j, m in enumerate(meses, start=len(FIXAS) + 1):
        c = ws.cell(row=ro, column=j, value=m.get('label', ''))
        c.font = F_HDR
        c.fill = P_HDR
        c.alignment = AC
        c.border = BD
    ro += 1

    # ── uma linha por fase
    primeira_fase = ro
    for idx, fase in enumerate(fases):
        linha_matriz = matriz[idx] if idx < len(matriz) else {}
        pcts = linha_matriz.get('percentuais_por_mes') or []
        soma_pct = sum(pcts) or 0
        valor = float(fase.get('valor_previsto') or 0)
        alt = P_LT if idx % 2 else None

        vals = [idx + 1, fase.get('label', ''), fase.get('inicio', ''),
                fase.get('fim', ''), fase.get('dur_dias', 0),
                (float(fase.get('pct_executado') or 0) / 100.0)]
        if tem_fin:
            vals.append(valor)
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=ro, column=i, value=v)
            c.font = F_N
            c.border = BD
            c.alignment = AL if i == 2 else AC
            if alt:
                c.fill = alt
            if i == 6:
                c.number_format = FMT_PCT
            if tem_fin and i == 7:
                c.number_format = FMT_MOEDA
                c.alignment = AR

        # Rateio da linha ANTES de escrever: cada fase tem que fechar exatamente
        # no valor que o cliente digitou. Arredondar célula a célula deixava
        # 1 centavo sobrando no total — e centavo que não fecha é o tipo de
        # detalhe que faz um banco devolver a papelada.
        cel_valores = {}
        if tem_fin and valor > 0 and soma_pct > 0:
            for j, p in enumerate(pcts[:len(meses)]):
                if p:
                    cel_valores[j] = round(valor * p / soma_pct, 2)
            resto = round(valor - sum(cel_valores.values()), 2)
            if resto and cel_valores:
                ult = max(cel_valores)
                cel_valores[ult] = round(cel_valores[ult] + resto, 2)

        for j, p in enumerate(pcts[:len(meses)]):
            col = len(FIXAS) + 1 + j
            c = ws.cell(row=ro, column=col)
            c.border = BD
            c.alignment = AR if tem_fin else AC
            if alt:
                c.fill = alt
            if not p:
                continue
            if tem_fin and valor > 0 and soma_pct > 0:
                c.value = cel_valores.get(j, 0)
                c.number_format = FMT_MOEDA
                c.font = F_N
            else:
                c.value = p / 100.0
                c.number_format = FMT_PCT
                c.font = F_N
        ro += 1
    ultima_fase = ro - 1

    # ── totais do mês + acumulado (só fazem sentido com dinheiro)
    if tem_fin and meses and ultima_fase >= primeira_fase:
        for rot, acumulado in (('DESEMBOLSO DO MÊS', False), ('ACUMULADO NA OBRA', True)):
            ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=len(FIXAS))
            c = ws.cell(row=ro, column=1, value=rot)
            c.font = F_TOT
            c.fill = P_TOT
            c.alignment = AR
            c.border = BD
            for j in range(len(meses)):
                col = len(FIXAS) + 1 + j
                L = get_column_letter(col)
                if acumulado:
                    # Fórmula VIVA de propósito: quem receber a planilha pode
                    # ajustar um mês e ver o acumulado se corrigir sozinho,
                    # em vez de encontrar número velho colado.
                    prim = get_column_letter(len(FIXAS) + 1)
                    f = f"=SUM({prim}{ro-1}:{L}{ro-1})"
                else:
                    f = f"=SUM({L}{primeira_fase}:{L}{ultima_fase})"
                c2 = ws.cell(row=ro, column=col, value=f)
                c2.font = F_BOLD
                c2.fill = P_TOT
                c2.number_format = FMT_MOEDA
                c2.alignment = AR
                c2.border = BD
            ro += 1

        # total geral informado, confrontável com a soma dos meses
        ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=len(FIXAS))
        c = ws.cell(row=ro, column=1, value='TOTAL INFORMADO POR VOCÊ')
        c.font = F_TOT
        c.fill = P_SUB
        c.alignment = AR
        c.border = BD
        colv = get_column_letter(len(FIXAS))
        c2 = ws.cell(row=ro, column=len(FIXAS) + 1,
                     value=f"=SUM({colv}{primeira_fase}:{colv}{ultima_fase})")
        c2.font = F_TOT
        c2.fill = P_SUB
        c2.number_format = FMT_MOEDA
        c2.alignment = AR
        c2.border = BD
        ro += 1

    # ── larguras e painel congelado
    larguras = [8, 42, 12, 12, 14, 13] + ([16] if tem_fin else [])
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for j in range(len(meses)):
        ws.column_dimensions[get_column_letter(len(FIXAS) + 1 + j)].width = 15 if tem_fin else 9
    ws.freeze_panes = ws.cell(row=linha_hdr + 1, column=3)

    wb.save(output_path)
    return output_path
