# -*- coding: utf-8 -*-
"""Gera os dois modelos XLSX que os posts do blog prometem para download.

🪤 01/08/2026: os posts de 23/08 (cronograma) e 13/09 (planilha SINAPI) prometiam
modelo em Excel NO TÍTULO, mas o botão de download servia o memorial descritivo
em PDF/DOCX — o único par de arquivos que existia. Este script cria os arquivos
que faltavam, com o conteúdo que cada post descreve.

Rodar: python blog/downloads/gen_modelos_xlsx.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AQUI = os.path.dirname(os.path.abspath(__file__))

F = 'Calibri'
TIT = Font(name=F, bold=True, size=14, color='0F172A')
SEC = Font(name=F, bold=True, size=11, color='FFFFFF')
HDR = Font(name=F, bold=True, size=10, color='0F172A')
NOR = Font(name=F, size=10)
NOTA = Font(name=F, size=9, italic=True, color='64748B')
P_SEC = PatternFill('solid', fgColor='0F172A')
P_HDR = PatternFill('solid', fgColor='E2E8F0')
P_YEL = PatternFill('solid', fgColor='FFF0A6')
P_TOT = PatternFill('solid', fgColor='E8ECF4')
_S = Side(style='thin', color='D0D5DD')
BD = Border(left=_S, right=_S, top=_S, bottom=_S)
AL = Alignment(horizontal='left', vertical='center', wrap_text=True)
AC = Alignment(horizontal='center', vertical='center', wrap_text=True)
AR = Alignment(horizontal='right', vertical='center')

# As 18 disciplinas na ordem construtiva, iguais às listadas no post de 13/09.
DISCIPLINAS = [
    "1 Serviços preliminares", "2 Movimento de terra", "3 Fundações", "4 Estrutura",
    "5 Vedações verticais", "6 Cobertura", "7 Impermeabilizações", "8 Esquadrias",
    "9 Vidros e ferragens", "10 Revest. internos", "11 Revest. externos", "12 Forros",
    "13 Pisos", "14 Pinturas", "15 Inst. hidrossanitárias", "16 Inst. elétricas",
    "17 AVAC e exaustão", "18 Complementares e limpeza",
]

COLS = [("ITEM", 10), ("CÓDIGO SINAPI", 16), ("DESCRIÇÃO DO SERVIÇO", 58), ("UN", 8),
        ("QUANT.", 12), ("PREÇO UNIT.", 14), ("PREÇO TOTAL", 16), ("OBSERVAÇÕES", 34)]


def _cabecalho(ws, titulo):
    ws.merge_cells('A1:H1')
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = TIT; c.alignment = AL
    ws.row_dimensions[1].height = 26
    ws.merge_cells('A2:H2')
    c = ws.cell(row=2, column=1, value=(
        'Preencha QUANT. com o que você mediu do projeto. As colunas de preço ficam em branco '
        'de propósito: quem precifica é o orçamentista. Declare no topo a data e o estado de '
        'referência do SINAPI e se usou a versão desonerada ou não.'))
    c.font = NOTA; c.alignment = AL
    ws.row_dimensions[2].height = 30


def _tabela(ws, linha_ini, n_linhas=14):
    for i, (nome, larg) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larg
        c = ws.cell(row=linha_ini, column=i, value=nome)
        c.font = HDR; c.fill = P_HDR; c.alignment = AC; c.border = BD
    prim = linha_ini + 1
    for r in range(prim, prim + n_linhas):
        for i in range(1, 9):
            cel = ws.cell(row=r, column=i)
            cel.border = BD; cel.font = NOR
            if i in (6, 7):
                cel.number_format = '#,##0.00'
            if i == 6:
                cel.fill = P_YEL
        ws.cell(row=r, column=5).number_format = '#,##0.00'
        ws.cell(row=r, column=7, value=f'=IF(E{r}="","",E{r}*F{r})')
    ult = prim + n_linhas - 1
    ws.cell(row=ult + 1, column=1, value='SUBTOTAL DA DISCIPLINA').font = HDR
    ws.merge_cells(start_row=ult + 1, start_column=1, end_row=ult + 1, end_column=6)
    t = ws.cell(row=ult + 1, column=7, value=f'=SUM(G{prim}:G{ult})')
    t.font = HDR; t.number_format = '#,##0.00'; t.fill = P_TOT; t.alignment = AR
    for i in range(1, 9):
        ws.cell(row=ult + 1, column=i).fill = P_TOT
        ws.cell(row=ult + 1, column=i).border = BD
    return ult + 1


def modelo_quantitativos():
    """Post 13/09 — XLSX com as 18 disciplinas em abas separadas + consolidação."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'INSTRUÇÕES'
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 104
    linhas = [
        ('Planilha modelo de quantitativos por disciplina — AI.arq', TIT),
        ('', NOR),
        ('Como usar', HDR),
        ('1. Cada aba é uma disciplina. Preencha só as que existem no seu projeto.', NOR),
        ('2. Coluna CÓDIGO SINAPI: use o código da composição do mês de referência que você baixou '
         'da Caixa. Se o item não tem composição compatível, escreva COT (cotação direta).', NOR),
        ('3. Coluna UN: use SEMPRE a unidade da composição SINAPI referenciada. Não invente.', NOR),
        ('4. Coluna QUANT.: o número medido do projeto. O PREÇO TOTAL se calcula sozinho.', NOR),
        ('5. Colunas de preço: deixe pro orçamentista. Ele aplica preço regional e BDI.', NOR),
        ('6. Em OBSERVAÇÕES, marque a origem de cada número: MEDIDO (extraído da planta) ou '
         'ESTIMADO (suposição a revisar). Use a palavra, não só cor.', NOR),
        ('', NOR),
        ('Declare no cabeçalho de cada aba', HDR),
        ('Data e mês de referência do SINAPI · Estado (UF) · Versão desonerada ou não-desonerada. '
         'Sem isso, ninguém consegue reconstruir a planilha depois.', NOR),
        ('', NOR),
        ('Aviso', HDR),
        ('Este é um modelo em branco, para você preencher. Não contém preços nem quantidades — '
         'quantitativo é medição do seu projeto, e preço é trabalho do orçamentista habilitado.', NOTA),
    ]
    r = 2
    for txt, fonte in linhas:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=2)
        c = ws.cell(row=r, column=2, value=txt)
        c.font = fonte; c.alignment = AL
        r += 1

    for disc in DISCIPLINAS:
        aba = wb.create_sheet(disc[:31])
        _cabecalho(aba, disc.split(' ', 1)[1].upper())
        _tabela(aba, 4)

    cons = wb.create_sheet('CONSOLIDAÇÃO')
    cons.column_dimensions['A'].width = 6
    cons.column_dimensions['B'].width = 46
    cons.column_dimensions['C'].width = 18
    _c = cons.cell(row=1, column=2, value='CONSOLIDAÇÃO POR DISCIPLINA'); _c.font = TIT
    for i, (nome, _) in enumerate([('#', 0), ('DISCIPLINA', 0), ('SUBTOTAL', 0)], start=1):
        c = cons.cell(row=3, column=i, value=nome)
        c.font = HDR; c.fill = P_HDR; c.border = BD; c.alignment = AC
    r = 4
    for disc in DISCIPLINAS:
        num, nome = disc.split(' ', 1)
        cons.cell(row=r, column=1, value=int(num)).border = BD
        cons.cell(row=r, column=2, value=nome).border = BD
        ref = f"'{disc[:31]}'!G19"
        cel = cons.cell(row=r, column=3, value=f'={ref}')
        cel.number_format = '#,##0.00'; cel.border = BD
        r += 1
    cons.cell(row=r, column=2, value='TOTAL CUSTO DIRETO').font = HDR
    t = cons.cell(row=r, column=3, value=f'=SUM(C4:C{r-1})')
    t.font = HDR; t.number_format = '#,##0.00'; t.fill = P_TOT
    cons.cell(row=r + 2, column=2, value='BDI (%) — preencher').font = HDR
    b = cons.cell(row=r + 2, column=3, value=0); b.number_format = '0.00%'; b.fill = P_YEL
    cons.cell(row=r + 3, column=2, value='TOTAL COM BDI').font = HDR
    g = cons.cell(row=r + 3, column=3, value=f'=C{r}*(1+C{r+2})')
    g.font = HDR; g.number_format = '#,##0.00'; g.fill = P_TOT
    cons.cell(row=r + 5, column=2, value=(
        'O BDI acima é campo do orçamentista. A fórmula do Acórdão TCU 2622/2013 é '
        '[(1+AC+S+R+G) x (1+DF) x (1+L)] / (1-T) - 1, com T = tributos sobre FATURAMENTO '
        '(PIS+COFINS+ISS, e CPRB quando aplicável). IRPJ e CSLL não entram — Súmula TCU 254/2010.'
    )).font = NOTA

    destino = os.path.join(AQUI, 'quantitativo-sinapi-planilha-modelo.xlsx')
    wb.save(destino)
    return destino


def modelo_cronograma():
    """Post 23/08 — Gantt físico até 18 meses + financeiro com acumulado (curva S)."""
    MESES = 18
    wb = Workbook()

    ws = wb.active
    ws.title = 'INSTRUÇÕES'
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 104
    linhas = [
        ('Cronograma físico-financeiro de obra — modelo AI.arq', TIT),
        ('', NOR),
        ('Você preenche 3 coisas', HDR),
        ('1. Na aba FÍSICO: o percentual de avanço previsto de cada disciplina, mês a mês. '
         'A coluna de verificação mostra se a linha fecha 100%.', NOR),
        ('2. Na aba FINANCEIRO: o valor total de cada disciplina (vem do seu orçamento).', NOR),
        ('3. Só isso. O desembolso mensal, o acumulado e a curva S se calculam sozinhos, '
         'distribuindo o valor conforme o avanço físico que você lançou.', NOR),
        ('', NOR),
        ('Sobre a curva S', HDR),
        ('A curva S é a linha do desembolso ACUMULADO ao longo do tempo — está na linha '
         '"% acumulado" da aba FINANCEIRO. Selecione essa linha e insira um gráfico de linhas '
         'para visualizá-la.', NOR),
        ('', NOR),
        ('Aviso', HDR),
        ('Este modelo NÃO calcula o valor da obra. Ele distribui no tempo um valor que você já '
         'tem. Quem apura custo e preço é o orçamentista habilitado.', NOTA),
    ]
    r = 2
    for txt, fonte in linhas:
        c = ws.cell(row=r, column=2, value=txt); c.font = fonte; c.alignment = AL
        r += 1

    disc_cron = [d.split(' ', 1)[1] for d in DISCIPLINAS]

    fis = wb.create_sheet('FÍSICO')
    fis.column_dimensions['A'].width = 30
    c = fis.cell(row=1, column=1, value='AVANÇO FÍSICO PREVISTO (% por mês)'); c.font = TIT
    fis.cell(row=2, column=1, value='Preencha em %. A última coluna avisa se a linha fecha 100%.').font = NOTA
    h = fis.cell(row=3, column=1, value='DISCIPLINA'); h.font = HDR; h.fill = P_HDR; h.border = BD
    for m in range(1, MESES + 1):
        col = 1 + m
        fis.column_dimensions[get_column_letter(col)].width = 7
        c = fis.cell(row=3, column=col, value=f'M{m}')
        c.font = HDR; c.fill = P_HDR; c.alignment = AC; c.border = BD
    ck = 2 + MESES
    fis.column_dimensions[get_column_letter(ck)].width = 12
    c = fis.cell(row=3, column=ck, value='SOMA'); c.font = HDR; c.fill = P_HDR; c.alignment = AC; c.border = BD
    for i, d in enumerate(disc_cron):
        r = 4 + i
        fis.cell(row=r, column=1, value=d).font = NOR
        fis.cell(row=r, column=1).border = BD
        for m in range(1, MESES + 1):
            cel = fis.cell(row=r, column=1 + m)
            cel.number_format = '0%'; cel.fill = P_YEL; cel.border = BD
        s = fis.cell(row=r, column=ck,
                     value=f'=IF(SUM(B{r}:{get_column_letter(1+MESES)}{r})=0,"",'
                           f'SUM(B{r}:{get_column_letter(1+MESES)}{r}))')
        s.number_format = '0%'; s.border = BD; s.font = HDR

    fin = wb.create_sheet('FINANCEIRO')
    fin.column_dimensions['A'].width = 30
    fin.column_dimensions['B'].width = 18
    c = fin.cell(row=1, column=1, value='DESEMBOLSO POR MÊS (calculado)'); c.font = TIT
    fin.cell(row=2, column=1, value='Preencha só a coluna VALOR TOTAL. O resto é fórmula.').font = NOTA
    for nome, col in (('DISCIPLINA', 1), ('VALOR TOTAL (R$)', 2)):
        c = fin.cell(row=3, column=col, value=nome)
        c.font = HDR; c.fill = P_HDR; c.border = BD; c.alignment = AC
    for m in range(1, MESES + 1):
        col = 2 + m
        fin.column_dimensions[get_column_letter(col)].width = 12
        c = fin.cell(row=3, column=col, value=f'M{m}')
        c.font = HDR; c.fill = P_HDR; c.alignment = AC; c.border = BD
    for i, d in enumerate(disc_cron):
        r = 4 + i
        fin.cell(row=r, column=1, value=d).font = NOR
        fin.cell(row=r, column=1).border = BD
        v = fin.cell(row=r, column=2); v.number_format = '#,##0.00'; v.fill = P_YEL; v.border = BD
        for m in range(1, MESES + 1):
            cel = fin.cell(row=r, column=2 + m,
                           value=f'=IF($B{r}="","",$B{r}*FÍSICO!{get_column_letter(1+m)}{r})')
            cel.number_format = '#,##0.00'; cel.border = BD
    ult = 3 + len(disc_cron)
    lin_mes = ult + 1
    fin.cell(row=lin_mes, column=1, value='DESEMBOLSO DO MÊS').font = HDR
    tot = fin.cell(row=lin_mes, column=2, value=f'=SUM(B4:B{ult})')
    tot.font = HDR; tot.number_format = '#,##0.00'; tot.fill = P_TOT
    for m in range(1, MESES + 1):
        L = get_column_letter(2 + m)
        c = fin.cell(row=lin_mes, column=2 + m, value=f'=SUM({L}4:{L}{ult})')
        c.number_format = '#,##0.00'; c.font = HDR; c.fill = P_TOT
    lin_ac = lin_mes + 1
    fin.cell(row=lin_ac, column=1, value='ACUMULADO').font = HDR
    for m in range(1, MESES + 1):
        L = get_column_letter(2 + m)
        ant = get_column_letter(1 + m)
        formula = f'={L}{lin_mes}' if m == 1 else f'={ant}{lin_ac}+{L}{lin_mes}'
        c = fin.cell(row=lin_ac, column=2 + m, value=formula)
        c.number_format = '#,##0.00'; c.font = HDR
    lin_pct = lin_ac + 1
    fin.cell(row=lin_pct, column=1, value='% ACUMULADO (curva S)').font = HDR
    for m in range(1, MESES + 1):
        L = get_column_letter(2 + m)
        c = fin.cell(row=lin_pct, column=2 + m,
                     value=f'=IF($B{lin_mes}=0,"",{L}{lin_ac}/$B{lin_mes})')
        c.number_format = '0.0%'; c.font = HDR
    fin.cell(row=lin_pct + 2, column=1, value=(
        'Selecione a linha "% ACUMULADO" e insira um gráfico de linhas: é a curva S da obra.'
    )).font = NOTA

    destino = os.path.join(AQUI, 'cronograma-fisico-financeiro-modelo.xlsx')
    wb.save(destino)
    return destino


if __name__ == '__main__':
    for f in (modelo_quantitativos(), modelo_cronograma()):
        print(f'  {os.path.basename(f)}  {os.path.getsize(f)/1024:.0f} KB')
