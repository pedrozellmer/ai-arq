# -*- coding: utf-8 -*-
"""Carrega a aba 'Analítico' do SINAPI Referência mensal para a tabela
sinapi_insumos do Supabase.

Estrutura da aba (após header na linha 10):
    Grupo | Código da Composição | Tipo Item | Código do Item |
    Descrição | Unidade | Coeficiente | Situação

Uso:
    python load_sinapi_analitico.py <xlsx_path> [--month YYYY-MM-DD] [--dry-run]
"""
import argparse
import json
import os
import sys
import urllib.request
from datetime import date

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

SUPABASE_URL = os.getenv("SUPABASE_URL", "https://kqjabzwgbfuivzlcfvvu.supabase.co")
SUPABASE_KEY = os.getenv("SUPABASE_KEY") or (
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9."
    "eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImtxamFiendnYmZ1aXZ6bGNmdnZ1Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzYwMDg5NzcsImV4cCI6MjA5MTU4NDk3N30."
    "48xSenZlDV0LfD94ZxwGvX41Kf9Je2n-ouZpJrrCSKI"
)


def parse_analitico(xlsx_path, sync_month="2026-03-01"):
    """Parseia a aba Analítico. Retorna lista de dicts."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # Procura aba "Analítico" (pode ter variações de acento)
    ws = None
    for sn in wb.sheetnames:
        if "anal" in sn.lower() and "custo" not in sn.lower():
            ws = wb[sn]
            break
    if ws is None:
        raise RuntimeError("Aba 'Analítico' não encontrada")

    # Acha linha do header (geralmente linha 10)
    header_row = None
    for i in range(1, 30):
        cells = [ws.cell(row=i, column=c).value for c in range(1, 9)]
        joined = " ".join(str(c) for c in cells if c).upper()
        if "GRUPO" in joined and "CÓDIGO" in joined and "COEFICIENTE" in joined:
            header_row = i
            break

    if not header_row:
        raise RuntimeError("Header 'Grupo | Código | Coeficiente' não encontrado")

    print(f"  Header na linha {header_row}")

    items = []
    current_comp = None

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or all(c is None or c == "" for c in row):
            continue
        grupo = row[0] if len(row) > 0 else None
        comp_codigo = row[1] if len(row) > 1 else None
        tipo = row[2] if len(row) > 2 else None
        cod_item = row[3] if len(row) > 3 else None
        desc = row[4] if len(row) > 4 else None
        unidade = row[5] if len(row) > 5 else None
        coef = row[6] if len(row) > 6 else None
        situacao = row[7] if len(row) > 7 else None

        if not comp_codigo:
            continue
        comp_codigo = str(comp_codigo).strip()
        if not comp_codigo or comp_codigo == "None":
            continue

        # Quando tipo_item vazio, é a linha-cabeçalho da composição principal
        # (não é insumo — é descrição da composição mãe). Pulamos.
        if not tipo:
            current_comp = comp_codigo
            continue

        # Valida tipo
        tipo_s = str(tipo).strip().upper()
        if tipo_s not in ("COMPOSICAO", "INSUMO"):
            continue

        try:
            coef_f = float(coef) if coef is not None else None
        except (ValueError, TypeError):
            coef_f = None

        items.append({
            "composicao_codigo": comp_codigo,
            "grupo": str(grupo).strip() if grupo else "",
            "tipo_item": tipo_s,
            "codigo_item": str(cod_item).strip() if cod_item else "",
            "descricao": str(desc).strip() if desc else "",
            "unidade": str(unidade).strip() if unidade else "",
            "coeficiente": coef_f,
            "situacao": str(situacao).strip() if situacao else "",
            "sync_month": sync_month,
        })

    return items


def insert_batch(rows):
    if not rows:
        return 0
    url = f"{SUPABASE_URL}/rest/v1/sinapi_insumos"
    body = json.dumps(rows).encode("utf-8")
    req = urllib.request.Request(url, data=body, method="POST")
    req.add_header("apikey", SUPABASE_KEY)
    req.add_header("Authorization", f"Bearer {SUPABASE_KEY}")
    req.add_header("Content-Type", "application/json")
    req.add_header("Prefer", "return=minimal")
    try:
        urllib.request.urlopen(req, timeout=30)
        return len(rows)
    except urllib.error.HTTPError as e:
        err_body = e.read().decode("utf-8", errors="replace")[:200]
        print(f"  [erro] {e.code}: {err_body}")
        return 0
    except Exception as e:
        print(f"  [erro] {e}")
        return 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="Caminho do SINAPI_Referência_AAAA_MM.xlsx")
    ap.add_argument("--month", default="2026-03-01")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--batch", type=int, default=1000)
    args = ap.parse_args()

    print(f"Parseando: {args.xlsx}")
    items = parse_analitico(args.xlsx, sync_month=args.month)
    print(f"\n{len(items)} insumos/subcomposições extraídos")

    # Stats
    tipos = {}
    for it in items:
        tipos[it["tipo_item"]] = tipos.get(it["tipo_item"], 0) + 1
    print(f"  Por tipo: {tipos}")
    n_comps = len(set(it["composicao_codigo"] for it in items))
    print(f"  Composições únicas: {n_comps}")

    if args.dry_run:
        print("\nDRY RUN — amostra dos 10 primeiros:")
        for it in items[:10]:
            print(f"  {it['composicao_codigo']:<8} | {it['tipo_item']:<10} | "
                  f"{it['codigo_item']:<8} | coef={it['coeficiente']} | "
                  f"{it['descricao'][:60]}")
        return

    # Insere em batches
    total = 0
    for i in range(0, len(items), args.batch):
        batch = items[i:i + args.batch]
        n = insert_batch(batch)
        total += n
        if (i + args.batch) % (args.batch * 10) == 0:
            print(f"  [{i + args.batch}/{len(items)}] inseridos: {total}")

    print(f"\nOK: {total} linhas no Supabase")


if __name__ == "__main__":
    main()
