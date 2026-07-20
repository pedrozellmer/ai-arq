# -*- coding: utf-8 -*-
"""Testes do revision_feedback ("onde a IA mais erra").

Gera uma planilha no formato REAL (chamando spreadsheet.generate_spreadsheet,
não uma imitação), simula as edições que um cliente faz (3 qty alteradas,
1 linha removida, 1 adicionada, 1 renomeada de leve) e prova pareamento,
deltas e agregados. Roda em segundos: `python tests/test_revision_feedback.py`.
"""
import io
import os
import sys
import tempfile

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))

from openpyxl import Workbook, load_workbook  # noqa: E402

import revision_feedback as rf  # noqa: E402
from models import ProjectData, BudgetItem, Confidence  # noqa: E402
from spreadsheet import generate_spreadsheet  # noqa: E402

_passed = 0
_failed = 0


def check(nome, cond, extra=""):
    global _passed, _failed
    if cond:
        _passed += 1
        print(f"  ok  {nome}")
    else:
        _failed += 1
        print(f"  XX  FALHOU: {nome}{('  [' + str(extra) + ']') if extra else ''}")


TMP = tempfile.mkdtemp(prefix="rf_test_")

# ═══════════════════════════════════════════════
#  Fixtures: planilha REAL gerada pelo spreadsheet.py
# ═══════════════════════════════════════════════

ITEMS = [
    BudgetItem(item_num="1.1", description="Demolição de alvenaria existente",
               unit="m2", quantity=45.0, confidence=Confidence.CONFIRMADO,
               discipline="Demolição e Remoção", origem="dxf_geom"),
    BudgetItem(item_num="1.2", description="Retirada de piso vinílico",
               unit="m2", quantity=120.0, confidence=Confidence.ESTIMADO,
               discipline="Demolição e Remoção"),
    BudgetItem(item_num="2.1", description="Alvenaria de vedação bloco cerâmico 14cm",
               unit="m2", quantity=88.5, confidence=Confidence.CONFIRMADO,
               discipline="Fechamentos Verticais", origem="dxf_geom"),
    BudgetItem(item_num="2.2", description="Parede drywall chapa dupla ST",
               unit="m2", quantity=60.0, confidence=Confidence.CONFIRMADO,
               discipline="Fechamentos Verticais", origem="dxf_geom"),
    BudgetItem(item_num="3.1", description="Pintura latex acrilica em paredes",
               unit="m2", quantity=300.0, confidence=Confidence.ESTIMADO,
               discipline="Revestimentos"),
    BudgetItem(item_num="4.1", description="Porcelanato acetinado 60x60 assentado",
               unit="m2", quantity=210.0, confidence=Confidence.ESTIMADO,
               discipline="Pisos e Rodapés"),
    BudgetItem(item_num="4.2", description="Rodapé poliestireno h=10cm",
               unit="m", quantity=95.0, confidence=Confidence.ESTIMADO,
               discipline="Pisos e Rodapés"),
    BudgetItem(item_num="5.1", description="Forro de gesso acartonado liso",
               unit="m2", quantity=150.0, confidence=Confidence.CONFIRMADO,
               discipline="Forros", origem="dxf_geom"),
]

PROJECT = ProjectData(name="Teste Revisão", architect="DTZ",
                      total_area=800.0, layout_area=520.0)

ORIG_XLSX = os.path.join(TMP, "original.xlsx")
generate_spreadsheet(PROJECT, ITEMS, ORIG_XLSX, typology="office")

# itens_originais como vêm da tabela project_items (fonte de verdade do motor)
ORIGINAIS = [{
    "item_num": it.item_num,
    "description": it.description,
    "unit": it.unit,
    "quantity": it.quantity,
    "confidence": it.confidence.value,
    "discipline": it.discipline,
} for it in ITEMS]


# ═══════════════════════════════════════════════
#  1. parse da planilha ORIGINAL (formato real, sem edição)
# ═══════════════════════════════════════════════

print("== parse_planilha_revisada: planilha real intocada ==")
parsed = rf.parse_planilha_revisada(ORIG_XLSX)
check("extrai exatamente os 8 itens", len(parsed) == 8, f"veio {len(parsed)}")
descrs = [p["description"] for p in parsed]
check("descrições preservadas",
      "Porcelanato acetinado 60x60 assentado" in descrs
      and "Demolição de alvenaria existente" in descrs)
check("premissas (0.x) ficam FORA",
      all(not p["item_num"].startswith("0.") for p in parsed)
      and all("Área" not in p["description"] for p in parsed))
check("sugestões (S.x) e resumo ficam FORA",
      all("Equipe técnica" not in p["description"] for p in parsed)
      and all("SUBTOTAL" not in p["description"] for p in parsed))
_pq = {p["description"]: p["quantity"] for p in parsed}
check("quantidades corretas", _pq.get("Retirada de piso vinílico") == 120.0
      and _pq.get("Alvenaria de vedação bloco cerâmico 14cm") == 88.5)
_pd = {p["description"]: p["discipline"] for p in parsed}
check("disciplina canônica igual à do banco",
      _pd.get("Porcelanato acetinado 60x60 assentado") == "Pisos e Rodapés"
      and _pd.get("Demolição de alvenaria existente") == "Demolição e Remoção",
      _pd)
_ps = {p["description"]: p["selo_medido"] for p in parsed}
check("selo medido/estimado lido da obs",
      _ps.get("Demolição de alvenaria existente") is True
      and _ps.get("Retirada de piso vinílico") is False)


# ═══════════════════════════════════════════════
#  2. simula a REVISÃO do cliente (edições reais de Excel)
# ═══════════════════════════════════════════════

REV_XLSX = os.path.join(TMP, "revisada.xlsx")
wb = load_workbook(ORIG_XLSX)
ws = wb["Orçamento"]


def find_row(desc):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=2).value or "").strip() == desc:
            return r
    raise AssertionError(f"linha não achada: {desc}")


# 3 quantidades alteradas (2 estimados + 1 confirmado)
ws.cell(row=find_row("Retirada de piso vinílico"), column=4).value = 90        # -25%
ws.cell(row=find_row("Porcelanato acetinado 60x60 assentado"), column=4).value = 240  # +14,29%
ws.cell(row=find_row("Parede drywall chapa dupla ST"), column=4).value = 55    # -8,33%
# 1 renomeada DE LEVE (acentos + palavra extra; qty igual → tem que parear e ser "mantido")
ws.cell(row=find_row("Pintura latex acrilica em paredes"), column=2).value = \
    "Pintura látex acrílica em paredes internas"
# 1 linha ADICIONADA pelo cliente (sem numeração na coluna A)
_r = find_row("Porcelanato acetinado 60x60 assentado")
ws.insert_rows(_r + 1)
ws.cell(row=_r + 1, column=2).value = "Manta acústica sob contrapiso"
ws.cell(row=_r + 1, column=3).value = "m2"
ws.cell(row=_r + 1, column=4).value = 50
# 1 linha REMOVIDA (estimado)
ws.delete_rows(find_row("Rodapé poliestireno h=10cm"))
wb.save(REV_XLSX)
wb.close()

print("== parse da planilha REVISADA (editada) ==")
revisados = rf.parse_planilha_revisada(REV_XLSX)
check("8 itens (8 - 1 removida + 1 adicionada)", len(revisados) == 8,
      f"veio {len(revisados)}: {[p['description'] for p in revisados]}")
_rd = {p["description"]: p for p in revisados}
check("linha adicionada sem item_num entrou",
      "Manta acústica sob contrapiso" in _rd
      and _rd["Manta acústica sob contrapiso"]["item_num"] == ""
      and _rd["Manta acústica sob contrapiso"]["quantity"] == 50.0)
check("linha adicionada herda a disciplina da seção",
      _rd["Manta acústica sob contrapiso"]["discipline"] == "Pisos e Rodapés")
check("linha removida sumiu", "Rodapé poliestireno h=10cm" not in _rd)
check("rename de leve preservado", "Pintura látex acrílica em paredes internas" in _rd)


# ═══════════════════════════════════════════════
#  3. comparar: pareamento, deltas, agregados
# ═══════════════════════════════════════════════

print("== comparar: pareamento e deltas ==")
res = rf.comparar(ORIGINAIS, revisados)
t = res["totais"]
check("n_originais=8 / n_revisados=8",
      t["n_originais"] == 8 and t["n_revisados"] == 8, t)
check("3 alterados", t["n_alterados"] == 3, t)
check("1 removido", t["n_removidos"] == 1, t)
check("1 adicionado", t["n_adicionados"] == 1, t)
check("4 mantidos (renomeada pareou e conta como mantido)",
      t["n_mantidos"] == 4, t)
check("pct_alterados = 37.5", t["pct_alterados"] == 37.5, t)

_by_desc = {x["descricao"]: x for x in res["itens"]}
check("renomeada de leve pareou como MANTIDO (não removido+adicionado)",
      _by_desc["Pintura latex acrilica em paredes"]["acao"] == "mantido")
check("delta retirada de piso = -25%",
      _by_desc["Retirada de piso vinílico"]["delta_pct"] == -25.0,
      _by_desc["Retirada de piso vinílico"])
check("delta porcelanato = +14.29%",
      _by_desc["Porcelanato acetinado 60x60 assentado"]["delta_pct"] == 14.29)
check("delta drywall = -8.33%",
      _by_desc["Parede drywall chapa dupla ST"]["delta_pct"] == -8.33)
check("rodapé marcado como removido",
      _by_desc["Rodapé poliestireno h=10cm"]["acao"] == "removido"
      and _by_desc["Rodapé poliestireno h=10cm"]["qty_revisada"] is None)
check("manta marcada como adicionado (sem confidence)",
      _by_desc["Manta acústica sob contrapiso"]["acao"] == "adicionado"
      and _by_desc["Manta acústica sob contrapiso"]["confidence"] is None)
check("mediana |delta| geral = 14.3 (de 25 / 14.29 / 8.33)",
      t["mediana_abs_delta_pct"] == 14.3, t)

print("== agregados por disciplina ==")
pdisc = res["por_disciplina"]
pr = pdisc.get("Pisos e Rodapés", {})
check("Pisos e Rodapés: 2 itens, 1 alterado, 1 removido, 1 adicionado",
      pr.get("n_itens") == 2 and pr.get("n_alterados") == 1
      and pr.get("n_removidos") == 1 and pr.get("n_adicionados") == 1, pr)
fv = pdisc.get("Fechamentos Verticais", {})
check("Fechamentos Verticais: 2 itens, 1 alterado, mediana 8.3",
      fv.get("n_itens") == 2 and fv.get("n_alterados") == 1
      and fv.get("mediana_abs_delta_pct") == 8.3, fv)

print("== a pergunta de ouro: recorte por confidence ==")
pc = res["por_confidence"]
conf = pc.get("confirmado", {})
est = pc.get("estimado", {})
check("confirmado: 4 itens, 1 alterado (25%), 0 removidos",
      conf.get("n_itens") == 4 and conf.get("n_alterados") == 1
      and conf.get("pct_alterados") == 25.0 and conf.get("n_removidos") == 0, conf)
check("estimado: 4 itens, 2 alterados (50%), 1 removido",
      est.get("n_itens") == 4 and est.get("n_alterados") == 2
      and est.get("pct_alterados") == 50.0 and est.get("n_removidos") == 1, est)
check("estimado apanha MAIS que confirmado (medido sobrevive)",
      est["pct_alterados"] > conf["pct_alterados"])
check("medianas por confidence: confirmado 8.3 / estimado 19.6",
      conf.get("mediana_abs_delta_pct") == 8.3
      and est.get("mediana_abs_delta_pct") == 19.6, (conf, est))


# ═══════════════════════════════════════════════
#  4. robustez: vazia / corrompida / listas vazias
# ═══════════════════════════════════════════════

print("== robustez ==")
EMPTY_XLSX = os.path.join(TMP, "vazia.xlsx")
Workbook().save(EMPTY_XLSX)
check("planilha vazia → [] sem explodir",
      rf.parse_planilha_revisada(EMPTY_XLSX) == [])

CORRUPT = os.path.join(TMP, "corrompida.xlsx")
with open(CORRUPT, "wb") as f:
    f.write(b"isto nao e um xlsx \x00\x01\x02")
check("arquivo corrompido → [] sem explodir",
      rf.parse_planilha_revisada(CORRUPT) == [])
check("arquivo inexistente → [] sem explodir",
      rf.parse_planilha_revisada(os.path.join(TMP, "nao_existe.xlsx")) == [])

vazio = rf.comparar([], [])
check("comparar([], []) → totais zerados sem explodir",
      vazio["totais"]["n_itens"] == 0
      and vazio["totais"]["pct_alterados"] is None
      and vazio["itens"] == [])

so_orig = rf.comparar(ORIGINAIS, [])
check("revisada vazia → tudo vira removido",
      so_orig["totais"]["n_removidos"] == 8)

# célula de qty com FÓRMULA sem cache não explode nem inventa número
check("_parse_qty('=D7*2') → None", rf._parse_qty("=D7*2") is None)
check("_parse_qty('1.234,56') → 1234.56", rf._parse_qty("1.234,56") == 1234.56)
check("_parse_qty(None) → None", rf._parse_qty(None) is None)

# fuzzy NÃO pareia coisas diferentes (conservador)
res_x = rf.comparar(
    [{"description": "Porta de madeira 80cm", "unit": "un", "quantity": 5,
      "confidence": "estimado", "discipline": "Portas e Ferragens"}],
    [{"description": "Janela de alumínio 120cm", "unit": "un", "quantity": 5,
      "discipline": "Portas e Ferragens"}])
check("fuzzy conservador: porta ≠ janela (vira removido + adicionado)",
      res_x["totais"]["n_removidos"] == 1 and res_x["totais"]["n_adicionados"] == 1)

# unidade incompatível bloqueia pareamento fuzzy
res_u = rf.comparar(
    [{"description": "Piso vinilico manta", "unit": "m2", "quantity": 100,
      "confidence": "estimado", "discipline": "Pisos e Rodapés"}],
    [{"description": "Piso vinilico manta rodape", "unit": "m", "quantity": 40,
      "discipline": "Pisos e Rodapés"}])
check("fuzzy: m² não pareia com metro linear",
      res_u["totais"]["n_removidos"] == 1 and res_u["totais"]["n_adicionados"] == 1)


# ═══════════════════════════════════════════════
#  5. salvar_feedback (insert stubado — sem rede)
# ═══════════════════════════════════════════════

print("== salvar_feedback (stub, sem rede) ==")
_captura = {}
_original_insert = rf._insert_supabase
rf._insert_supabase = lambda table, record: (_captura.update(
    {"table": table, "record": record}) or True)
try:
    ok = rf.salvar_feedback("job-teste-123", res, arquivo="revisada.xlsx")
    check("retorna True e mira a tabela revision_feedback",
          ok and _captura.get("table") == "revision_feedback")
    rec = _captura.get("record", {})
    check("record com contagens e agregados",
          rec.get("job_id") == "job-teste-123"
          and rec.get("n_alterados") == 3 and rec.get("n_removidos") == 1
          and rec.get("n_adicionados") == 1
          and rec.get("mediana_abs_delta_pct") == 14.3
          and isinstance(rec.get("por_disciplina"), dict)
          and isinstance(rec.get("por_confidence"), dict), rec.keys())
    # cap de detalhe: planilha gigante não entope o jsonb
    res_gigante = {"totais": {}, "itens": [{"acao": "mantido"}] * 999}
    rf.salvar_feedback("job-gigante", res_gigante)
    check("itens capados em MAX_ITENS_DETALHE",
          len(_captura["record"]["itens"]) == rf.MAX_ITENS_DETALHE)
    check("resultado vazio → False sem explodir",
          rf.salvar_feedback("job-x", None) is False)
finally:
    rf._insert_supabase = _original_insert


# ═══════════════════════════════════════════════
#  6. resumo_para_admin
# ═══════════════════════════════════════════════

print("== resumo_para_admin ==")
rows = [
    {"job_id": "job-a", "created_at": "2026-07-19T12:00:00Z",
     "arquivo": "rev_a.xlsx", "n_originais": 8, "n_revisados": 8,
     "n_alterados": 3, "n_removidos": 1, "n_adicionados": 1,
     "itens": res["itens"]},
    # row legada sem detalhe: entra só nas contagens globais
    {"job_id": "job-b", "created_at": "2026-07-18T12:00:00Z",
     "arquivo": "rev_b.xlsx", "n_originais": 10, "n_revisados": 10,
     "n_alterados": 2, "n_removidos": 0, "n_adicionados": 0, "itens": []},
]
resumo = rf.resumo_para_admin(rows)
check("n_planilhas = 2", resumo["n_planilhas"] == 2)
check("totais somam detalhe + legado (18 itens, 5 alterados)",
      resumo["totais"]["n_itens"] == 18
      and resumo["totais"]["n_alterados"] == 5, resumo["totais"])
check("por_disciplina é lista ordenada por % alterados",
      isinstance(resumo["por_disciplina"], list)
      and len(resumo["por_disciplina"]) >= 4
      and resumo["por_disciplina"][0]["pct_alterados"]
      >= (resumo["por_disciplina"][-1]["pct_alterados"] or 0))
_rc = resumo["por_confidence"]
check("recorte confidence no resumo (estimado 50% × confirmado 25%)",
      _rc["estimado"]["pct_alterados"] == 50.0
      and _rc["confirmado"]["pct_alterados"] == 25.0, _rc)
check("ultimas com no máx. 10 e campos do painel",
      len(resumo["ultimas"]) == 2
      and resumo["ultimas"][0]["job_id"] == "job-a")
check("resumo de lista vazia → estado vazio sem explodir",
      rf.resumo_para_admin([])["n_planilhas"] == 0
      and rf.resumo_para_admin(None)["n_planilhas"] == 0)


# ═══════════════════════════════════════════════

print(f"\n{'=' * 46}\n{_passed} ok, {_failed} falhas")
sys.exit(1 if _failed else 0)
