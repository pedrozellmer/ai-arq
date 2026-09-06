# -*- coding: utf-8 -*-
"""Financeiro da obra — exportação .xlsx e PDF (05/09/2026).

O arquivo sai do site e chega em quem nunca viu a tela: banco, cliente final,
orçamentista. Por isso cada guarda aqui cobra um FATO que o arquivo tem que
carregar sozinho:
  • os números são os DA TELA — `montar_dados_export` repete os predicados de
    financeiro.html (pago / emAberto / vencido / 30 dias / Contratado) um a um,
    conferidos aqui contra uma conta feita à mão;
  • nº5 — valor ausente é célula VAZIA (nunca 0) e "—" no PDF; a ressalva "o
    AI.arq não precifica" está DENTRO do .xlsx e do PDF; as somas dizem que são
    só do que tem valor;
  • nº7 — o vencimento por fase vira data pela fase do cronograma (início/fim),
    com a categoria de reserva; sem cronograma sai "fase sem data", não uma data
    inventada;
  • rotas: async com tudo pesado no threadpool; 0 lançamentos → 404 com texto
    em português (não um arquivo vazio); admin monta cronograma/branding SEM o
    JWT dele (a RLS do cronograma é do dono); banco fora → 502, não 404.
🧪 Controles: a régua de 30 dias INCLUI os vencidos (como na tela); o total geral
é a soma dos SUBTOTAIS (somar a coluna inteira contaria os subtotais de novo).
"""
import os
import sys
import types
from datetime import date

import pytest

_AQUI = os.path.dirname(os.path.abspath(__file__))
_BACKEND = os.path.dirname(_AQUI)
sys.path.insert(0, _BACKEND)
sys.path.insert(0, _AQUI)

import main  # noqa: E402
import financeiro_export as fe  # noqa: E402
from _corpo import corpo_de, fonte, sem_comentarios  # noqa: E402

HOJE = date(2026, 9, 5)
JOB = "job-export-1"
REQ = types.SimpleNamespace(headers={"Authorization": "Bearer jwt-do-cliente"})

FASES = [
    {"label": "Pisos", "inicio": "2026-09-10", "fim": "2026-10-05", "ordem": 2},
    {"label": "Complementares", "inicio": "2026-11-01", "fim": "2026-12-01", "ordem": 3},
    {"label": "Demolição", "inicio": "2026-08-01", "fim": "2026-08-20", "ordem": 1},
    {"label": "Sem datas", "inicio": None, "fim": None},
]


def _l(**k):
    base = {"id": "x", "job_id": JOB, "escopo": "obra", "origem": "quantitativo", "categoria": "Pisos",
            "descricao": "Porcelanato 60x60", "fornecedor": "Cerâmica Boa", "valor": 1000, "forma_pagamento": "3×",
            "venc_tipo": "fase", "venc_fase": "Pisos", "venc_quando": "inicio", "venc_data": None,
            "status": "contratado", "pago_em": None, "origem_quantidade": 1062.0, "origem_unidade": "m2"}
    base.update(k)
    return base


ROWS = [
    _l(descricao="Porcelanato 60x60", valor=1000, status="contratado"),                       # vence 10/09 (5 d) → até 30
    _l(descricao="Rejunte", valor=200, status="aprovado", venc_quando="fim"),                   # vence 05/10 (30 d) → até 30
    _l(descricao="Luminária", categoria="Complementares", venc_fase="Complementares",
       valor=500, status="pago", pago_em="2026-09-01"),                                         # pago
    _l(descricao="Pintura", categoria="Acabamentos", venc_fase="Acabamentos",
       valor=300, status="contratado"),                                                        # fase sem data → sem data
    _l(descricao="Projeto elétrico", categoria="Complementares", venc_fase="Complementares",
       valor=None, status="cotado", origem="livre"),                                            # sem valor
    _l(descricao="Entulho", categoria="Demolição", venc_fase="Demolição", venc_quando="fim",
       valor=150, status="aprovado"),                                                           # vence 20/08 → VENCIDO
    _l(descricao="Marcenaria", categoria="Acabamentos", venc_tipo="data", venc_data="2026-09-30",
       valor=800, status="enviado"),                                                            # aguardando o cliente
]


# ══════════════════════════════════════════════════════════════════════════
#  os números são os da tela
# ══════════════════════════════════════════════════════════════════════════
def test_kpis_batem_com_a_conta_feita_a_mao():
    d = fe.montar_dados_export(ROWS, FASES, HOJE)
    k = d["kpis"]
    assert k["contratado"] == 1800.0 and k["contratado_n"] == 3, "contratado+pago: 1000+500+300 (Pintura conta)"
    assert k["pago"] == 500.0 and k["pago_n"] == 1 and k["pago_pct"] == 28
    # em aberto = não pago, com valor, contratado/aprovado: Porcelanato, Rejunte, Pintura, Entulho
    assert k["a_pagar_30"] == 1350.0 and k["a_pagar_30_n"] == 3, "10/09, 05/10 e o vencido 20/08; Pintura não tem data"
    assert k["vencidos_n"] == 1 and k["vencidos"] == 150.0
    assert k["sem_data_n"] == 1, "Pintura: em aberto sem data"
    assert k["aguardando"] == 800.0 and k["aguardando_n"] == 1
    assert k["sem_valor_n"] == 1 and k["total_com_valor"] == 2950.0 and k["n"] == 7


def test_CONTROLE_a_regua_de_30_dias_INCLUI_os_vencidos_como_na_tela():
    d = fe.montar_dados_export([_l(valor=10, status="contratado", venc_tipo="data", venc_data="2026-01-01")], [], HOJE)
    assert d["kpis"]["a_pagar_30_n"] == 1 and d["kpis"]["vencidos_n"] == 1


def test_valor_ausente_fica_None_nunca_zero_e_nao_entra_em_soma():
    d = fe.montar_dados_export([_l(valor=None, status="contratado")], FASES, HOJE)
    l = d["linhas"][0]
    assert l["valor"] is None and l["em_aberto"] is False, "sem valor não está 'em aberto' (tela: temValor)"
    assert d["kpis"]["contratado"] == 0.0 and d["kpis"]["contratado_n"] == 1


def test_vencimento_por_fase_vira_data_pela_fase_com_categoria_de_reserva():
    fases, ordem = fe.fases_do_cronograma(FASES)
    assert ordem == ["Demolição", "Pisos", "Complementares"], "por `ordem` como a tela; fase sem as DUAS datas não entra"
    sem_ordem = fe.fases_do_cronograma([{"label": "B", "inicio": "2026-02-01", "fim": "2026-03-01"},
                                        {"label": "A", "inicio": "2026-01-01", "fim": "2026-02-01"}])[1]
    assert sem_ordem == ["A", "B"], "sem `ordem`, pela data de início"
    assert fe.vencimento(_l(venc_quando="inicio"), fases) == (date(2026, 9, 10), "início da fase Pisos")
    assert fe.vencimento(_l(venc_quando="fim"), fases) == (date(2026, 10, 5), "fim da fase Pisos")
    assert fe.vencimento(_l(venc_fase="", categoria="Pisos"), fases)[0] == date(2026, 9, 10), "categoria de reserva"
    assert fe.vencimento(_l(venc_fase="Acabamentos", categoria="Acabamentos"), fases) == (None, "fase sem data")
    # fase pedida sem data mas categoria com data: a tela cai na categoria — e a regra nomeia a fase que DATOU
    assert fe.vencimento(_l(venc_fase="Acabamentos", categoria="Pisos"), fases) == (date(2026, 9, 10), "início da fase Pisos")
    assert fe.vencimento(_l(venc_tipo="data", venc_data="2026-09-30"), fases) == (date(2026, 9, 30), "data fixa")
    assert fe.vencimento(_l(venc_tipo="data", venc_data="lixo"), fases) == (None, "data fixa")


def test_grupos_na_ordem_do_cronograma_e_o_resto_depois():
    d = fe.montar_dados_export(ROWS, FASES, HOJE)
    # fases_custom cru não vem ordenado: a tela ordena por `ordem`, depois `inicio` — Demolição tem ordem 1
    assert [g["categoria"] for g in d["grupos"]] == ["Demolição", "Pisos", "Complementares", "Acabamentos"]
    comp = [g for g in d["grupos"] if g["categoria"] == "Complementares"][0]
    assert comp["n"] == 2 and comp["n_sem_valor"] == 1 and comp["total"] == 500.0
    assert d["tem_cronograma"] is True
    assert fe.montar_dados_export(ROWS, [], HOJE)["tem_cronograma"] is False


def test_pct_pago_arredonda_meio_pra_cima_como_o_Math_round_da_tela():
    """Revisão 05/09: round() do Python é 'banqueiro' (28,5 → 28); a tela usa Math.round (→ 29)."""
    d = fe.montar_dados_export([_l(valor=100, status="contratado"), _l(valor=100, status="contratado"),
                                _l(valor=57, status="pago", venc_tipo="data", venc_data="2026-09-01")], [], HOJE)
    assert d["kpis"]["contratado"] == 257.0 and d["kpis"]["pago_pct"] == 22
    d2 = fe.montar_dados_export([_l(valor=143, status="contratado"), _l(valor=57, status="pago")], [], HOJE)
    assert d2["kpis"]["pago_pct"] == 29, "57/200 = 28,5 → 29 (não 28)"
    d3 = fe.montar_dados_export([_l(valor=875, status="contratado"), _l(valor=125, status="pago")], [], HOJE)
    assert d3["kpis"]["pago_pct"] == 13, "125/1000 = 12,5 → 13"


def test_grupo_em_que_ninguem_tem_valor_nao_vira_R0(tmp_path):
    """Revisão 05/09 (nº5): categoria só com linhas sem valor saía 'R$ 0,00' — a tela mostra '—'."""
    rows = [_l(descricao="Projeto elétrico", categoria="Complementares", venc_fase="Complementares", valor=None, status="cotado"),
            _l(descricao="Projeto hidráulico", categoria="Complementares", venc_fase="Complementares", valor=None, status="cotado"),
            _l(descricao="Porcelanato", valor=1000, status="contratado")]
    d = fe.montar_dados_export(rows, FASES, HOJE)
    comp = [g for g in d["grupos"] if g["categoria"] == "Complementares"][0]
    assert comp["total"] is None and comp["todos_sem_valor"] is True
    html = fe.montar_html_financeiro(d, {})
    assert "Complementares <small>· 2 lançamentos · sem valor informado</small></td><td class=\"num mudo\">—<" in html
    from openpyxl import load_workbook
    p = str(tmp_path / "g.xlsx")
    fe.gerar_financeiro_xlsx(d, p, {})
    linhas = [[c.value for c in r] for r in load_workbook(p)["Financeiro"].iter_rows()]
    grp = [r for r in linhas if isinstance(r[0], str) and r[0].startswith("Complementares  ·")][0]
    assert grp[fe._COL_VALOR - 1] is None, "subtotal VAZIO, não 0 nem fórmula"
    total = [r for r in linhas if isinstance(r[0], str) and r[0].startswith("TOTAL DOS LANÇAMENTOS")][0]
    assert str(total[fe._COL_VALOR - 1]).count("+") == 0 and str(total[fe._COL_VALOR - 1]).startswith("="), (
        "total = só o subtotal de Pisos; o grupo sem valor fica fora da cadeia")
    # e quando NINGUÉM tem valor, nem o total geral existe
    d0 = fe.montar_dados_export(rows[:2], FASES, HOJE)
    assert d0["kpis"]["total_com_valor"] is None
    fe.gerar_financeiro_xlsx(d0, p, {})
    linhas0 = [[c.value for c in r] for r in load_workbook(p)["Financeiro"].iter_rows()]
    total0 = [r for r in linhas0 if isinstance(r[0], str) and r[0].startswith("TOTAL DOS LANÇAMENTOS")][0]
    assert total0[fe._COL_VALOR - 1] is None
    assert 'class="num mudo">—<' in fe.montar_html_financeiro(d0, {})


def test_rodape_do_pdf_escapa_string_css_nao_html():
    """Revisão 05/09: '&' no nome do projeto saía '&amp;' literal no rodapé de todas as páginas."""
    d = fe.montar_dados_export(ROWS, FASES, HOJE)
    html = fe.montar_html_financeiro(d, {"project_name": 'Casa & "Jardim" \\ 2'})
    i = html.find("@bottom-center")
    rodape = html[i:html.find("}", i)]
    assert 'Casa & \\"Jardim\\" \\\\ 2' in rodape, "string CSS: aspas e barra escapadas, & intacto"
    assert "&amp;" not in rodape
    assert 'Casa &amp; &quot;Jardim&quot; \\ 2' in html, "no corpo HTML o escape é o de HTML"


# ══════════════════════════════════════════════════════════════════════════
#  .xlsx
# ══════════════════════════════════════════════════════════════════════════
def _xlsx(tmp_path, rows=ROWS, fases=FASES, branding=None):
    from openpyxl import load_workbook
    d = fe.montar_dados_export(rows, fases, HOJE)
    p = str(tmp_path / "fin.xlsx")
    fe.gerar_financeiro_xlsx(d, p, branding or {"project_name": "Casa Teste", "architect_name": "Studio X"})
    return load_workbook(p)["Financeiro"]


def _linhas(ws):
    return [[c.value for c in r] for r in ws.iter_rows()]


def test_xlsx_tem_titulo_ressalva_e_os_4_numeros():
    ws = _xlsx(pytest.importorskip("pathlib").Path(os.environ.get("TMP") or os.environ.get("TEMP") or "/tmp"))
    texto = "\n".join(str(c) for r in _linhas(ws) for c in r if c is not None)
    assert "FINANCEIRO DA OBRA" in texto and "Casa Teste" in texto and "Studio X" in texto
    assert "O AI.arq não precifica obra" in texto, "a ressalva viaja DENTRO do arquivo"
    assert "1 lançamento sem valor informado" in texto
    for rot in ("Contratado", "Pago", "A pagar até 30 dias", "Aguardando o cliente"):
        assert rot in texto


def test_xlsx_valor_ausente_e_celula_VAZIA_e_subtotal_e_total_sao_formulas(tmp_path):
    ws = _xlsx(tmp_path)
    linhas = _linhas(ws)
    col_v = fe._COL_VALOR - 1
    sem_valor = [r for r in linhas if r[2] == "Projeto elétrico"]
    assert len(sem_valor) == 1 and sem_valor[0][col_v] is None, "nº5: sem valor é VAZIO, não 0"
    com_valor = [r for r in linhas if r[2] == "Porcelanato 60x60"][0]
    # openpyxl devolve a data como datetime — compara o dia, não o tipo
    assert com_valor[col_v] == 1000.0 and str(com_valor[6])[:10] == "2026-09-10" and com_valor[7] == "início da fase Pisos"
    subtotais = [r[col_v] for r in linhas if isinstance(r[0], str) and r[0].startswith("Pisos  ·")]
    assert subtotais and str(subtotais[0]).startswith("=SUM("), "subtotal por grupo é fórmula VIVA"
    total = [r for r in linhas if isinstance(r[0], str) and r[0].startswith("TOTAL DOS LANÇAMENTOS")][0]
    assert str(total[col_v]).startswith("=") and "+" in str(total[col_v]) and "SUM" not in str(total[col_v]), (
        "total geral = soma dos SUBTOTAIS (a coluna inteira contaria os subtotais de novo)")
    assert ws.freeze_panes is not None and int(str(ws.freeze_panes)[1:]) <= 10, (
        "revisão 05/09: o bloco congelado engolia a tela num notebook (16 linhas); agora ≤ 9 de topo")


def test_xlsx_sem_cronograma_avisa_e_nao_inventa_data(tmp_path):
    ws = _xlsx(tmp_path, fases=[])
    texto = "\n".join(str(c) for r in _linhas(ws) for c in r if c is not None)
    assert "ainda não tem cronograma gerado" in texto
    assert all(r[7] in (None, "fase sem data", "data fixa", "REGRA DO VENCIMENTO") for r in _linhas(ws) if r[7])


# ══════════════════════════════════════════════════════════════════════════
#  PDF (HTML testado sem WeasyPrint)
# ══════════════════════════════════════════════════════════════════════════
def test_html_do_pdf_carrega_ressalva_kpis_grupos_e_traco_no_sem_valor():
    d = fe.montar_dados_export(ROWS, FASES, HOJE)
    html = fe.montar_html_financeiro(d, {"project_name": "Casa <Teste>", "brand_color": "#abc"})
    assert "Casa &lt;Teste&gt;" in html and "#aabbcc" in html, "escapa HTML; cor da marca normalizada"
    assert "não precifica obra" in html
    for rot in ("Contratado", "Pago", "A pagar até 30 dias", "Aguardando o cliente"):
        assert rot in html
    assert "R$ 1.000,00" in html and "R$ 2.950,00" in html
    assert 'class="num mudo">—<' in html, "sem valor é traço, não R$ 0,00"
    assert "fase sem data" in html and "início da fase Pisos" in html and "data fixa" in html
    assert "vencido" in html and "Aguardando o cliente" in html
    assert html.count('class="grp"') == 4


def test_cor_da_marca_invalida_cai_no_indigo_da_casa():
    assert fe._accent({"brand_color": "azul"}) == "#4F46E5" and fe._accent({}) == "#4F46E5"
    assert fe._accent({"brand_color": "#1D4ED8"}) == "#1d4ed8"


def test_brl_formata_pt_br():
    assert fe._brl(1234567.5) == "R$ 1.234.567,50" and fe._brl(0) == "R$ 0,00" and fe._brl(None) == "—"


# ══════════════════════════════════════════════════════════════════════════
#  as rotas
# ══════════════════════════════════════════════════════════════════════════
class _Banco:
    def __init__(self, rows):
        self.rows = rows
        self.chamadas = []

    def __call__(self, request, method, path, body=None, params=None, prefer=None, timeout=15):
        self.chamadas.append({"m": method, "path": path})
        if method == "GET" and f"/{main._FIN_TABELA}?" in path:
            return self.rows
        return (200, [])


@pytest.fixture
def casa(monkeypatch):
    monkeypatch.setattr(main, "_require_project_owner", lambda request, job_id: "uid-dono")
    monkeypatch.setattr(main, "_fin_eh_admin", lambda request, owner: False)
    monkeypatch.setattr(main, "_log_error", lambda *a, **k: None)
    monkeypatch.setattr(main, "_fin_cronograma_salvo", lambda req, job_id: (200, None))
    chamadas = {}
    monkeypatch.setattr(main, "_get_branding_context",
                        lambda job_id, request=None: chamadas.setdefault("branding", []).append(request) or
                        {"project_name": "Casa Teste", "brand_color": "#1D4ED8"})
    return chamadas


def _banco(monkeypatch, resp):
    b = _Banco(resp)
    monkeypatch.setattr(main, "_supa_rest_as_user", b)
    return b


def test_xlsx_route_devolve_arquivo_com_nome_do_projeto(monkeypatch, casa):
    _banco(monkeypatch, (200, ROWS))
    r = main.financeiro_export_xlsx(JOB, REQ)
    assert r.media_type == main._FIN_XLSX_MIME and r.filename == "financeiro_obra_Casa_Teste.xlsx"
    from openpyxl import load_workbook
    ws = load_workbook(r.path)["Financeiro"]
    assert ws["A1"].value.startswith("FINANCEIRO DA OBRA")
    assert casa["branding"] == [REQ], "dono: branding lido com o JWT dele"


def test_pdf_route_devolve_pdf_e_o_render_roda_com_os_dados_montados(monkeypatch, casa):
    _banco(monkeypatch, (200, ROWS))
    visto = {}
    monkeypatch.setattr(fe, "render_financeiro_pdf_bytes",
                        lambda dados, branding=None: visto.update(n=dados["kpis"]["n"]) or b"%PDF-1.4 fake")
    r = main.financeiro_export_pdf(JOB, REQ)
    assert r.media_type == "application/pdf" and r.filename == "financeiro_obra_Casa_Teste.pdf"
    assert visto["n"] == 7 and open(r.path, "rb").read().startswith(b"%PDF")


def test_sem_lancamento_e_404_em_portugues_e_nao_arquivo_vazio(monkeypatch, casa):
    _banco(monkeypatch, (200, []))
    with pytest.raises(main.HTTPException) as ex:
        main.financeiro_export_xlsx(JOB, REQ)
    assert ex.value.status_code == 404 and "adicione o primeiro" in ex.value.detail
    # admin não pode adicionar — não recebe uma instrução impossível
    monkeypatch.setattr(main, "_fin_eh_admin", lambda request, owner: True)
    monkeypatch.setattr(main, "_fin_get", lambda request, path, eh_admin, timeout=10: (200, []))
    with pytest.raises(main.HTTPException) as ex2:
        main.financeiro_export_xlsx(JOB, REQ)
    assert ex2.value.status_code == 404 and "adicione" not in ex2.value.detail and "ainda não tem lançamentos" in ex2.value.detail


def test_banco_fora_e_502_nao_404(monkeypatch, casa):
    _banco(monkeypatch, (500, None))
    with pytest.raises(main.HTTPException) as ex:
        main.financeiro_export_pdf(JOB, REQ)
    assert ex.value.status_code == 502


def test_admin_monta_cronograma_e_branding_SEM_o_jwt_dele(monkeypatch, casa):
    """A RLS do cronograma é do dono: com o JWT do admin o export sairia 'fase sem data' — mentira."""
    _banco(monkeypatch, (200, ROWS))
    monkeypatch.setattr(main, "_fin_eh_admin", lambda request, owner: True)
    monkeypatch.setattr(main, "_fin_get", lambda request, path, eh_admin, timeout=10: (200, ROWS))
    vistos = []
    monkeypatch.setattr(main, "_fin_cronograma_salvo",
                        lambda req, job_id: vistos.append(req) or (200, {"fases_custom": FASES}))
    monkeypatch.setattr(main, "_build_cronograma_for_export",
                        lambda job_id, request=None: (_ for _ in ()).throw(AssertionError("não recalcular com fases_custom")))
    dados, branding = main._fin_montar_export(REQ, JOB, True)
    assert vistos == [None] and casa["branding"] == [None]
    assert dados["tem_cronograma"] is True and branding["project_name"] == "Casa Teste"


def test_leitura_do_cronograma_salvo_distingue_nao_existe_de_nao_li(monkeypatch):
    """`_supabase_get_cronograma` engole erro em None — aqui o STATUS volta junto."""
    chamadas = []
    monkeypatch.setattr(main, "_supa_rest_as_user",
                        lambda request, method, path, **k: chamadas.append(("user", path)) or (200, [{"id": 1}]))
    monkeypatch.setattr(main, "_supa_rest_service",
                        lambda method, path, **k: chamadas.append(("service", path)) or (500, None))
    assert main._fin_cronograma_salvo(REQ, JOB) == (200, {"id": 1})
    assert main._fin_cronograma_salvo(None, JOB) == (500, None), "admin lê pelo service_role; falha volta como status"
    assert all(f"job_id=eq.{JOB}" in p for _, p in chamadas) and [o for o, _ in chamadas] == ["user", "service"]
    monkeypatch.setattr(main, "_supa_rest_as_user", lambda request, method, path, **k: (200, []))
    assert main._fin_cronograma_salvo(REQ, JOB) == (200, None), "lista vazia = não existe (200, None)"


def test_cronograma_ilegivel_e_502_nao_um_arquivo_que_afirma_sem_cronograma(monkeypatch, casa):
    """Revisão 05/09 (alta): timeout no /cronogramas virava 'este projeto ainda não tem cronograma'
    dentro de um documento que vai pro banco — mentira com HTTP 200."""
    _banco(monkeypatch, (200, ROWS))
    monkeypatch.setattr(main, "_fin_cronograma_salvo", lambda req, job_id: (500, None))
    with pytest.raises(main.HTTPException) as ex:
        main._fin_montar_export(REQ, JOB, False)
    assert ex.value.status_code == 502 and "cronograma" in ex.value.detail


def test_fases_vem_do_fases_custom_CRU_como_a_tela_e_so_sem_ele_do_recalculado(monkeypatch, casa):
    """Revisão 05/09 (alta): a tela lê saved.fases_custom cru; o recalculado aplica cascade de
    dependência e reescreve fim de fase de um dia — moveria um vencimento que a tela não move."""
    _banco(monkeypatch, (200, ROWS))
    cru = [{"label": "Pisos", "inicio": "2026-10-01", "fim": "2026-10-20", "ordem": 1}]
    recalc = [{"label": "Pisos", "inicio": "2026-10-16", "fim": "2026-11-04", "ordem": 1}]
    monkeypatch.setattr(main, "_fin_cronograma_salvo", lambda req, job_id: (200, {"fases_custom": cru}))
    monkeypatch.setattr(main, "_build_cronograma_for_export",
                        lambda job_id, request=None: (_ for _ in ()).throw(AssertionError("não recalcular com fases_custom")))
    dados, _ = main._fin_montar_export(REQ, JOB, False)
    porc = [l for l in dados["linhas"] if l["descricao"] == "Porcelanato 60x60"][0]
    assert porc["venc"] == date(2026, 10, 1), "a data da tela, não a recalculada"
    # salvo SEM fases_custom (só config): o /full da tela → recalculado; e o branding que ele já
    # devolve é REAPROVEITADO (não se lê projects/profiles/logo duas vezes por clique)
    monkeypatch.setattr(main, "_fin_cronograma_salvo", lambda req, job_id: (200, {"data_inicio": "2026-10-01"}))
    monkeypatch.setattr(main, "_build_cronograma_for_export",
                        lambda job_id, request=None: ({"fases": recalc}, {"project_name": "Casa Teste"}))
    dados2, branding2 = main._fin_montar_export(REQ, JOB, False)
    porc2 = [l for l in dados2["linhas"] if l["descricao"] == "Porcelanato 60x60"][0]
    assert porc2["venc"] == date(2026, 10, 16) and branding2["project_name"] == "Casa Teste"
    assert casa["branding"] == [REQ], "a única leitura de branding foi a da 1ª parte — aqui veio do cronograma"


def test_cronograma_so_config_e_projeto_sem_itens_fica_sem_datas_como_a_tela(monkeypatch, casa):
    """/full devolve 404 'Projeto sem itens' → a tela fica sem datas; o arquivo idem. Outro erro → 502."""
    _banco(monkeypatch, (200, ROWS))
    monkeypatch.setattr(main, "_fin_cronograma_salvo", lambda req, job_id: (200, {"data_inicio": "2026-10-01"}))

    def _404(job_id, request=None):
        raise main.HTTPException(404, "Projeto sem itens")
    monkeypatch.setattr(main, "_build_cronograma_for_export", _404)
    dados, _ = main._fin_montar_export(REQ, JOB, False)
    assert dados["tem_cronograma"] is False and dados["kpis"]["n"] == 7

    def _500(job_id, request=None):
        raise main.HTTPException(500, "Erro ao buscar items")
    monkeypatch.setattr(main, "_build_cronograma_for_export", _500)
    with pytest.raises(main.HTTPException) as ex:
        main._fin_montar_export(REQ, JOB, False)
    assert ex.value.status_code == 502

    def _boom(job_id, request=None):
        raise RuntimeError("x")
    monkeypatch.setattr(main, "_build_cronograma_for_export", _boom)
    with pytest.raises(main.HTTPException) as ex2:
        main._fin_montar_export(REQ, JOB, False)
    assert ex2.value.status_code == 502


def test_branding_sem_nome_do_projeto_e_502_nao_Projeto_sem_nome(monkeypatch, casa):
    """Revisão 05/09: `_get_branding_context` engole timeout e devolve '' — o PDF sairia
    'Projeto sem nome' pro cliente, com HTTP 200 e sem rastro."""
    _banco(monkeypatch, (200, ROWS))
    monkeypatch.setattr(main, "_get_branding_context", lambda job_id, request=None: {"project_name": ""})
    with pytest.raises(main.HTTPException) as ex:
        main._fin_montar_export(REQ, JOB, False)
    assert ex.value.status_code == 502 and "dados do projeto" in ex.value.detail


def test_hoje_da_tela_e_aceito_so_perto_de_brasilia():
    from datetime import datetime, timedelta
    brasilia = (datetime.utcnow() - timedelta(hours=3)).date()
    ontem = brasilia - timedelta(days=1)
    assert main._fin_hoje(ontem.isoformat()) == ontem, "Manaus às 23h30 ainda está em ontem"
    assert main._fin_hoje((brasilia + timedelta(days=1)).isoformat()) == brasilia + timedelta(days=1)
    assert main._fin_hoje((brasilia - timedelta(days=30)).isoformat()) == brasilia, "longe demais: não é fuso, é erro"
    assert main._fin_hoje("") == brasilia and main._fin_hoje("lixo") == brasilia and main._fin_hoje(None) == brasilia


def test_forma_rotas_de_export_sao_SINCRONAS_e_passam_pelo_montador():
    """Revisão 05/09: `async def` faria 3 chamadas de rede (dono + JWT ×2) NO LAÇO antes do
    primeiro await — com --workers 1 o site inteiro para. `def` roda no threadpool do FastAPI."""
    src = sem_comentarios(fonte("main.py"))
    for rota in ("financeiro_export_xlsx", "financeiro_export_pdf"):
        assert f"\ndef {rota}(" in src and f"async def {rota}(" not in src
        c = corpo_de(rota)
        assert "_fin_montar_export(request, job_id, eh_admin, hoje)" in c
        assert "run_in_threadpool" not in c, "rota síncrona não despacha de novo"
