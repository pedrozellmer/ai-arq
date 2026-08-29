# -*- coding: utf-8 -*-
"""Modelo de QUADRO DE ÁREAS — planilha estrutural (29/08/2026).

Nasceu do estudo do blog de 29/08: o formato que traz gente é ARQUIVO PRA
BAIXAR (memorial, cronograma), e "quadro de áreas" é lacuna total nos 26 posts.

🚫 REGRA QUE GOVERNA O ARQUIVO: a planilha entrega a ESTRUTURA e as FÓRMULAS.
   Nenhum coeficiente de área equivalente vem preenchido — a NBR 12721:2006 é
   quem dá os coeficientes, e estampar um valor aqui seria afirmar norma de
   memória (o caminho do incidente de bibliografia inventada). As células de
   coeficiente ficam AMARELAS (preencher) com nota apontando pra norma.

🔑 A planilha carrega a lição do motor: TODA aba tem linha de conferência
   (soma dos pavimentos = total declarado). Quadro que não fecha consigo mesmo
   é a causa nº2 de recusa em prefeitura segundo o nosso próprio post do
   memorial CAU ("quadro de áreas inconsistente").

Rodar da raiz: python blog/downloads/gen_quadro_areas.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AQUI = os.path.dirname(os.path.abspath(__file__))
SAIDA = os.path.join(AQUI, "quadro-de-areas-modelo.xlsx")

# ── paleta do AI.arq (a mesma do cronograma) ────────────────────────────────
INDIGO = "4F46E5"
CINZA_CLARO = "F3F4F6"
AMARELO = "FEF3C7"          # célula que o usuário preenche
FINO = Side(style="thin", color="D1D5DB")
BORDA = Border(left=FINO, right=FINO, top=FINO, bottom=FINO)

F_TITULO = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
F_CAB = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
F_NOTA = Font(name="Calibri", italic=True, size=9, color="6B7280")
F_TOTAL = Font(name="Calibri", bold=True, size=11)


def _cabecalho(ws, titulo, colunas, larguras):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(colunas))
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = F_TITULO
    c.fill = PatternFill("solid", fgColor=INDIGO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for j, (nome, larg) in enumerate(zip(colunas, larguras), start=1):
        cel = ws.cell(row=3, column=j, value=nome)
        cel.font = F_CAB
        cel.fill = PatternFill("solid", fgColor="6366F1")
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = BORDA
        ws.column_dimensions[get_column_letter(j)].width = larg


def _linhas_vazias(ws, ini, n, cols, amarelas=()):
    for i in range(n):
        for j in range(1, cols + 1):
            cel = ws.cell(row=ini + i, column=j)
            cel.border = BORDA
            if j in amarelas:
                cel.fill = PatternFill("solid", fgColor=AMARELO)


wb = Workbook()

# ═══ ABA 1 — LEIA-ME ════════════════════════════════════════════════════════
ws = wb.active
ws.title = "LEIA-ME"
ws.column_dimensions["A"].width = 100
textos = [
    ("QUADRO DE ÁREAS — MODELO (ai.arq.br)", True),
    ("", False),
    ("Como usar:", True),
    ("1. Aba POR PAVIMENTO: uma linha por ambiente/uso, separando área coberta e descoberta.", False),
    ("2. Aba PRIVATIVA × COMUM: uma linha por unidade autônoma (apartamento, sala, loja).", False),
    ("3. Aba ÁREA EQUIVALENTE: só para incorporação imobiliária. As células AMARELAS de", False),
    ("   coeficiente ficam vazias de propósito: os coeficientes são definidos pela", False),
    ("   ABNT NBR 12721:2006 — leia a norma (ou o CUB do seu Sinduscon) antes de preencher.", False),
    ("   Esta planilha NÃO traz nenhum coeficiente embutido.", False),
    ("", False),
    ("Conferência automática:", True),
    ("Cada aba tem linha de TOTAL com soma automática e uma célula de CONFERÊNCIA para você", False),
    ("digitar o total declarado no projeto. Se as duas divergirem, o quadro não fecha —", False),
    ("e quadro de áreas inconsistente entre planta, memorial e quadro é motivo clássico de", False),
    ("recusa na prefeitura.", False),
    ("", False),
    ("Este modelo é estrutural e gratuito. Revise com profissional habilitado (CAU/CREA)", False),
    ("antes de protocolar. Fonte dos conceitos de área: ABNT NBR 12721:2006.", False),
]
for i, (t, negrito) in enumerate(textos, start=1):
    c = ws.cell(row=i, column=1, value=t)
    c.font = Font(name="Calibri", bold=negrito, size=11 if negrito else 10)

# ═══ ABA 2 — POR PAVIMENTO ══════════════════════════════════════════════════
ws = wb.create_sheet("POR PAVIMENTO")
cols = ["Pavimento", "Ambiente / uso", "Área coberta (m²)", "Área descoberta (m²)",
        "Total da linha (m²)"]
_cabecalho(ws, "QUADRO DE ÁREAS POR PAVIMENTO", cols, [16, 34, 16, 18, 16])
INI, N = 4, 30
_linhas_vazias(ws, INI, N, len(cols), amarelas=(1, 2, 3, 4))
for i in range(N):
    r = INI + i
    ws.cell(row=r, column=5, value=f"=SUM(C{r}:D{r})").border = BORDA
FIM = INI + N - 1
r = FIM + 1
ws.cell(row=r, column=2, value="TOTAL").font = F_TOTAL
for j, col in ((3, "C"), (4, "D"), (5, "E")):
    cel = ws.cell(row=r, column=j, value=f"=SUM({col}{INI}:{col}{FIM})")
    cel.font = F_TOTAL
    cel.border = BORDA
    cel.fill = PatternFill("solid", fgColor=CINZA_CLARO)
r += 2
ws.cell(row=r, column=2, value="Total declarado no projeto (digite):").font = F_NOTA
ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor=AMARELO)
ws.cell(row=r, column=3).border = BORDA
ws.cell(row=r + 1, column=2, value="Diferença (tem que dar zero):").font = F_NOTA
ws.cell(row=r + 1, column=3, value=f"=E{FIM + 1}-C{r}").font = F_TOTAL

# ═══ ABA 3 — PRIVATIVA × COMUM ══════════════════════════════════════════════
ws = wb.create_sheet("PRIVATIVA × COMUM")
cols = ["Unidade", "Privativa coberta (m²)", "Privativa descoberta (m²)",
        "Comum coberta (m²)", "Comum descoberta (m²)", "Total da unidade (m²)"]
_cabecalho(ws, "ÁREAS POR UNIDADE — REAL PRIVATIVA E DE USO COMUM", cols,
           [18, 20, 22, 18, 20, 18])
INI, N = 4, 24
_linhas_vazias(ws, INI, N, len(cols), amarelas=(1, 2, 3, 4, 5))
for i in range(N):
    r = INI + i
    ws.cell(row=r, column=6, value=f"=SUM(B{r}:E{r})").border = BORDA
FIM = INI + N - 1
r = FIM + 1
ws.cell(row=r, column=1, value="TOTAL").font = F_TOTAL
for j in range(2, 7):
    col = get_column_letter(j)
    cel = ws.cell(row=r, column=j, value=f"=SUM({col}{INI}:{col}{FIM})")
    cel.font = F_TOTAL
    cel.border = BORDA
    cel.fill = PatternFill("solid", fgColor=CINZA_CLARO)
ws.cell(row=r + 2, column=1,
        value="Os conceitos de área real, privativa e de uso comum são da ABNT NBR "
              "12721:2006 — use as definições da norma, não as de corretagem.").font = F_NOTA

# ═══ ABA 4 — ÁREA EQUIVALENTE (incorporação) ════════════════════════════════
ws = wb.create_sheet("ÁREA EQUIVALENTE")
cols = ["Descrição (parte da edificação)", "Área real (m²)",
        "Coeficiente (PREENCHER — ver NBR 12721)", "Área equivalente (m²)"]
_cabecalho(ws, "ÁREA EQUIVALENTE DE CONSTRUÇÃO — SÓ PARA INCORPORAÇÃO", cols,
           [40, 16, 34, 20])
INI, N = 4, 20
_linhas_vazias(ws, INI, N, len(cols), amarelas=(1, 2, 3))
for i in range(N):
    r = INI + i
    ws.cell(row=r, column=4, value=f"=IF(C{r}=\"\",\"\",B{r}*C{r})").border = BORDA
FIM = INI + N - 1
r = FIM + 1
ws.cell(row=r, column=1, value="TOTAL").font = F_TOTAL
for j, col in ((2, "B"), (4, "D")):
    cel = ws.cell(row=r, column=j, value=f"=SUM({col}{INI}:{col}{FIM})")
    cel.font = F_TOTAL
    cel.border = BORDA
    cel.fill = PatternFill("solid", fgColor=CINZA_CLARO)
ws.cell(row=r + 2, column=1,
        value="⚠ Os coeficientes de equivalência NÃO vêm preenchidos: eles são "
              "definidos pela ABNT NBR 12721:2006. Preencha a coluna C lendo a "
              "norma ou a orientação do Sinduscon do seu estado (CUB).").font = F_NOTA

wb.save(SAIDA)
print("gerado:", SAIDA, os.path.getsize(SAIDA), "bytes")
