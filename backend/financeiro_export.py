# -*- coding: utf-8 -*-
"""Financeiro da obra — exportação em .xlsx e PDF (etapa 1, 05/09/2026).

A tela `financeiro.html` mostra os lançamentos por categoria, cruzados com o
cronograma (o vencimento "amarrado à fase" vira data pela fase do Gantt) e os
quatro números do topo: Contratado, Pago, A pagar até 30 dias, Aguardando o
cliente. Este módulo produz os MESMOS números e a MESMA tabela em arquivo, pra
o arquiteto mandar pro cliente ou pro banco sem depender da tela.

Regras que valem aqui, escritas DENTRO do arquivo (o arquivo sai do site e chega
em quem nunca viu o contexto):
  • nº5 — o AI.arq não precifica: todo valor foi digitado pelo arquiteto ou veio
    de uma cotação que ele mesmo subiu. Valor ausente fica VAZIO (nunca 0) e as
    somas dizem que são só do que tem valor;
  • nº7 — a tabela sai do mesmo quantitativo/cronograma da tela; nada é
    recalculado aqui com regra própria: `montar_dados_export` repete, em Python,
    exatamente os predicados da tela (`pago`, `emAberto`, `vencido`, 30 dias);
  • estilo — a planilha IMPORTA os estilos de `spreadsheet.py` (como o cronograma):
    cópia diverge no primeiro ajuste de paleta.

O PDF é um documento A4 retrato, no molde do memorial (WeasyPrint, CSS inline),
com a cor da marca do escritório. Import do weasyprint fica dentro da função.
"""
from __future__ import annotations

import base64
import math
import os
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Mesma paleta/tipografia do quantitativo — importada, nunca duplicada.
from spreadsheet import (
    F_TITLE, F_SUB, F_HDR, F_N, F_BOLD, F_TOT, F_NOTE,
    P_SUB, P_HDR, P_TOT, P_LT,
    AC, AL, AR, BD,
)

FMT_MOEDA = 'R$ #,##0.00'
FMT_DATA = 'DD/MM/YYYY'

STATUS_LABEL = {
    "cotado": "Cotado", "enviado": "Aguardando o cliente", "aprovado": "Aprovado",
    "contratado": "Contratado", "pago": "Pago",
}
ORIGEM_LABEL = {"quantitativo": "Quantitativo", "comparativo": "Comparativo", "livre": "Linha livre"}

AVISO_NAO_PRECIFICA = (
    "Os valores deste documento foram informados por você ou vieram das cotações que você "
    "mesmo subiu no Comparativo. O AI.arq não precifica obra: ele só organiza por categoria, "
    "cruza com o cronograma e mostra o que vence, o que está pago e o que o cliente ainda não "
    "aprovou. Confira antes de enviar a terceiros."
)


# ══════════════════════════════════════════════════════════════════════════
#  datas e predicados — os MESMOS da tela
# ══════════════════════════════════════════════════════════════════════════
def _data(s) -> Optional[date]:
    """'AAAA-MM-DD' (ou ISO mais longo) → date; qualquer coisa torta → None."""
    if isinstance(s, date) and not isinstance(s, datetime):
        return s
    if isinstance(s, datetime):
        return s.date()
    try:
        return date.fromisoformat(str(s or "")[:10])
    except Exception:
        return None


def _br(d: Optional[date]) -> str:
    return d.strftime("%d/%m/%Y") if d else ""


def fases_do_cronograma(fases_lista: Optional[List[Dict]]) -> Tuple[Dict[str, Tuple[date, date]], List[str]]:
    """{label: (inicio, fim)} + ordem dos labels — só fase com as DUAS datas
    (a tela faz o mesmo: `if (!lab || !f.inicio || !f.fim) return`)."""
    fases: Dict[str, Tuple[date, date]] = {}
    ordem: List[str] = []
    # mesma ordenação da tela: `ordem`, depois `inicio` (fases_custom cru não vem ordenado)
    def _ordem(f):
        try:
            return float(f.get("ordem") or 0)
        except Exception:
            return 0.0             # ordem torta não pode derrubar o export inteiro
    lista = sorted((f for f in (fases_lista or []) if isinstance(f, dict)),
                   key=lambda f: (_ordem(f), str(f.get("inicio") or "")))
    for f in lista:
        lab = str((f or {}).get("label") or "").strip()
        ini, fim = _data((f or {}).get("inicio")), _data((f or {}).get("fim"))
        if not lab or not ini or not fim or lab in fases:
            continue
        fases[lab] = (ini, fim)
        ordem.append(lab)
    return fases, ordem


def vencimento(l: Dict, fases: Dict[str, Tuple[date, date]]) -> Tuple[Optional[date], str]:
    """(data, regra). Regra em português, pra sair no arquivo do lado da data:
    'data fixa' · 'início da fase X' · 'fim da fase X' · 'fase sem data'.
    Fase por label, com a categoria de reserva — igual à tela (`vencData`)."""
    if str(l.get("venc_tipo") or "fase") == "data":
        return _data(l.get("venc_data")), "data fixa"
    # 🩸 auditoria 06/09: a categoria era reserva mesmo quando a fase FOI escolhida e sumiu do Gantt —
    # o arquivo herdava calado a data de outra fase. Reserva só quando não há fase informada.
    cat = str(l.get("categoria") or "").strip()
    lab = str(l.get("venc_fase") or "").strip() or cat
    f = fases.get(lab)
    if not f:
        return None, "fase sem data"
    quando = "fim" if str(l.get("venc_quando") or "inicio") == "fim" else "inicio"
    return (f[1] if quando == "fim" else f[0]), f"{'fim' if quando == 'fim' else 'início'} da fase {lab}"


def montar_dados_export(rows: List[Dict], fases_lista: Optional[List[Dict]], hoje: date) -> Dict:
    """Linhas enriquecidas + grupos por categoria + os 4 KPIs da tela.

    Predicados copiados da tela, um a um:
      pago      = status == 'pago'
      emAberto  = não pago E tem valor E status em (contratado, aprovado)
      vencido   = emAberto E tem data E data < hoje
      até 30 d  = emAberto E tem data E (data - hoje) <= 30 dias   (inclui os vencidos)
      Contratado (KPI) = status em (contratado, pago) — conta linhas, soma só quem tem valor
    """
    fases, ordem = fases_do_cronograma(fases_lista)
    linhas: List[Dict] = []
    for l in rows or []:
        valor = l.get("valor")
        try:
            valor = None if valor is None or valor == "" else round(float(valor), 2)
        except Exception:
            valor = None
        st = str(l.get("status") or "cotado")
        venc, regra = vencimento(l, fases)
        pago = st == "pago"
        em_aberto = (not pago) and valor is not None and st in ("contratado", "aprovado")
        vencido = em_aberto and venc is not None and venc < hoje
        ate_30 = em_aberto and venc is not None and (venc - hoje).days <= 30
        linhas.append({
            "categoria": str(l.get("categoria") or "").strip() or "Sem categoria",
            "descricao": str(l.get("descricao") or "").strip(),
            "origem": ORIGEM_LABEL.get(str(l.get("origem") or "livre"), "Linha livre"),
            "origem_quantidade": l.get("origem_quantidade"),
            "origem_unidade": str(l.get("origem_unidade") or "").strip(),
            "fornecedor": str(l.get("fornecedor") or "").strip(),
            "forma_pagamento": str(l.get("forma_pagamento") or "").strip(),
            "valor": valor,
            "status": st,
            "status_label": STATUS_LABEL.get(st, st),
            "venc": venc, "venc_br": _br(venc), "venc_regra": regra,
            "pago_em": _data(l.get("pago_em")) if pago else None,
            "pago": pago, "em_aberto": em_aberto, "vencido": vencido, "ate_30": ate_30,
        })

    # categorias: na ordem das fases do cronograma; as que não são fase, depois, na ordem em que aparecem
    vistas: List[str] = []
    for l in linhas:
        if l["categoria"] not in vistas:
            vistas.append(l["categoria"])
    cats = [c for c in ordem if c in vistas] + [c for c in vistas if c not in ordem]
    grupos = []
    for c in cats:
        ls = [l for l in linhas if l["categoria"] == c]
        com_valor = [l["valor"] for l in ls if l["valor"] is not None]
        # 🔒 nº5: grupo em que NINGUÉM tem valor não soma "R$ 0,00" — o total é AUSENTE (a tela mostra "—")
        grupos.append({"categoria": c, "linhas": ls, "n": len(ls),
                       "total": (round(sum(com_valor), 2) if com_valor else None),
                       "n_sem_valor": len(ls) - len(com_valor), "todos_sem_valor": not com_valor})

    def _soma(ls):
        return round(sum(l["valor"] for l in ls if l["valor"] is not None), 2)

    def _pct(parte, todo):
        # meio-pra-cima como o Math.round da tela (o round() do Python é "banqueiro": 28,5 → 28)
        return int(math.floor(100.0 * parte / todo + 0.5)) if todo else None

    contr = [l for l in linhas if l["status"] in ("contratado", "pago")]
    pagos = [l for l in linhas if l["pago"]]
    em30 = [l for l in linhas if l["ate_30"]]
    vencidos = [l for l in em30 if l["vencido"]]
    aguard = [l for l in linhas if l["status"] == "enviado"]
    sem_data = [l for l in linhas if l["em_aberto"] and l["venc"] is None]

    def _kpi(balde):
        """🔒 nº5, igual à tela: balde só com linha sem valor não vira R$ 0 — o número é AUSENTE."""
        return _soma(balde) if any(l["valor"] is not None for l in balde) else None
    vc, vp = _soma(contr), _soma(pagos)
    kpis = {
        "contratado": _kpi(contr), "contratado_n": len(contr), "contratado_sem_valor": sum(1 for l in contr if l["valor"] is None),
        "pago": _kpi(pagos), "pago_n": len(pagos),
        "pago_pct": _pct(vp, vc),
        "a_pagar_30": _kpi(em30), "a_pagar_30_n": len(em30),
        "vencidos_n": len(vencidos), "vencidos": _soma(vencidos),
        "sem_data_n": len(sem_data),
        "aguardando": _kpi(aguard), "aguardando_n": len(aguard),
        "aguardando_sem_valor": sum(1 for l in aguard if l["valor"] is None),
        "sem_valor_n": sum(1 for l in linhas if l["valor"] is None),
        # ausente quando NENHUMA linha tem valor (nº5: não existe "total R$ 0,00" de nada informado)
        "total_com_valor": (_soma(linhas) if any(l["valor"] is not None for l in linhas) else None),
        "n": len(linhas),
    }
    return {"linhas": linhas, "grupos": grupos, "kpis": kpis, "hoje": hoje,
            "tem_cronograma": bool(ordem), "fases_ordem": ordem}


# ══════════════════════════════════════════════════════════════════════════
#  .xlsx — mesmo padrão visual do quantitativo e do cronograma
# ══════════════════════════════════════════════════════════════════════════
COLS = ['Nº', 'CATEGORIA / FASE', 'ITEM', 'ORIGEM', 'FORNECEDOR', 'FORMA DE PAGAMENTO',
        'VENCIMENTO', 'REGRA DO VENCIMENTO', 'STATUS', 'PAGO EM', 'VALOR (R$)']
_COL_VALOR = len(COLS)          # 11
_LARGURAS = [6, 22, 48, 13, 24, 18, 13, 26, 20, 12, 16]


def _merge_texto(ws, ro, texto, font, fill=None, altura=None, n_cols=_COL_VALOR):
    ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=n_cols)
    c = ws.cell(row=ro, column=1, value=texto)
    c.font = font
    c.alignment = AL
    if fill:
        c.fill = fill
    if altura:
        ws.row_dimensions[ro].height = altura
    return ro + 1


def gerar_financeiro_xlsx(dados: Dict, output_path: str, branding: Optional[Dict] = None) -> str:
    """Escreve o .xlsx e devolve o caminho. `dados` = saída de montar_dados_export."""
    branding = branding or {}
    k = dados["kpis"]
    wb = Workbook()
    ws = wb.active
    ws.title = 'Financeiro'
    ws.sheet_properties.tabColor = '4F46E5'

    # 🪤 O bloco de cima fica CONGELADO junto com o cabeçalho da tabela — cada linha aqui é uma
    # linha a menos de lançamento visível num notebook. Por isso: identificação numa linha só,
    # e os 4 números da tela numa linha (mais a dos subtítulos). Cabeçalho da tabela ≤ linha 9.
    ro = 1
    ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=_COL_VALOR)
    c = ws.cell(row=ro, column=1, value='FINANCEIRO DA OBRA — LANÇAMENTOS E DESEMBOLSO')
    c.font = F_TITLE
    c.alignment = AL
    ro += 1
    ident = " · ".join(f"{rot}: {val}" for rot, val in (
        ('Projeto', branding.get('project_name') or 'Projeto sem nome'),
        ('Escritório', branding.get('architect_name') or branding.get('company') or ''),
        ('Cliente', branding.get('client_name') or ''),
        ('Emitido em', _br(dados.get("hoje"))),
    ) if val)
    ro = _merge_texto(ws, ro, ident, F_N)

    # a ressalva (nº5) DENTRO do arquivo, como no cronograma e no quantitativo
    ro = _merge_texto(ws, ro, AVISO_NAO_PRECIFICA, F_NOTE, altura=30)
    if k["sem_valor_n"]:
        ro = _merge_texto(ws, ro, (
            f"ATENÇÃO — {k['sem_valor_n']} lançamento{'s' if k['sem_valor_n'] != 1 else ''} "
            f"sem valor informado: as somas abaixo são só do que tem valor, NÃO é o custo da obra inteira."
        ), F_SUB, fill=P_SUB)
    if not dados.get("tem_cronograma"):
        ro = _merge_texto(ws, ro, (
            "Este projeto ainda não tem cronograma gerado: vencimento amarrado à fase sai como "
            "'fase sem data'. Gere o cronograma e exporte de novo pra ver as datas."
        ), F_NOTE)

    # ── os 4 números da tela, numa linha (rótulo | valor) e os subtítulos na linha de baixo
    pago_sub = (f"{k['pago_pct']}% do contratado" if k["pago_pct"] is not None else "—")
    a30_sub = (f"{k['vencidos_n']} já vencido{'s' if k['vencidos_n'] > 1 else ''} · "
               f"{k['a_pagar_30_n'] - k['vencidos_n']} a vencer" if k["vencidos_n"]
               else f"{k['a_pagar_30_n']} a vencer")
    if k["sem_data_n"]:
        a30_sub += f" · {k['sem_data_n']} sem data"
    _sufixo = lambda n: f" · {n} sem valor" if n else ""
    kpis = (
        ('Contratado', k["contratado"], f"{k['contratado_n']} lançamento{'s' if k['contratado_n'] != 1 else ''}"
                                        + _sufixo(k.get("contratado_sem_valor", 0))),
        ('Pago', k["pago"], pago_sub),
        ('A pagar até 30 dias', k["a_pagar_30"], a30_sub),
        ('Aguardando o cliente', k["aguardando"], f"{k['aguardando_n']} aguardando o cliente"
                                                  + _sufixo(k.get("aguardando_sem_valor", 0))),
    )
    slots = ((1, 2, 3), (4, 5, 6), (7, 8, 9), (10, 10, 11))     # (rótulo de, rótulo até, coluna do valor)
    linha_kpi, linha_sub = ro, ro + 1
    for (c1, c2, cv), (rot, val, sub) in zip(slots, kpis):
        if c2 > c1:
            ws.merge_cells(start_row=linha_kpi, start_column=c1, end_row=linha_kpi, end_column=c2)
        ws.merge_cells(start_row=linha_sub, start_column=c1, end_row=linha_sub, end_column=cv)
        c = ws.cell(row=linha_kpi, column=c1, value=rot)
        c.font = F_BOLD
        c.fill = P_SUB
        c.alignment = AL
        c.border = BD
        v = ws.cell(row=linha_kpi, column=cv, value=(val if val is not None else "—"))
        v.font = F_BOLD
        v.fill = P_SUB
        if val is not None:
            v.number_format = FMT_MOEDA      # 🔒 nº5: ausente é "—", nunca R$ 0,00
        v.alignment = AR
        v.border = BD
        s = ws.cell(row=linha_sub, column=c1, value=sub)
        s.font = F_NOTE
        s.alignment = AL
    ws.row_dimensions[linha_kpi].height = 30
    ro += 3

    # ── cabeçalho da tabela
    linha_hdr = ro
    for i, h in enumerate(COLS, start=1):
        c = ws.cell(row=ro, column=i, value=h)
        c.font = F_HDR
        c.fill = P_HDR
        c.alignment = AC
        c.border = BD
    ro += 1

    # ── grupos por categoria, subtotal por grupo (fórmula VIVA), total geral pelos subtotais
    n = 0
    subtotais: List[str] = []
    colv = get_column_letter(_COL_VALOR)
    for g in dados["grupos"]:
        ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=_COL_VALOR - 1)
        if g["todos_sem_valor"]:
            sufixo = "  ·  sem valor informado"
        elif g["n_sem_valor"]:
            sufixo = f"  ·  {g['n_sem_valor']} sem valor"
        else:
            sufixo = ""
        c = ws.cell(row=ro, column=1, value=f"{g['categoria']}  ·  {g['n']} lançamento{'s' if g['n'] != 1 else ''}{sufixo}")
        c.font = F_SUB
        c.fill = P_SUB
        c.alignment = AL
        c.border = BD
        prim = ro + 1
        ult = ro + len(g["linhas"])
        ct = ws.cell(row=ro, column=_COL_VALOR)
        ct.font = F_SUB
        ct.fill = P_SUB
        ct.alignment = AR
        ct.border = BD
        if not g["todos_sem_valor"]:
            # 🔒 nº5: grupo sem nenhum valor fica com o subtotal VAZIO (não "R$ 0,00") e fora do total
            ct.value = f"=SUM({colv}{prim}:{colv}{ult})"
            ct.number_format = FMT_MOEDA
            subtotais.append(f"{colv}{ro}")
        ro += 1
        for idx, l in enumerate(g["linhas"]):
            n += 1
            alt = P_LT if idx % 2 else None
            vals = [n, l["categoria"], l["descricao"], l["origem"], l["fornecedor"],
                    l["forma_pagamento"], l["venc"], l["venc_regra"], l["status_label"],
                    l["pago_em"], l["valor"]]
            for i, v in enumerate(vals, start=1):
                c = ws.cell(row=ro, column=i, value=v)
                c.font = F_N
                c.border = BD
                c.alignment = AL if i in (2, 3, 5, 6, 8, 9) else AC
                if alt:
                    c.fill = alt
                if i in (7, 10) and isinstance(v, date):
                    c.number_format = FMT_DATA
                if i == _COL_VALOR:
                    c.alignment = AR
                    if v is None:
                        c.value = None            # 🔒 nº5: sem valor é VAZIO, nunca 0
                    else:
                        c.number_format = FMT_MOEDA
            ro += 1

    # ── total geral (soma dos subtotais — nunca da coluna inteira, que teria os subtotais dentro)
    ws.merge_cells(start_row=ro, start_column=1, end_row=ro, end_column=_COL_VALOR - 1)
    c = ws.cell(row=ro, column=1, value='TOTAL DOS LANÇAMENTOS COM VALOR'
                + (f"  ({k['sem_valor_n']} sem valor não entram)" if k["sem_valor_n"] else ''))
    c.font = F_TOT
    c.fill = P_TOT
    c.alignment = AR
    c.border = BD
    ct = ws.cell(row=ro, column=_COL_VALOR)
    ct.font = F_TOT
    ct.fill = P_TOT
    ct.alignment = AR
    ct.border = BD
    if subtotais:
        ct.value = "=" + "+".join(subtotais)
        ct.number_format = FMT_MOEDA
    # sem nenhum subtotal (ninguém tem valor) a célula fica VAZIA — não "R$ 0,00"
    ro += 1

    for i, w in enumerate(_LARGURAS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # por COORDENADA: a 1ª linha abaixo do cabeçalho é o grupo (mesclada), e o openpyxl
    # não aceita MergedCell como âncora do painel
    ws.freeze_panes = f"C{linha_hdr + 1}"
    wb.save(output_path)
    return output_path


# ══════════════════════════════════════════════════════════════════════════
#  PDF — A4 retrato, molde do memorial (WeasyPrint, CSS inline), cor da marca
# ══════════════════════════════════════════════════════════════════════════
def _esc(s) -> str:
    return (str(s if s is not None else "").replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def _css_str(s) -> str:
    """Texto dentro de `content: "..."` no CSS: escape de STRING CSS, não de HTML —
    '&amp;' ali sairia literal no rodapé de todas as páginas ('Casa &amp; Jardim')."""
    return (str(s if s is not None else "").replace("\\", "\\\\").replace('"', '\\"')
            .replace("\n", " ").replace("\r", " "))


def _brl(v: Optional[float]) -> str:
    if v is None:
        return "—"
    s = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


def _brl0(v: Optional[float]) -> str:
    """Sem centavos — é como a tela mostra os 4 números do topo (BRL com maximumFractionDigits:0),
    e cabe na caixa do KPI do PDF a partir de R$ 1 milhão (auditoria 06/09)."""
    if v is None:
        return "—"
    return "R$ " + f"{int(round(v)):,}".replace(",", ".")


def _accent(branding: Dict) -> str:
    h = str((branding or {}).get("brand_color") or "").strip().lower()
    body = h[1:] if h.startswith("#") else ""
    if len(body) == 3 and all(ch in "0123456789abcdef" for ch in body):
        body = "".join(ch * 2 for ch in body)
    if len(body) == 6 and all(ch in "0123456789abcdef" for ch in body):
        return "#" + body
    return "#4F46E5"


def _logo_uri(branding: Dict) -> str:
    path = (branding or {}).get("logo_local_path")
    if not path or not isinstance(path, str) or not os.path.exists(path):
        return ""
    try:
        ext = os.path.splitext(path)[1].lower()
        mime = {".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                ".gif": "image/gif", ".webp": "image/webp", ".svg": "image/svg+xml"}.get(ext, "image/png")
        with open(path, "rb") as fh:
            return f"data:{mime};base64,{base64.b64encode(fh.read()).decode('ascii')}"
    except Exception:
        return ""


_ST_COR = {"pago": "#047857", "contratado": "#1D4ED8", "aprovado": "#6D28D9",
           "enviado": "#B45309", "cotado": "#475569"}


def montar_html_financeiro(dados: Dict, branding: Optional[Dict] = None) -> str:
    """HTML completo do PDF (testável sem WeasyPrint)."""
    branding = branding or {}
    k = dados["kpis"]
    acc = _accent(branding)
    logo = _logo_uri(branding)
    hoje_br = _br(dados.get("hoje"))

    cab = []
    if logo:
        cab.append(f'<img class="logo" src="{logo}" alt="">')
    cab.append(f'<div class="tit"><h1>Financeiro da obra</h1>'
               f'<div class="proj">{_esc(branding.get("project_name") or "Projeto sem nome")}</div></div>')
    meta = [x for x in (
        ("Escritório", branding.get("architect_name") or branding.get("company") or ""),
        ("Cliente", branding.get("client_name") or ""),
        ("Emitido em", hoje_br),
    ) if x[1]]
    meta_html = "".join(f'<div><span>{_esc(r)}</span>{_esc(v)}</div>' for r, v in meta)

    pago_sub = (f"{k['pago_pct']}% do contratado" if k["pago_pct"] is not None else "—")
    a30_sub = (f"{k['vencidos_n']} já vencido{'s' if k['vencidos_n'] > 1 else ''} · "
               f"{k['a_pagar_30_n'] - k['vencidos_n']} a vencer" if k["vencidos_n"]
               else f"{k['a_pagar_30_n']} a vencer")
    if k["sem_data_n"]:
        a30_sub += f" · {k['sem_data_n']} sem data"
    _suf = lambda n: f" · {n} sem valor" if n else ""
    kpis = "".join(
        f'<div class="kpi {cls}"><div class="k">{_esc(rot)}</div><div class="v">{_esc(_brl0(val))}</div>'
        f'<div class="s">{_esc(sub)}</div></div>'
        for rot, val, sub, cls in (
            ("Contratado", k["contratado"], f"{k['contratado_n']} lançamento{'s' if k['contratado_n'] != 1 else ''}"
                                            + _suf(k.get("contratado_sem_valor", 0)), ""),
            ("Pago", k["pago"], pago_sub, "ok"),
            ("A pagar até 30 dias", k["a_pagar_30"], a30_sub, "alerta" if k["vencidos_n"] else ""),
            ("Aguardando o cliente", k["aguardando"], f"{k['aguardando_n']} aguardando o cliente"
                                                      + _suf(k.get("aguardando_sem_valor", 0)), ""),
        ))

    avisos = [f'<p class="nota">{_esc(AVISO_NAO_PRECIFICA)}</p>']
    if k["sem_valor_n"]:
        avisos.append(f'<p class="nota forte">{k["sem_valor_n"]} lançamento'
                      f'{"s" if k["sem_valor_n"] != 1 else ""} sem valor informado: as somas são só do que '
                      f'tem valor, não o custo da obra inteira.</p>')
    if not dados.get("tem_cronograma"):
        avisos.append('<p class="nota">Projeto sem cronograma gerado: vencimento amarrado à fase sai como '
                      '"fase sem data".</p>')

    partes = []
    for g in dados["grupos"]:
        if g["todos_sem_valor"]:
            sufixo = " · sem valor informado"
        elif g["n_sem_valor"]:
            sufixo = f" · {g['n_sem_valor']} sem valor"
        else:
            sufixo = ""
        partes.append(
            f'<tr class="grp"><td colspan="4">{_esc(g["categoria"])} <small>· {g["n"]} lançamento'
            f'{"s" if g["n"] != 1 else ""}{sufixo}</small></td>'
            f'<td class="num{"" if g["total"] is not None else " mudo"}">{_esc(_brl(g["total"]))}</td></tr>')
        for l in g["linhas"]:
            venc = (f'{_esc(l["venc_br"])}<br><small>{_esc(l["venc_regra"])}</small>' if l["venc"]
                    else f'<small class="semdata">{_esc(l["venc_regra"])}</small>')
            forn = _esc(l["fornecedor"]) or '<small class="mudo">—</small>'
            if l["forma_pagamento"]:
                forn += f'<br><small>{_esc(l["forma_pagamento"])}</small>'
            st = (f'<span class="st" style="color:{_ST_COR.get(l["status"], "#475569")}">{_esc(l["status_label"])}</span>'
                  + (f'<br><small>em {_esc(_br(l["pago_em"]))}</small>' if l["pago_em"] else ""))
            if l["vencido"]:
                st += '<br><small class="venc">vencido</small>'
            partes.append(
                f'<tr><td>{_esc(l["descricao"])}<br><small class="mudo">{_esc(l["origem"])}</small></td>'
                f'<td>{forn}</td><td>{venc}</td><td>{st}</td>'
                f'<td class="num{"" if l["valor"] is not None else " mudo"}">{_esc(_brl(l["valor"]))}</td></tr>')
    partes.append(
        f'<tr class="tot"><td colspan="4">Total dos lançamentos com valor'
        f'{(" — " + str(k["sem_valor_n"]) + " sem valor não entram") if k["sem_valor_n"] else ""}</td>'
        f'<td class="num{"" if k["total_com_valor"] is not None else " mudo"}">{_esc(_brl(k["total_com_valor"]))}</td></tr>')

    return f"""<!DOCTYPE html><html><head><meta charset="utf-8"><title>Financeiro da obra</title><style>
@page {{
  size: A4; margin: 1.8cm 1.6cm 2.0cm;
  @top-center {{ content: "Valores informados pelo arquiteto ou vindos das cotações dele — o AI.arq não precifica obra."; color: #64748B; font-size: 7.5pt; }}
  @bottom-center {{ content: "Financeiro da obra · {_css_str(branding.get('project_name') or '')} · gerado pelo AI.arq em {_css_str(hoje_br)} · página " counter(page) " de " counter(pages); color: #6B7280; font-size: 7.5pt; }}
}}
/* Inter é a fonte que a imagem de produção TEM (os templates do cronograma usam a mesma);
   Helvetica/Arial não existem lá e cairiam no DejaVu, mais largo. `gap` em flex só chegou no
   WeasyPrint 65 — produção roda 62.3 — por isso o espaçamento é por margem nos filhos. */
body {{ font-family: 'Inter', 'IBM Plex Sans', 'DejaVu Sans', Helvetica, Arial, sans-serif; font-size: 9.5pt; color: #0F172A; line-height: 1.35; }}
.cab {{ display: flex; align-items: center; border-bottom: 2.5pt solid {acc}; padding-bottom: 8pt; margin-bottom: 8pt; }}
.logo {{ max-height: 42pt; max-width: 120pt; margin-right: 14pt; }}
h1 {{ font-size: 18pt; margin: 0; letter-spacing: -0.2pt; }}
.proj {{ font-size: 11.5pt; color: #334155; margin-top: 1pt; }}
.meta {{ display: flex; color: #475569; font-size: 8.5pt; margin-bottom: 10pt; }}
.meta > div + div {{ margin-left: 18pt; }}
.meta span {{ display: block; text-transform: uppercase; letter-spacing: .6pt; font-size: 6.5pt; color: #94A3B8; }}
.kpis {{ display: flex; margin: 4pt 0 10pt; }}
.kpi {{ flex: 1; border: 1pt solid #E2E8F0; border-top: 3pt solid {acc}; border-radius: 6pt; padding: 7pt 9pt; }}
.kpi + .kpi {{ margin-left: 8pt; }}
.kpi.ok {{ border-top-color: #059669; }} .kpi.alerta {{ border-top-color: #D97706; }}
.kpi .k {{ font-size: 7pt; text-transform: uppercase; letter-spacing: .6pt; color: #64748B; font-weight: 700; }}
.kpi .v {{ font-size: 12.5pt; font-weight: 700; margin: 2pt 0 1pt; white-space: nowrap; }}
.kpi .s {{ font-size: 7.5pt; color: #64748B; }}
.nota {{ font-size: 8pt; color: #64748B; margin: 0 0 4pt; }}
.nota.forte {{ color: #92400E; background: #FFF7ED; border-left: 3pt solid #F59E0B; padding: 4pt 7pt; }}
table {{ width: 100%; border-collapse: collapse; margin-top: 8pt; }}
th {{ text-align: left; font-size: 7pt; text-transform: uppercase; letter-spacing: .5pt; color: #475569; background: #F1F5F9; padding: 5pt 6pt; border-bottom: 1pt solid #CBD5E1; }}
td {{ padding: 5pt 6pt; border-bottom: 0.6pt solid #E2E8F0; vertical-align: top; }}
tr {{ page-break-inside: avoid; }}
/* o título da categoria não fica sozinho no pé da página, nem o TOTAL abre página sozinho */
tr.grp {{ break-after: avoid; }}
tr.tot {{ break-before: avoid; }}
tr.grp td {{ background: #EEF2FF; font-weight: 700; color: #1E1B4B; border-bottom: 1pt solid #C7D2FE; padding-top: 6pt; }}
tr.grp small {{ font-weight: 400; color: #64748B; }}
tr.tot td {{ background: #E8ECF4; font-weight: 700; border-top: 1.5pt solid #94A3B8; }}
.num {{ text-align: right; white-space: nowrap; font-variant-numeric: tabular-nums; }}
small {{ font-size: 7.5pt; color: #64748B; }}
.mudo {{ color: #94A3B8; }} .semdata {{ color: #B45309; }} .venc {{ color: #B91C1C; font-weight: 700; }}
.st {{ font-weight: 600; }}
th.num, td.num {{ text-align: right; }}
</style></head><body>
<div class="cab">{''.join(cab)}</div>
<div class="meta">{meta_html}</div>
<div class="kpis">{kpis}</div>
{''.join(avisos)}
<table>
<thead><tr><th style="width:38%">Item</th><th style="width:20%">Fornecedor · forma de pagamento</th><th style="width:16%">Vencimento</th><th style="width:13%">Status</th><th class="num" style="width:13%">Valor (R$)</th></tr></thead>
<tbody>{''.join(partes)}</tbody>
</table>
</body></html>"""


def render_financeiro_pdf_bytes(dados: Dict, branding: Optional[Dict] = None) -> bytes:
    """HTML → PDF (WeasyPrint). Import aqui dentro, como no memorial e no cronograma."""
    from weasyprint import HTML
    return HTML(string=montar_html_financeiro(dados, branding)).write_pdf()
