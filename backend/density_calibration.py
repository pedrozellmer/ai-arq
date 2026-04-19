# -*- coding: utf-8 -*-
"""Calibração por densidade — aprende RATIOS (qty/área) de orçamentos
históricos e ALERTA sobre anomalias em projetos novos.

Regra do produto (feedback_calibracao_por_densidade.md):
- Armazena densidades por tipologia (ex.: luminárias/m² em "office"), NÃO
  quantidades absolutas.
- Projetos novos NUNCA copiam valores de outros; densidade fora de ±2σ
  gera ALERTA (confidence → estimado + observação laranja).
- Calibração NUNCA promove confidence pra "confirmado".

Tabela Supabase `density_benchmarks`:
  (item_type, typology, unit)  [chave composta única]
  mean, stddev, min_value, max_value, n_projects, updated_at
"""
import json
import re
import time
import statistics
import unicodedata
import urllib.request
from typing import Optional

from openpyxl import load_workbook


SUPABASE_URL = "https://kqjabzwgbfuivzlcfvvu.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImtxamFiendnYmZ1aXZ6bGNmdnZ1Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzYwMDg5NzcsImV4cCI6MjA5MTU4NDk3N30.48xSenZlDV0LfD94ZxwGvX41Kf9Je2n-ouZpJrrCSKI"

_benchmarks_cache = {"data": None, "timestamp": 0}
_CACHE_TTL = 300  # 5 min


def _normalize_item_type(description: str) -> str:
    """Normaliza descrição pra chave estável — mesma lógica do calibrator."""
    if not description:
        return ""
    text = description.lower().strip()
    text = unicodedata.normalize("NFD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text[:50]


def _parse_float(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", ".").strip()
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def extract_density_ratios(xlsx_path: str, area_m2: float, typology: str) -> list[dict]:
    """Lê XLSX histórico e computa qty/área pra cada item com qty numérica.

    Retorna lista de dicts: {item_type, description, unit, qty, density, typology}.
    Tolerante a layouts variados — ignora linhas sem desc ou qty inválida.
    """
    if area_m2 <= 0 or not xlsx_path:
        return []

    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"[density] Erro ao abrir {xlsx_path}: {e}")
        return []

    ratios: list[dict] = []
    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
        except Exception:
            continue
        for row in ws.iter_rows(min_row=1, max_col=10, values_only=True):
            if not row or len(row) < 4:
                continue
            # Heurística flexível: procura uma célula-string (desc) e uma
            # célula numérica (qty) dentro das 10 primeiras colunas. Units
            # costumam estar adjacentes à descrição.
            desc = None
            desc_idx = -1
            for i, cell in enumerate(row[:6]):
                if isinstance(cell, str) and len(cell.strip()) >= 5:
                    desc = cell.strip()
                    desc_idx = i
                    break
            if not desc:
                continue

            unit = ""
            qty = 0.0
            # Procura unit (string curta com "m", "m2", "un", "ml"...) e qty
            # (número) nas colunas após a descrição
            for cell in row[desc_idx + 1:desc_idx + 5]:
                if isinstance(cell, str):
                    s = cell.strip().lower()
                    if s in ("m", "m2", "m²", "un", "ml", "pc", "kg", "l", "h", "dia", "vb"):
                        unit = s.replace("m2", "m²")
                elif isinstance(cell, (int, float)) and cell > 0 and qty == 0:
                    qty = float(cell)
            if qty <= 0:
                continue

            density = qty / area_m2
            if density <= 0 or density > 100:  # filtro de outliers grosseiros
                continue

            # Filtra linhas de totais/headers/seções (conteúdo muito genérico)
            desc_lower = desc.lower()
            if any(w in desc_lower for w in (
                "total geral", "subtotal", "soma", "resumo",
                "descrição", "descricao", "item", "discriminação",
            )) and len(desc.strip()) < 25:
                continue

            ratios.append({
                "item_type": _normalize_item_type(desc),
                "description": desc,
                "unit": unit or "vb",
                "qty": round(qty, 4),
                "density": round(density, 6),
                "typology": typology,
            })

    wb.close()
    return ratios


def _supabase_upsert(table: str, record: dict, on_conflict: str) -> bool:
    """Upsert (insert or update) via Supabase REST."""
    try:
        url = f"{SUPABASE_URL}/rest/v1/{table}?on_conflict={on_conflict}"
        body = json.dumps(record).encode("utf-8")
        req = urllib.request.Request(url, data=body, method="POST")
        req.add_header("apikey", SUPABASE_KEY)
        req.add_header("Authorization", f"Bearer {SUPABASE_KEY}")
        req.add_header("Content-Type", "application/json")
        req.add_header("Prefer", "resolution=merge-duplicates,return=minimal")
        urllib.request.urlopen(req, timeout=10)
        return True
    except Exception as e:
        print(f"[density] Upsert error: {e}")
        return False


def _supabase_select(table: str, query: str = "select=*") -> list[dict]:
    try:
        url = f"{SUPABASE_URL}/rest/v1/{table}?{query}"
        req = urllib.request.Request(url, method="GET")
        req.add_header("apikey", SUPABASE_KEY)
        req.add_header("Authorization", f"Bearer {SUPABASE_KEY}")
        req.add_header("Accept", "application/json")
        resp = urllib.request.urlopen(req, timeout=10)
        return json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        print(f"[density] Select error: {e}")
        return []


def ingest_budget(xlsx_path: str, area_m2: float, typology: str,
                  project_label: str = "") -> dict:
    """Parseia XLSX, classifica cada item via LLM (`classifier.classify_item`)
    e recomputa benchmarks em 3 níveis: família, grupo e capítulo.

    Cada item do XLSX vira:
    - 1 linha em `density_ingest_raw` com familia_id, atributos folha
      extraídos (cor, PD, marca, dim…) e confidence da classificação.
    - Densidades agregadas em `density_benchmarks` com level={familia|grupo|
      capitulo}. Alertas futuros cascateiam leaf → familia → grupo → capitulo.

    Retorna resumo com items_parsed, classificados, benchmarks_updated
    por nível.
    """
    ratios = extract_density_ratios(xlsx_path, area_m2, typology)
    if not ratios:
        return {"items_parsed": 0, "classified": 0,
                "benchmarks_updated_familia": 0,
                "benchmarks_updated_grupo": 0,
                "benchmarks_updated_capitulo": 0,
                "error": "nenhuma linha válida extraída do XLSX"}

    # 1) Classifica cada item via LLM (pode ser lento — N chamadas API)
    classified = 0
    for r in ratios:
        try:
            from classifier import classify_item
            cls = classify_item(r["description"], r["unit"])
        except Exception as e:
            print(f"[density] classifier error: {e}")
            cls = {"familia_id": None, "grupo_id": None, "capitulo_id": None,
                   "confidence": 0, "attributes": {}, "reasoning": f"err: {e}"}
        r["familia_id"] = cls.get("familia_id")
        r["grupo_id"] = cls.get("grupo_id")
        r["capitulo_id"] = cls.get("capitulo_id")
        r["classification_confidence"] = cls.get("confidence") or 0
        r["attributes"] = cls.get("attributes") or {}
        if r["familia_id"]:
            classified += 1

    # 2) Escreve raw com classificação
    for r in ratios:
        raw_record = {
            "typology": typology,
            "item_type": r["item_type"],
            "description": r["description"][:200],
            "unit": r["unit"],
            "qty": r["qty"],
            "area_m2": area_m2,
            "density": r["density"],
            "project_label": project_label[:100],
            "familia_id": r["familia_id"],
            "classification_confidence": r["classification_confidence"],
            "attributes": r["attributes"],
        }
        _supabase_upsert(
            "density_ingest_raw", raw_record,
            "typology,item_type,unit,project_label",
        )

    # 3) Recomputa benchmarks em 3 níveis (familia, grupo, capitulo).
    # A agregação é: pra cada (typology, nivel_id, unit), busca todas as linhas
    # raw que pertencem àquele nó, soma densidades por project_label (mesmo
    # projeto com várias linhas na mesma família vira a soma), e computa
    # mean/stddev através dos projetos.
    updated_fam = _recompute_level(typology, "familia")
    updated_grp = _recompute_level(typology, "grupo")
    updated_cap = _recompute_level(typology, "capitulo")

    _benchmarks_cache["data"] = None
    _benchmarks_cache["timestamp"] = 0

    return {
        "items_parsed": len(ratios),
        "classified": classified,
        "benchmarks_updated_familia": updated_fam,
        "benchmarks_updated_grupo": updated_grp,
        "benchmarks_updated_capitulo": updated_cap,
        "benchmarks_updated": updated_fam + updated_grp + updated_cap,
        "new_item_types": 0,  # mantém chave por retrocompatibilidade
    }


def _recompute_level(typology: str, level: str) -> int:
    """Recomputa benchmarks pra um nível da árvore (familia|grupo|capitulo)
    agregando todas as linhas raw pela classificação + unit. Escreve em
    density_benchmarks. Retorna nº de benchmarks upsertados."""
    id_col = {"familia": "familia_id", "grupo": "grupo_id", "capitulo": "capitulo_id"}[level]

    # Puxa todas as raws classificadas desta tipologia (com classificação não-nula)
    hist = _supabase_select(
        "density_ingest_raw",
        f"select=description,density,project_label,unit,familia_id"
        f"&typology=eq.{urllib.request.quote(typology)}"
        f"&familia_id=not.is.null",
    )
    if not hist:
        return 0

    # Precisa resolver familia_id → grupo_id/capitulo_id pra agregar nos níveis
    # superiores. Busca todos os familia→grupo→capitulo numa query só.
    familias = _supabase_select(
        "catalog_familia",
        "select=id,grupo:grupo_id(id,capitulo_id)",
    )
    fam_to_grp_cap: dict = {}
    for f in familias:
        grupo = f.get("grupo") or {}
        if isinstance(grupo, dict):
            fam_to_grp_cap[f["id"]] = {
                "grupo_id": grupo.get("id"),
                "capitulo_id": grupo.get("capitulo_id"),
            }

    # Agrupa: (nivel_id, unit) → {project_label: soma_density}
    buckets: dict = {}
    for h in hist:
        fid = h.get("familia_id")
        if fid is None:
            continue
        if level == "familia":
            nid = fid
        elif level == "grupo":
            nid = (fam_to_grp_cap.get(fid) or {}).get("grupo_id")
        else:
            nid = (fam_to_grp_cap.get(fid) or {}).get("capitulo_id")
        if nid is None:
            continue
        unit = (h.get("unit") or "").strip()
        lab = h.get("project_label") or ""
        d = _parse_float(h.get("density"))
        if d <= 0:
            continue
        key = (nid, unit)
        per_project = buckets.setdefault(key, {})
        per_project[lab] = per_project.get(lab, 0.0) + d

    # Escreve benchmarks. Usa `item_type` como rótulo derivado do nível pra
    # manter a unique key existente (typology, item_type, unit).
    updated = 0
    for (nid, unit), per_project in buckets.items():
        densities = [d for d in per_project.values() if d > 0]
        if not densities:
            continue
        n = len(densities)
        mean = statistics.mean(densities)
        stddev = statistics.stdev(densities) if n >= 2 else 0.0
        item_type_label = f"{level}:{nid}"
        record = {
            "typology": typology,
            "item_type": item_type_label,
            "unit": unit,
            "mean": round(mean, 6),
            "stddev": round(stddev, 6),
            "min_value": round(min(densities), 6),
            "max_value": round(max(densities), 6),
            "n_projects": n,
            "level": level,
            "updated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        }
        # Popula o FK correto no nível
        if level == "familia":
            record["familia_id"] = nid
        elif level == "grupo":
            record["grupo_id"] = nid
        else:
            record["capitulo_id"] = nid

        if _supabase_upsert("density_benchmarks", record, "typology,item_type,unit"):
            updated += 1
    return updated

    # Invalida cache
    _benchmarks_cache["data"] = None
    _benchmarks_cache["timestamp"] = 0

    return {
        "items_parsed": len(ratios),
        "benchmarks_updated": updated,
        "new_item_types": new_types,
    }


def reclassify_raws(typology: Optional[str] = None,
                    only_unclassified: bool = True,
                    limit: Optional[int] = None) -> dict:
    """Classifica linhas raw existentes via LLM e recomputa benchmarks.

    Útil pra "ativar" raws antigos ingeridos antes do classificador existir.
    Por padrão só toca linhas com `familia_id IS NULL` (idempotente).

    Args:
        typology: filtra por tipologia. None = todas.
        only_unclassified: se True, pula raws já classificados.
        limit: máximo de raws pra processar (útil pra batch incremental).

    Retorna resumo: {raws_total, classified, skipped, recomputed_benchmarks}.
    """
    from classifier import classify_item

    base = ("select=id,description,unit,typology,project_label,familia_id"
            "&order=ingested_at.asc")
    if typology:
        base += f"&typology=eq.{urllib.request.quote(typology)}"
    if only_unclassified:
        base += "&familia_id=is.null"
    if limit:
        base += f"&limit={int(limit)}"

    raws = _supabase_select("density_ingest_raw", base)
    classified = 0
    skipped = 0
    typologies_touched = set()

    for raw in raws:
        rid = raw.get("id")
        desc = raw.get("description") or ""
        unit = raw.get("unit") or ""
        typ = raw.get("typology") or "office"
        if not desc or not rid:
            skipped += 1
            continue

        try:
            cls = classify_item(desc, unit)
        except Exception as e:
            print(f"[reclassify] error on raw {rid}: {e}")
            skipped += 1
            continue

        familia_id = cls.get("familia_id")
        patch: dict = {
            "classification_confidence": cls.get("confidence") or 0,
            "attributes": cls.get("attributes") or {},
        }
        if familia_id is not None:
            patch["familia_id"] = familia_id
            classified += 1
            typologies_touched.add(typ)

        # PATCH in-place pelo id (único)
        try:
            url = f"{SUPABASE_URL}/rest/v1/density_ingest_raw?id=eq.{rid}"
            body = json.dumps(patch).encode("utf-8")
            req = urllib.request.Request(url, data=body, method="PATCH")
            req.add_header("apikey", SUPABASE_KEY)
            req.add_header("Authorization", f"Bearer {SUPABASE_KEY}")
            req.add_header("Content-Type", "application/json")
            req.add_header("Prefer", "return=minimal")
            urllib.request.urlopen(req, timeout=15)
        except Exception as e:
            print(f"[reclassify] patch error on raw {rid}: {e}")
            skipped += 1
            continue

    # Recomputa benchmarks pra cada tipologia tocada
    benchmarks_total = 0
    for typ in typologies_touched:
        for level in ("familia", "grupo", "capitulo"):
            benchmarks_total += _recompute_level(typ, level)

    _benchmarks_cache["data"] = None
    _benchmarks_cache["timestamp"] = 0

    return {
        "raws_total": len(raws),
        "classified": classified,
        "skipped": skipped,
        "typologies_touched": sorted(typologies_touched),
        "benchmarks_recomputed": benchmarks_total,
    }


def get_benchmarks(typology: Optional[str] = None) -> dict:
    """Busca benchmarks do Supabase — indexado por (item_type, unit)."""
    now = time.time()
    if (_benchmarks_cache["data"] is not None and
            (now - _benchmarks_cache["timestamp"]) < _CACHE_TTL):
        cached = _benchmarks_cache["data"]
        if typology:
            return {k: v for k, v in cached.items() if v.get("typology") == typology}
        return cached

    query = "select=*"
    if typology:
        query += f"&typology=eq.{urllib.request.quote(typology)}"
    rows = _supabase_select("density_benchmarks", query)
    index: dict = {}
    for r in rows:
        key = (r.get("item_type", ""), r.get("unit", ""))
        index[key] = {
            "typology": r.get("typology", ""),
            "mean": _parse_float(r.get("mean")),
            "stddev": _parse_float(r.get("stddev")),
            "min_value": _parse_float(r.get("min_value")),
            "max_value": _parse_float(r.get("max_value")),
            "n_projects": int(r.get("n_projects") or 0),
        }
    _benchmarks_cache["data"] = index
    _benchmarks_cache["timestamp"] = now
    return index


def check_density_anomaly(item, project_area_m2: float,
                          benchmarks: Optional[dict] = None,
                          typology: str = "office") -> tuple[bool, str]:
    """Verifica se a densidade do item é anômala.

    Retorna (is_anomaly, motivo). Motivo vazio se ok ou sem benchmark.
    Só ALERTA — nunca sobe confidence pra confirmado.

    Critérios:
    - Precisa de n_projects >= 2 (benchmark estatisticamente pobre abaixo disso)
    - Anomalia alta: density > mean + 2σ OU density > max * 1.5
    - Anomalia baixa: density < mean - 2σ (só se threshold > 0)
    """
    if project_area_m2 <= 0:
        return False, ""
    try:
        qty = float(item.quantity or 0)
    except Exception:
        return False, ""
    if qty <= 0:
        return False, ""

    if benchmarks is None:
        benchmarks = get_benchmarks(typology=typology)
    if not benchmarks:
        return False, ""

    # OTIMIZAÇÃO: se nenhum benchmark tem n>=2, NUNCA vamos disparar alerta
    # (a função sempre retorna False). Pula a chamada de classify_item (~3s
    # de LLM) — caso contrário desperdiçamos tempo+tokens em todo job.
    if not any((b.get("n_projects") or 0) >= 2 for b in benchmarks.values()):
        return False, ""

    # CASCATA: classifica o item e busca benchmark do mais específico pro
    # mais genérico — família → grupo → capítulo. Preserva especificidade
    # quando há dados; ainda dá sinal útil nos níveis superiores.
    try:
        from classifier import classify_item
        cls = classify_item(item.description, item.unit or "")
    except Exception:
        cls = {"familia_id": None, "grupo_id": None, "capitulo_id": None}

    unit = (item.unit or "").strip()
    bench = None
    level_hit = None
    for level, nid in (
        ("familia",  cls.get("familia_id")),
        ("grupo",    cls.get("grupo_id")),
        ("capitulo", cls.get("capitulo_id")),
    ):
        if nid is None:
            continue
        candidate = benchmarks.get((f"{level}:{nid}", unit))
        if candidate and (candidate.get("n_projects") or 0) >= 2:
            bench = candidate
            level_hit = level
            break
    if bench is None:
        return False, ""

    n = bench["n_projects"]

    density = qty / project_area_m2
    mean = bench["mean"]
    stddev = bench["stddev"]
    max_v = bench["max_value"]

    if mean <= 0:
        return False, ""

    level_label = {"familia": "família", "grupo": "grupo", "capitulo": "capítulo"}.get(
        level_hit or "familia", "família")

    # Alta
    threshold_hi = mean + 2 * stddev if stddev > 0 else mean * 2.5
    if density > threshold_hi or (max_v > 0 and density > max_v * 1.5):
        factor = density / mean
        return True, (
            f"densidade {density:.3f} {unit}/m² é {factor:.1f}× maior que "
            f"o típico (nível {level_label}: {mean:.3f} ± {stddev:.3f}, "
            f"n={n} projetos) — possível dupla contagem"
        )
    # Baixa (só se faz sentido)
    threshold_lo = mean - 2 * stddev if stddev > 0 else mean / 2.5
    if threshold_lo > 0 and density < threshold_lo:
        factor = mean / density if density > 0 else 0
        return True, (
            f"densidade {density:.3f} {unit}/m² é {factor:.1f}× menor que "
            f"o típico (nível {level_label}: {mean:.3f} ± {stddev:.3f}, "
            f"n={n} projetos) — possível subcontagem"
        )

    return False, ""


def ddl_statements() -> list[str]:
    """Retorna as declarações SQL das tabelas do Supabase — referência/doc."""
    return [
        """CREATE TABLE IF NOT EXISTS density_benchmarks (
            typology text NOT NULL,
            item_type text NOT NULL,
            unit text NOT NULL,
            mean double precision NOT NULL,
            stddev double precision DEFAULT 0,
            min_value double precision DEFAULT 0,
            max_value double precision DEFAULT 0,
            n_projects integer DEFAULT 1,
            updated_at timestamptz DEFAULT now(),
            PRIMARY KEY (typology, item_type, unit)
        );""",
        """CREATE TABLE IF NOT EXISTS density_ingest_raw (
            id bigserial PRIMARY KEY,
            typology text NOT NULL,
            item_type text NOT NULL,
            description text,
            unit text NOT NULL,
            qty double precision,
            area_m2 double precision,
            density double precision,
            project_label text,
            ingested_at timestamptz DEFAULT now(),
            UNIQUE (typology, item_type, unit, project_label)
        );""",
    ]
