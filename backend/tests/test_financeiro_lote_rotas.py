# -*- coding: utf-8 -*-
"""As rotas do lote: baixar o modelo, conferir e só então gravar.

06/09/2026. O Pedro pediu duas coisas no mesmo dia: *"podemos baixar a planilha padrão para preencher
os preços e subir depois, em lote"* e *"sempre deixa aviso pro usuário né"*. As duas viram FATO aqui:

  • `/modelo.xlsx` sai com a coluna de valor VAZIA (nº5) e a âncora de cada item (nº7);
  • `/lote/conferir` NÃO grava nada — devolve a frase em português, os avisos e o total;
  • `/lote/aplicar` relê a planilha do zero: o servidor nunca confia numa lista de ações vinda do
    cliente (senão dava pra mandar "cria lançamento no projeto do vizinho");
  • admin LÊ o financeiro do cliente mas não escreve — as duas rotas do lote recusam (LGPD nº6);
  • arquivo torto (não-xlsx, vazio, gigante) é MENSAGEM em português, nunca 500;
  • uma linha que o banco recusa não derruba as outras: volta na lista de falhas com o motivo.
🧪 Controles: conferir com o banco fora é 502 (não "planilha vazia"); e a rota de aplicar chama a
mesma conferência, não uma segunda régua.
"""
import io
import os
import sys
import types

import pytest

_AQUI = os.path.dirname(os.path.abspath(__file__))
_BACKEND = os.path.dirname(_AQUI)
sys.path.insert(0, _BACKEND)
sys.path.insert(0, _AQUI)

import main  # noqa: E402
import financeiro_lote as fl  # noqa: E402
from _corpo import corpo_de, fonte, sem_comentarios  # noqa: E402

JOB = "job-lote-1"
ITEM_A = "11111111-1111-4111-8111-111111111111"
REQ = types.SimpleNamespace(headers={"Authorization": "Bearer jwt"}, state=types.SimpleNamespace())

ITENS = [{"id": ITEM_A, "description": "Porcelanato 60x60", "quantity": 1062.0, "unit": "m2",
          "discipline": "Pisos", "sort_order": 1}]


@pytest.fixture
def casa(monkeypatch):
    monkeypatch.setattr(main, "_require_project_owner", lambda request, job_id: "uid-dono")
    monkeypatch.setattr(main, "_fin_eh_admin", lambda request, owner: False)
    monkeypatch.setattr(main, "_log_error", lambda *a, **k: None)
    monkeypatch.setattr(main, "_fin_nome_do_projeto", lambda req, job_id: (200, "Casa Teste"))


def _banco(monkeypatch, lanc=(200, []), itens=(200, None)):
    itens = (200, ITENS) if itens == (200, None) else itens
    escritas = []

    def fake(request, method, path, body=None, params=None, prefer=None, timeout=15):
        if method == "GET" and "/project_items?" in path:
            return itens
        if method == "GET" and f"/{main._FIN_TABELA}?" in path:
            return lanc
        escritas.append({"m": method, "path": path, "body": body})
        return (201, None) if method == "POST" else (204, None)
    monkeypatch.setattr(main, "_supa_rest_as_user", fake)
    return escritas


def _planilha_bytes(linhas):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = fl.ABA
    ws.append(["FINANCEIRO DA OBRA — PREENCHER OS PREÇOS"])
    ws.append(fl.COLS)
    for l in linhas:
        ws.append(l)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════════════════════════════════
#  o modelo
# ══════════════════════════════════════════════════════════════════════════
def test_modelo_sai_com_nome_do_projeto_e_valor_vazio(monkeypatch, casa):
    _banco(monkeypatch)
    r = main.financeiro_modelo_lote(JOB, REQ)
    assert r.media_type == main._FIN_XLSX_MIME
    assert r.filename == "financeiro_preencher_Casa_Teste.xlsx", "o nome diz o que é: pra preencher"
    from openpyxl import load_workbook
    ws = load_workbook(r.path)[fl.ABA]
    linha = [r_[6].value for r_ in ws.iter_rows() if r_[2].value == "Porcelanato 60x60"]
    assert linha == [None], "🔒 nº5: a coluna VALOR sai vazia"


def test_modelo_recusa_projeto_vazio_e_projeto_gigante(monkeypatch, casa):
    _banco(monkeypatch, lanc=(200, []), itens=(200, []))
    with pytest.raises(main.HTTPException) as ex:
        main.financeiro_modelo_lote(JOB, REQ)
    assert ex.value.status_code == 404 and "não há o que preencher" in ex.value.detail
    _banco(monkeypatch, itens=(200, [dict(ITENS[0], id=str(i)) for i in range(1000)]))
    with pytest.raises(main.HTTPException) as ex2:
        main.financeiro_modelo_lote(JOB, REQ)
    assert ex2.value.status_code == 409


def test_modelo_com_banco_fora_e_502(monkeypatch, casa):
    _banco(monkeypatch, lanc=(500, None))
    with pytest.raises(main.HTTPException) as ex:
        main.financeiro_modelo_lote(JOB, REQ)
    assert ex.value.status_code == 502


# ══════════════════════════════════════════════════════════════════════════
#  conferir — o aviso que o Pedro pediu, ANTES de gravar
# ══════════════════════════════════════════════════════════════════════════
def test_conferir_NAO_grava_e_diz_o_que_vai_acontecer(monkeypatch, casa):
    escritas = _banco(monkeypatch)
    dados = _planilha_bytes([[1, "Pisos", "Porcelanato 60x60", 1062, "m2", "Cerâmica Boa",
                              "48.000,00", "3x", "", "Contratado", "i:%s" % ITEM_A]])
    conf = main._fin_lote_conferencia(REQ, JOB, False, dados)
    assert conf["status"] == "ok"
    assert conf["frase"] == "1 lançamento novo"
    assert conf["total"] == "R$ 48.000,00", "o total em dinheiro aparece ANTES de gravar"
    assert conf["acoes"][0]["corpo"]["origem_ref_id"] == ITEM_A
    assert escritas == [], "🚫 conferir NÃO escreve nada no banco"


def test_conferir_com_planilha_torta_devolve_mensagem_nao_erro(monkeypatch, casa):
    _banco(monkeypatch)
    conf = main._fin_lote_conferencia(REQ, JOB, False, b"isto nao e uma planilha")
    assert conf["status"] == "erro" and "não consegui abrir" in conf["erro"]
    assert conf["acoes"] == []


def test_CONTROLE_conferir_com_banco_fora_e_502_nao_planilha_vazia(monkeypatch, casa):
    _banco(monkeypatch, lanc=(503, None))
    with pytest.raises(main.HTTPException) as ex:
        main._fin_lote_conferencia(REQ, JOB, False, _planilha_bytes([]))
    assert ex.value.status_code == 502


# ══════════════════════════════════════════════════════════════════════════
#  aplicar — grava, e o que falha volta com motivo
# ══════════════════════════════════════════════════════════════════════════
def test_aplicar_grava_com_o_retrato_da_origem(monkeypatch, casa):
    escritas = _banco(monkeypatch)
    monkeypatch.setattr(main, "_fin_retrato_da_origem",
                        lambda request, job_id, origem, ref, pos=None: {
                            "descricao": "Porcelanato 60x60", "origem_ref_id": ref,
                            "origem_quantidade": 1062.0, "origem_unidade": "m2"})
    dados = _planilha_bytes([[1, "Pisos", "Porcelanato 60x60", 1062, "m2", "Cerâmica Boa",
                              "48.000,00", "3x", "", "Contratado", "i:%s" % ITEM_A]])
    conf = main._fin_lote_conferencia(REQ, JOB, False, dados)
    saida = main._fin_lote_gravar(REQ, JOB, conf)
    assert saida["criados"] == 1 and saida["atualizados"] == 0 and saida["falhas"] == []
    assert saida["frase"] == "1 lançamento criado"
    corpo = [e for e in escritas if e["m"] == "POST"][0]["body"]
    assert corpo["job_id"] == JOB and corpo["escopo"] == "obra"
    assert corpo["origem_quantidade"] == 1062.0 and corpo["origem_unidade"] == "m2", (
        "🔗 nº7: o retrato da origem entra no MESMO insert — é ele que faz o aviso 'item mudou'")
    assert corpo["valor"] == 48000.0


def test_uma_linha_recusada_nao_derruba_as_outras(monkeypatch, casa):
    chamadas = {"n": 0}

    def fake(request, method, path, body=None, params=None, prefer=None, timeout=15):
        if method == "GET" and "/project_items?" in path:
            return (200, ITENS)
        if method == "GET":
            return (200, [])
        chamadas["n"] += 1
        return (400, None) if chamadas["n"] == 1 else (201, None)
    monkeypatch.setattr(main, "_supa_rest_as_user", fake)
    monkeypatch.setattr(main, "_fin_retrato_da_origem",
                        lambda request, job_id, origem, ref, pos=None: {"descricao": "x"})
    conf = {"acoes": [
        {"acao": "cria", "n": 1, "item": "Primeira", "corpo": {
            "escopo": "obra", "origem": "livre", "descricao": "Primeira", "categoria": "Pisos",
            "valor": 10.0, "status": "cotado", "venc_tipo": "fase", "venc_fase": "Pisos",
            "venc_quando": "inicio"}},
        {"acao": "cria", "n": 2, "item": "Segunda", "corpo": {
            "escopo": "obra", "origem": "livre", "descricao": "Segunda", "categoria": "Pisos",
            "valor": 20.0, "status": "cotado", "venc_tipo": "fase", "venc_fase": "Pisos",
            "venc_quando": "inicio"}},
    ], "avisos": []}
    saida = main._fin_lote_gravar(REQ, JOB, conf)
    assert saida["criados"] == 1 and len(saida["falhas"]) == 1
    assert saida["falhas"][0]["n"] == 1 and saida["falhas"][0]["item"] == "Primeira"
    assert "1 lançamento criado" in saida["frase"] and "1 linha não entrou" in saida["frase"]


def test_atualizar_normaliza_a_linha_MESCLADA(monkeypatch, casa):
    """A coerência (fase OU data, pago exige valor) é julgada na linha inteira, como no PATCH da tela."""
    escritas = _banco(monkeypatch)
    atual = {"id": "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa", "job_id": JOB, "escopo": "obra",
             "categoria": "Pisos", "descricao": "Rejunte", "origem": "livre", "fornecedor": "",
             "valor": 900.0, "forma_pagamento": "", "venc_tipo": "fase", "venc_fase": "Pisos",
             "venc_quando": "inicio", "venc_data": None, "status": "contratado", "pago_em": None}
    conf = {"acoes": [{"acao": "atualiza", "id": atual["id"], "n": 1, "item": "Rejunte",
                       "campos": {"valor": 1500.0}, "atual": atual}], "avisos": []}
    saida = main._fin_lote_gravar(REQ, JOB, conf)
    assert saida["atualizados"] == 1 and saida["falhas"] == []
    patch = [e for e in escritas if e["m"] == "PATCH"][0]
    assert f"job_id=eq.{JOB}" in patch["path"], "🔒 nº2: escrita por id filtra o job na URL"
    assert patch["body"]["valor"] == 1500.0 and patch["body"]["categoria"] == "Pisos"


# ══════════════════════════════════════════════════════════════════════════
#  forma: o que o fonte tem que garantir
# ══════════════════════════════════════════════════════════════════════════
def test_aplicar_RELE_a_planilha_e_nao_confia_em_lista_do_cliente():
    c = corpo_de("financeiro_lote_aplicar")
    assert "_fin_lote_ler_upload(file)" in c and "_fin_lote_conferencia" in c, (
        "aplicar relê o arquivo e refaz a conferência — lista de ações vinda do cliente seria porta "
        "pra gravar no projeto de outro")
    assert "conf.get(\"erro\")" in c


def test_as_duas_rotas_do_lote_recusam_o_admin_e_conferem_o_dono():
    for rota in ("financeiro_lote_conferir", "financeiro_lote_aplicar"):
        c = corpo_de(rota)
        assert "_require_project_owner(request, job_id)" in c
        assert "_fin_so_o_dono_escreve(request, owner)" in c, "LGPD nº6: admin lê, não escreve"
    modelo = corpo_de("financeiro_modelo_lote")
    assert "_require_project_owner(request, job_id)" in modelo
    assert "_fin_so_o_dono_escreve" not in modelo, "baixar o modelo é leitura — o admin pode"


def test_o_trabalho_pesado_do_lote_roda_fora_do_laco():
    for rota in ("financeiro_lote_conferir", "financeiro_lote_aplicar"):
        c = corpo_de(rota)
        assert "await run_in_threadpool(_fin_lote_conferencia" in c, rota
    assert "await run_in_threadpool(_fin_lote_gravar" in corpo_de("financeiro_lote_aplicar")


def test_upload_recusa_o_que_nao_e_xlsx_com_texto_em_portugues():
    c = corpo_de("_fin_lote_ler_upload")
    assert '.endswith(".xlsx")' in c and "salve como .xlsx no Excel" in c
    assert "_FIN_LOTE_TETO_MB" in c and "chegou vazia" in c


def test_a_tela_confere_antes_de_gravar_e_mostra_os_avisos():
    html = fonte("financeiro.html")
    js = "\n".join(__import__("re").findall(r"<script>(.*?)</script>", html, __import__("re").S))
    assert "onchange=\"conferirLote()\"" in html, "escolher o arquivo já mostra a conferência"
    assert "id=\"btn-lote-aplicar\" type=\"button\" hidden" in html, (
        "o botão de gravar nasce ESCONDIDO — só aparece depois da conferência")
    assert "document.getElementById('btn-lote-aplicar').hidden = !(d.acoes && d.acoes.length)" in js
    assert "lote-avisos" in js and "⚠" in js, "os avisos linha a linha chegam ao arquiteto"
    assert "Somando ${d.total}" in js or "Somando " in js
    # e o evento de uso existe dos dois lados
    MAIN = sem_comentarios(fonte("main.py"))
    for ev in ("fin_lote_abriu", "fin_lote_modelo", "fin_lote_aplicou"):
        assert f'"{ev}"' in MAIN and f"trackEvent('{ev}'" in js, ev
