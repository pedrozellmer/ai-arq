# -*- coding: utf-8 -*-
"""A planilha não pode APAGAR item diferente que começa igual.

🩸 O gerador deduplicava pelos primeiros 50 caracteres da descrição, com um
`continue` MUDO. Dois itens distintos cujo texto começa igual: o segundo nunca
chegava no arquivo que o cliente baixa — e continuava aparecendo na tela de
revisão e no banco. Ninguém via o buraco.

📏 MEDIDO NO ACERVO em 01/09/2026: 106 grupos colidem pelo prefixo de 50
caracteres. Em apenas **5** a descrição INTEIRA é igual (duplicata de verdade).
Nos outros **101** são itens distintos — **144 linhas perdidas em 37 de 130
jobs (28%)**, sendo 138 delas com quantidade preenchida.

Casos REAIS colhidos do banco (todos com prefixo idêntico e item diferente):

    Sapata S1 (160×160×60cm) 14,75 m³   ×   S2 (100×100×40cm) 4,40 m³
    Contrapiso área interna 593 m²      ×   varanda/circulação 276,91 m²
    Rodapé pav. superior 42 ml          ×   pav. inferior 45 ml
    Divisória gesso Standard 15 m²      ×   umidade 8 m²  ×  fogo 6 m²
    Porta interna 8 un                  ×   10 un

🪤 O par S1/S2 JÁ ESTAVA ESCRITO no comentário do próprio código, desde 24/08 —
e o conserto daquele dia tapou só o caminho `origem='revisao_cliente'`. O caso
geral, que é 28% dos jobs, ficou de pé mais uma semana.

🔑 Agora só cai fora o que é EXATAMENTE igual (descrição + unidade +
quantidade), e o que cair fica registrado. A dedup de verdade é a de montante;
esta é a última rede — e rede que apaga o que não devia é pior que rede nenhuma.
"""
import os
import sys

_BACKEND = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _BACKEND)

from models import BudgetItem, ProjectData, Confidence  # noqa: E402
from spreadsheet import generate_spreadsheet            # noqa: E402
from openpyxl import load_workbook                      # noqa: E402


def _it(desc, qtd, unid, disc="Estrutura"):
    return BudgetItem(item_num="", description=desc, unit=unid, quantity=qtd,
                      discipline=disc, confidence=Confidence.ESTIMADO)


# ── Os pares REAIS, como estão no banco ────────────────────────────────────
_SAPATAS = [
    _it("Escavação manual/mecânica de cavas para sapatas tipo S1 (160×160×60cm) "
        "— solo em condições normais", 14.75, "m³"),
    _it("Escavação manual/mecânica de cavas para sapatas tipo S2 (100×100×40cm) "
        "— solo em condições normais", 4.40, "m³"),
]
_CONTRAPISO = [
    _it("Regularização e contrapiso sobre terreno — concreto magro 5cm — área "
        "interna dos compartimentos", 593.0, "m²"),
    _it("Regularização e contrapiso sobre terreno — concreto magro 5cm — varanda "
        "e circulação externa", 276.91, "m²"),
]
_DIVISORIA = [
    _it("Demolição de divisória em gesso — placa de gesso do tipo Standard",
        15.0, "m²", "Demolição"),
    _it("Demolição de divisória em gesso — placa de gesso do tipo resistente à "
        "umidade — chapa verde", 8.0, "m²", "Demolição"),
    _it("Demolição de divisória em gesso — placa de gesso do tipo resistente ao "
        "fogo — chapa rosa", 6.0, "m²", "Demolição"),
]


def _gera(itens, tmp_path):
    saida = str(tmp_path / "p.xlsx")
    generate_spreadsheet(ProjectData(name="teste"), itens, saida)
    wb = load_workbook(saida)
    texto = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for c in row:
                if c is not None:
                    texto.append(str(c))
    return " || ".join(texto)


def test_CONTROLE_os_pares_REALMENTE_colidem_no_corte_de_50():
    """🧪 Controle positivo do CENÁRIO: se os prefixos não fossem iguais, este
    arquivo inteiro estaria guardando um problema que não existe."""
    for par in (_SAPATAS, _CONTRAPISO, _DIVISORIA):
        chaves = {i.description.lower().strip()[:50] for i in par}
        assert len(chaves) == 1, (
            "o par deixou de colidir no corte de 50 caracteres — o teste perdeu "
            "o sentido: %s" % chaves)


def test_sapata_S1_e_S2_saem_as_DUAS(tmp_path):
    """🩸 O par que estava escrito no comentário do código desde 24/08."""
    txt = _gera(_SAPATAS, tmp_path)
    assert "S1" in txt and "S2" in txt, "sumiu uma das sapatas"
    assert "14.75" in txt or "14,75" in txt
    assert "4.4" in txt or "4,4" in txt, (
        "os 4,40 m³ da sapata S2 não chegaram na planilha")


def test_contrapiso_interno_e_externo_saem_os_DOIS(tmp_path):
    txt = _gera(_CONTRAPISO, tmp_path)
    assert "593" in txt, "sumiu a área interna"
    assert "276.91" in txt or "276,91" in txt, (
        "os 276,91 m² de contrapiso da varanda sumiram da planilha")


def test_as_TRES_divisorias_de_gesso_saem(tmp_path):
    """Standard, resistente à umidade e resistente ao fogo têm preço diferente."""
    txt = _gera(_DIVISORIA, tmp_path)
    for esperado in ("Standard", "umidade", "fogo"):
        assert esperado in txt, "sumiu a divisória '%s'" % esperado
    for q in ("15", "8", "6"):
        assert q in txt


def test_duplicata_EXATA_continua_sendo_descartada(tmp_path):
    """Controle negativo: a rede não pode deixar de existir."""
    ident = "Alvenaria de vedação em bloco cerâmico — paredes novas"
    itens = [_it(ident, 40.0, "m²"), _it(ident, 40.0, "m²"), _it(ident, 40.0, "m²")]
    saida = str(tmp_path / "d.xlsx")
    generate_spreadsheet(ProjectData(name="t"), itens, saida)
    wb = load_workbook(saida)
    ws = [w for w in wb.worksheets if w.title != "Resumo Comparativo"][0]
    n = sum(1 for row in ws.iter_rows(values_only=True)
            if any(c is not None and ident in str(c) for c in row))
    assert n == 1, "a duplicata EXATA apareceu %d vezes — a rede sumiu" % n


def test_mesma_descricao_com_quantidade_DIFERENTE_nao_e_duplicata(tmp_path):
    """🪤 Duas leituras do mesmo texto com números diferentes são informação,
    não repetição: apagar uma esconde a divergência do cliente."""
    d = "Porta interna — tipo, dimensão e material a especificar — inclui batente"
    txt = _gera([_it(d, 8, "un", "Esquadrias"), _it(d, 10, "un", "Esquadrias")],
                tmp_path)
    assert "8" in txt and "10" in txt


def test_item_da_revisao_do_cliente_continua_intocado(tmp_path):
    """🪤 Regra dura nº7: o que veio da revisão do cliente nunca é descartado —
    nem quando é byte a byte igual a outra linha."""
    d = "Piso vinílico — sala de reunião"
    a = _it(d, 30.0, "m²", "Pisos")
    b = _it(d, 30.0, "m²", "Pisos")
    b.origem = "revisao_cliente"
    saida = str(tmp_path / "r.xlsx")
    generate_spreadsheet(ProjectData(name="t"), [a, b], saida)
    wb = load_workbook(saida)
    ws = [w for w in wb.worksheets if w.title != "Resumo Comparativo"][0]
    n = sum(1 for row in ws.iter_rows(values_only=True)
            if any(c is not None and d in str(c) for c in row))
    assert n == 2, ("a linha da revisão do cliente foi apagada (apareceu %d vez) "
                    "— fura a regra nº7" % n)


def test_o_descarte_deixa_RASTRO(capsys, tmp_path):
    """Até 01/09 esta rede apagava em silêncio. Nunca mais."""
    ident = "Forro de gesso acartonado — placa 12,5mm"
    generate_spreadsheet(ProjectData(name="t"),
                         [_it(ident, 12.0, "m²", "Forros"),
                          _it(ident, 12.0, "m²", "Forros")],
                         str(tmp_path / "l.xlsx"))
    saida = capsys.readouterr().out
    assert "dedup final descartou" in saida, (
        "descartou linha sem registrar nada:\n" + saida)
