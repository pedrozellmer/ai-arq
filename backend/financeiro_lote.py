# -*- coding: utf-8 -*-
"""Financeiro da obra — preencher os preços EM LOTE (a planilha vai e volta).

06/09/2026, Pedro: *"podemos baixar a planilha padrão para preencher os preços e subir depois né,
em lote, sacou? facilitar o trabalho do arquiteto"*.

O arquiteto não trabalha item a item numa tela: ele manda a planilha pro fornecedor, recebe de volta
com preço, e sem isto teria que redigitar tudo aqui dentro. Aqui a planilha SAI com uma linha por
item medido (mais os lançamentos que já existem) e VOLTA virando lançamento.

Regras da casa que mandam neste arquivo:
  • nº5 — o AI.arq NÃO precifica. As colunas de dinheiro saem VAZIAS; quem preenche é o arquiteto ou
    o fornecedor dele. Célula vazia continua vazia (nunca 0).
  • nº7 — tudo interligado: cada linha criada guarda o RETRATO do item de origem (quantidade a 4
    casas e unidade) no mesmo insert, igual ao lançamento feito na tela. É esse retrato que faz o
    aviso "item mudou no quantitativo" funcionar depois.
  • nº2 — isolamento: a âncora de cada linha é resolvida SEMPRE dentro do job_id do projeto.
  • "1.234" vale MIL em pt-BR. O parser de valor é UM SÓ, aqui, no servidor.
  • 🪤 O `/add-file` RECRIA os ids dos itens: planilha baixada antes de re-subir a planta aponta pra
    item que não existe mais. Isso não pode virar erro seco nem linha fantasma — vira aviso na
    conferência e a linha entra como livre, com o valor que a pessoa digitou.

Fluxo: `gerar_modelo_xlsx` → o arquiteto preenche → `ler_planilha_lote` → `conferir_lote`
(o que VAI acontecer, em português) → o servidor grava só depois do OK.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from spreadsheet import (
    F_TITLE, F_SUB, F_HDR, F_N, F_NOTE,
    P_HDR, P_LT, P_YEL,
    AC, AL, AR, BD,
)
from financeiro_export import FMT_MOEDA, FMT_DATA, STATUS_LABEL, _brl

ABA = "Preencher preços"
# 🔑 A âncora. Ela é o que liga a linha da planilha ao que existe no banco:
#   l:<uuid>  lançamento que JÁ existe (a pessoa está editando)
#   i:<uuid>  item do quantitativo ainda sem lançamento (vai criar)
#   vazio     linha que a pessoa acrescentou à mão (vira linha livre)
_RE_ANCORA = re.compile(r"^([li]):([0-9a-fA-F-]{36})$")

COLS = ['Nº', 'CATEGORIA', 'ITEM', 'QUANTIDADE', 'UN.',
        'FORNECEDOR', 'VALOR (R$)', 'FORMA DE PAGAMENTO', 'VENCIMENTO', 'STATUS', 'REFERÊNCIA']
_LARGURAS = [6, 22, 52, 13, 7, 26, 16, 22, 13, 20, 40]
_COL_REF = len(COLS)          # 11 — a última, escondida
_STATUS_OPCOES = ["Cotado", "Aguardando o cliente", "Aprovado", "Contratado", "Pago"]
_STATUS_POR_LABEL = {v.lower(): k for k, v in STATUS_LABEL.items()}

INSTRUCOES = (
    "Preencha as colunas FORNECEDOR, VALOR, FORMA DE PAGAMENTO, VENCIMENTO e STATUS. "
    "Deixe em branco o que ainda não sabe — linha sem valor entra como \"sem valor informado\", "
    "nunca como zero. Não apague nem reordene as colunas, e não mexa na coluna REFERÊNCIA "
    "(escondida): é ela que liga cada linha ao item do seu projeto. Pode acrescentar linhas no fim "
    "para o que não está no quantitativo — elas entram como linha livre."
)


# ══════════════════════════════════════════════════════════════════════════
#  IDA — o modelo pra preencher
# ══════════════════════════════════════════════════════════════════════════
def _texto(v) -> str:
    return "" if v is None else str(v).strip()


def _qtd(v):
    try:
        f = float(v)
    except Exception:
        return None
    return None if f != f else round(f, 4)      # NaN fora


def montar_linhas_do_modelo(itens: List[Dict], lancamentos: List[Dict]) -> List[Dict]:
    """As linhas do modelo, na ordem em que o arquiteto espera vê-las:
    1) o que ele JÁ lançou (pra conferir/corrigir), 2) os itens medidos que ainda não viraram
    lançamento (pra preencher). Item que já tem lançamento não aparece duas vezes."""
    por_item = {}
    for l in lancamentos or []:
        if l.get("origem") == "quantitativo" and l.get("origem_ref_id"):
            por_item.setdefault(str(l["origem_ref_id"]), []).append(l)

    linhas = []
    for l in lancamentos or []:
        linhas.append({
            "ancora": "l:%s" % l.get("id"),
            "categoria": _texto(l.get("categoria")),
            "item": _texto(l.get("descricao")),
            "quantidade": _qtd(l.get("origem_quantidade")),
            "unidade": _texto(l.get("origem_unidade")),
            "fornecedor": _texto(l.get("fornecedor")),
            "valor": l.get("valor"),
            "forma": _texto(l.get("forma_pagamento")),
            "venc": l.get("venc_data") if l.get("venc_tipo") == "data" else None,
            "status": STATUS_LABEL.get(_texto(l.get("status")) or "cotado", "Cotado"),
            "ja_existe": True,
        })
    for it in itens or []:
        if str(it.get("id")) in por_item:
            continue                      # já virou lançamento: não repete
        linhas.append({
            "ancora": "i:%s" % it.get("id"),
            "categoria": _texto(it.get("discipline")),
            "item": _texto(it.get("description")),
            "quantidade": _qtd(it.get("quantity")),
            "unidade": _texto(it.get("unit")),
            "fornecedor": "", "valor": None, "forma": "", "venc": None, "status": "",
            "ja_existe": False,
        })
    return linhas


def gerar_modelo_xlsx(linhas: List[Dict], output_path: str, branding: Optional[Dict] = None,
                      hoje: Optional[date] = None) -> str:
    """Escreve o modelo .xlsx e devolve o caminho."""
    branding = branding or {}
    wb = Workbook()
    ws = wb.active
    ws.title = ABA
    ws.sheet_properties.tabColor = '4F46E5'

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_COL_REF - 1)
    c = ws.cell(row=1, column=1, value='FINANCEIRO DA OBRA — PREENCHER OS PREÇOS')
    c.font = F_TITLE
    c.alignment = AL

    ident = " · ".join(x for x in (
        "Projeto: %s" % (branding.get("project_name") or "Projeto sem nome"),
        ("Escritório: %s" % branding["architect_name"]) if branding.get("architect_name") else "",
        "Baixado em %s" % (hoje or date.today()).strftime("%d/%m/%Y"),
    ) if x)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_COL_REF - 1)
    c = ws.cell(row=2, column=1, value=ident)
    c.font = F_N
    c.alignment = AL

    # a ressalva (nº5) e as instruções viajam DENTRO do arquivo: ele sai do site e chega
    # em fornecedor que nunca viu a nossa tela
    for texto, fonte_, fill, altura in (
        ("O AI.arq não precifica obra: os valores desta planilha são seus ou dos seus fornecedores. "
         "As quantidades vieram do seu quantitativo e estão aqui só para consulta.", F_NOTE, None, 26),
        (INSTRUCOES, F_SUB, P_YEL, 42),
    ):
        ws.merge_cells(start_row=ws.max_row + 1, start_column=1,
                       end_row=ws.max_row + 1, end_column=_COL_REF - 1)
        cc = ws.cell(row=ws.max_row, column=1, value=texto)
        cc.font = fonte_
        cc.alignment = AL
        if fill:
            cc.fill = fill
        ws.row_dimensions[ws.max_row].height = altura

    ro = ws.max_row + 2
    linha_hdr = ro
    for i, h in enumerate(COLS, start=1):
        cc = ws.cell(row=ro, column=i, value=h)
        cc.font = F_HDR
        cc.fill = P_HDR
        cc.alignment = AC
        cc.border = BD
    ro += 1

    primeira = ro
    for n, l in enumerate(linhas, start=1):
        alt = P_LT if n % 2 == 0 else None
        vals = [n, l["categoria"], l["item"], l["quantidade"], l["unidade"],
                l["fornecedor"], l["valor"], l["forma"],
                (_data(l["venc"]) if l.get("venc") else None), l["status"], l["ancora"]]
        for i, v in enumerate(vals, start=1):
            cc = ws.cell(row=ro, column=i, value=v)
            cc.font = F_N
            cc.border = BD
            cc.alignment = AL if i in (2, 3, 6, 8, 10) else (AR if i in (4, 7) else AC)
            if alt:
                cc.fill = alt
            if i == 4 and v is not None:
                cc.number_format = '#,##0.####'
            if i == 7:
                cc.number_format = FMT_MOEDA          # vazio continua vazio (nº5)
                if v is None:
                    cc.fill = P_YEL                   # amarelo = "preencha aqui"
            if i == 9:
                cc.number_format = FMT_DATA
        ro += 1
    ultima = ro - 1

    # lista fechada no STATUS: menos erro de digitação na volta
    if ultima >= primeira:
        dv = DataValidation(type="list", formula1='"%s"' % ",".join(_STATUS_OPCOES),
                            allow_blank=True, showDropDown=False)
        dv.error = "Escolha um dos status da lista."
        dv.prompt = "Deixe em branco para 'Cotado'."
        ws.add_data_validation(dv)
        col = get_column_letter(10)
        dv.add("%s%d:%s%d" % (col, primeira, col, ultima))

    for i, w in enumerate(_LARGURAS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # a âncora fica escondida: é maquinaria, não informação pro fornecedor
    ws.column_dimensions[get_column_letter(_COL_REF)].hidden = True
    ws.freeze_panes = "C%d" % (linha_hdr + 1)
    wb.save(output_path)
    return output_path


# ══════════════════════════════════════════════════════════════════════════
#  VOLTA — ler o que o arquiteto preencheu
# ══════════════════════════════════════════════════════════════════════════
def _data(v) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except Exception:
            continue
    return None


def valor_do_texto(v) -> Tuple[Optional[float], Optional[str]]:
    """(valor, erro). O parser de dinheiro é UM SÓ e mora aqui.

    🪤 "1.234" vale MIL em pt-BR e um-vírgula-dois em inglês — errar isso é erro de 1000× EM
    DINHEIRO. A régua: se tem vírgula, a vírgula é o decimal e o ponto é milhar. Sem vírgula, um
    ponto só com 1 ou 2 casas depois é decimal ("1234.5"); qualquer outro ponto é milhar.
    Célula vazia devolve (None, None) — ausência, nunca zero (nº5).
    """
    if v is None:
        return None, None
    if isinstance(v, bool):
        return None, "valor inválido"
    if isinstance(v, (int, float)):
        f = float(v)
        if f != f:
            return None, "valor inválido"
        return (round(f, 2), None) if f >= 0 else (None, "valor negativo")
    s = str(v).strip()
    if not s:
        return None, None
    s = re.sub(r"(?i)r\$|\s| ", "", s)
    if not s:
        return None, None
    if not re.fullmatch(r"-?[\d.,]+", s):
        return None, "não parece um valor"
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "." in s:
        inteiro, _, dec = s.rpartition(".")
        s = (inteiro.replace(".", "") + "." + dec) if len(dec) in (1, 2) else s.replace(".", "")
    try:
        f = float(s)
    except Exception:
        return None, "não parece um valor"
    if f != f or f in (float("inf"), float("-inf")):
        return None, "valor inválido"
    if f < 0:
        return None, "valor negativo"
    if f > 999999999999.99:
        return None, "valor acima do teto"
    return round(f, 2), None


def _status_do_texto(v) -> Tuple[Optional[str], Optional[str]]:
    s = str(v or "").strip().lower()
    if not s:
        return None, None
    if s in _STATUS_POR_LABEL:
        return _STATUS_POR_LABEL[s], None
    if s in STATUS_LABEL:                      # aceita o valor cru também
        return s, None
    if s.startswith("aguard"):
        return "enviado", None
    return None, "status desconhecido"


def ler_planilha_lote(caminho: str, max_linhas: int = 3000) -> Dict:
    """Lê o .xlsx preenchido e devolve {linhas, erro_fatal}. Não decide nada — só transcreve.

    Cada linha vira {n, ancora, tipo, ref, categoria, item, fornecedor, valor, forma, venc, status,
    problemas[]}. Nada de exceção pra fora: planilha torta é MENSAGEM, não erro 500."""
    try:
        wb = load_workbook(caminho, data_only=True, read_only=True)
    except Exception as e:
        return {"linhas": [], "erro_fatal": "não consegui abrir a planilha (%s). Ela precisa ser "
                                            ".xlsx, o mesmo arquivo que você baixou." % type(e).__name__}
    try:
        ws = wb[ABA] if ABA in wb.sheetnames else wb[wb.sheetnames[0]]
        cabecalho, dados = None, []
        for linha in ws.iter_rows(values_only=True):
            if linha is None:
                continue
            celulas = ["" if c is None else c for c in linha]
            if cabecalho is None:
                brutos = [str(c).strip().upper() for c in celulas]
                if "ITEM" in brutos and ("VALOR (R$)" in brutos or "REFERÊNCIA" in brutos):
                    cabecalho = brutos
                continue
            if not any(str(c).strip() for c in celulas):
                continue
            dados.append(celulas)
            if len(dados) > max_linhas:
                return {"linhas": [], "erro_fatal":
                        "a planilha tem mais de %d linhas — mande em partes." % max_linhas}
    finally:
        try:
            wb.close()
        except Exception:
            pass

    if cabecalho is None:
        return {"linhas": [], "erro_fatal": "não achei o cabeçalho da planilha. Use o modelo que "
                                            "você baixou aqui, sem apagar a linha de títulos."}

    def col(nome, padrao=None):
        return cabecalho.index(nome) if nome in cabecalho else padrao

    i_cat, i_item = col("CATEGORIA"), col("ITEM")
    i_forn, i_val = col("FORNECEDOR"), col("VALOR (R$)")
    i_forma, i_venc = col("FORMA DE PAGAMENTO"), col("VENCIMENTO")
    i_st, i_ref = col("STATUS"), col("REFERÊNCIA")

    def celula(linha, i):
        return linha[i] if (i is not None and i < len(linha)) else None

    linhas = []
    for n, bruta in enumerate(dados, start=1):
        problemas = []
        ancora = str(celula(bruta, i_ref) or "").strip()
        m = _RE_ANCORA.match(ancora)
        tipo, ref = (m.group(1), m.group(2)) if m else ("novo", None)
        if ancora and not m:
            problemas.append("referência ilegível — a linha entra como livre")
        valor, err_v = valor_do_texto(celula(bruta, i_val))
        if err_v:
            problemas.append(err_v)
        status, err_s = _status_do_texto(celula(bruta, i_st))
        if err_s:
            problemas.append(err_s)
        venc_bruto = celula(bruta, i_venc)
        venc = _data(venc_bruto)
        if venc_bruto not in (None, "") and venc is None:
            problemas.append("data de vencimento ilegível")
        linhas.append({
            "n": n, "ancora": ancora, "tipo": tipo, "ref": ref,
            "categoria": _texto(celula(bruta, i_cat)),
            "item": _texto(celula(bruta, i_item)),
            "fornecedor": _texto(celula(bruta, i_forn))[:200],
            "valor": valor,
            "forma": _texto(celula(bruta, i_forma))[:120],
            "venc": venc,
            "status": status,
            "problemas": problemas,
        })
    return {"linhas": linhas, "erro_fatal": None}


def conferir_lote(linhas: List[Dict], itens_vivos: Dict[str, Dict],
                  lancamentos_atuais: Dict[str, Dict]) -> Dict:
    """O que VAI acontecer, em português, ANTES de gravar qualquer coisa.

    `itens_vivos`: {id do item: item} — o quantitativo de AGORA (é ele que dá o retrato, nº7).
    `lancamentos_atuais`: {id do lançamento: linha} — pra saber o que muda de verdade.
    Devolve {acoes, resumo, avisos}: `acoes` é o que o servidor grava depois do OK.
    """
    acoes, avisos = [], []
    n_novo = n_atualiza = n_igual = n_ignorada = 0
    perdeu_origem = 0

    for l in linhas:
        preencheu = bool(l["fornecedor"] or l["forma"] or l["venc"] or l["status"]) or l["valor"] is not None
        tipo, ref = l["tipo"], l["ref"]

        if tipo == "l" and ref in lancamentos_atuais:
            atual = lancamentos_atuais[ref]
            campos = {}
            if l["valor"] != (None if atual.get("valor") is None else round(float(atual["valor"]), 2)):
                campos["valor"] = l["valor"]
            if l["fornecedor"] != (atual.get("fornecedor") or ""):
                campos["fornecedor"] = l["fornecedor"]
            if l["forma"] != (atual.get("forma_pagamento") or ""):
                campos["forma_pagamento"] = l["forma"]
            if l["status"] and l["status"] != atual.get("status"):
                campos["status"] = l["status"]
            if l["venc"] and (atual.get("venc_tipo") != "data" or str(atual.get("venc_data") or "")[:10] != l["venc"].isoformat()):
                campos["venc_tipo"] = "data"
                campos["venc_data"] = l["venc"].isoformat()
            # apagar o valor de linha PAGA exige o status junto (mesma regra da tela)
            if campos.get("valor") is None and "valor" in campos and atual.get("status") == "pago" and "status" not in campos:
                campos["status"] = "contratado"
            if campos:
                # a linha ATUAL vai junto: o servidor normaliza a linha MESCLADA (a coerência —
                # fase OU data, pago exige valor — é julgada na linha inteira, como no PATCH da tela)
                acoes.append({"acao": "atualiza", "id": ref, "campos": campos, "atual": atual,
                              "item": l["item"], "n": l["n"]})
                n_atualiza += 1
            else:
                n_igual += 1
            continue

        if tipo == "l":
            avisos.append("linha %d: o lançamento não existe mais no projeto — entra como novo" % l["n"])
            tipo = "novo"

        if not preencheu:
            n_ignorada += 1
            continue

        if tipo == "i" and ref in itens_vivos:
            it = itens_vivos[ref]
            acoes.append({"acao": "cria", "n": l["n"], "item": l["item"], "corpo": {
                "escopo": "obra", "origem": "quantitativo", "origem_ref_id": ref,
                "categoria": l["categoria"] or (it.get("discipline") or "").strip() or "Sem categoria",
                "fornecedor": l["fornecedor"], "valor": l["valor"],
                "forma_pagamento": l["forma"], "status": l["status"] or "cotado",
                **_vencimento(l, it),
            }})
            n_novo += 1
            continue

        if tipo == "i":
            # 🪤 o /add-file recria os ids: a planilha é de antes de re-subir a planta
            perdeu_origem += 1
            avisos.append("linha %d (%s): este item não está mais no quantitativo com a mesma "
                          "identidade — entra como linha livre, com o valor que você preencheu"
                          % (l["n"], l["item"][:60] or "sem nome"))
        acoes.append({"acao": "cria", "n": l["n"], "item": l["item"], "corpo": {
            "escopo": "obra", "origem": "livre",
            "descricao": l["item"] or "Lançamento sem descrição",
            "categoria": l["categoria"] or "Sem categoria",
            "fornecedor": l["fornecedor"], "valor": l["valor"],
            "forma_pagamento": l["forma"], "status": l["status"] or "cotado",
            **_vencimento(l, None),
        }})
        n_novo += 1

    sem_valor = sum(1 for l in linhas if l["valor"] is None and (l["fornecedor"] or l["status"] or l["forma"]))
    problemas = [("linha %d: %s" % (l["n"], "; ".join(l["problemas"]))) for l in linhas if l["problemas"]]
    return {
        "acoes": acoes,
        "avisos": avisos[:40] + problemas[:40],
        "resumo": {
            "linhas": len(linhas), "cria": n_novo, "atualiza": n_atualiza,
            "sem_mudanca": n_igual, "ignoradas": n_ignorada,
            "sem_valor": sem_valor, "sem_origem": perdeu_origem,
            "com_problema": len(problemas),
        },
        "frase": _frase(n_novo, n_atualiza, n_igual, n_ignorada, perdeu_origem),
    }


def _vencimento(l: Dict, item: Optional[Dict]) -> Dict:
    """Data fixa quando a pessoa preencheu; senão, amarrado à fase com o nome da categoria
    (o mesmo padrão do lançamento feito na tela)."""
    if l.get("venc"):
        return {"venc_tipo": "data", "venc_data": l["venc"].isoformat()}
    fase = l.get("categoria") or ((item or {}).get("discipline") or "").strip()
    return {"venc_tipo": "fase", "venc_fase": fase, "venc_quando": "inicio"}


def _frase(cria: int, atualiza: int, igual: int, ignorada: int, sem_origem: int) -> str:
    partes = []
    if cria:
        partes.append("%d lançamento%s novo%s" % (cria, "s" if cria > 1 else "", "s" if cria > 1 else ""))
    if atualiza:
        partes.append("%d atualizad%s" % (atualiza, "os" if atualiza > 1 else "o"))
    if igual:
        partes.append("%d sem mudança" % igual)
    if ignorada:
        partes.append("%d em branco (ignorada%s)" % (ignorada, "s" if ignorada > 1 else ""))
    if not partes:
        return "Não achei nada preenchido nesta planilha."
    frase = ", ".join(partes[:-1]) + (" e " + partes[-1] if len(partes) > 1 else partes[0])
    if sem_origem:
        frase += " — %d sem vínculo com o quantitativo" % sem_origem
    return frase[0].upper() + frase[1:]


def total_das_acoes(acoes: List[Dict]) -> str:
    """Quanto some de dinheiro nesta subida — pro arquiteto conferir antes de confirmar."""
    soma = 0.0
    tem = False
    for a in acoes:
        v = (a.get("corpo") or a.get("campos") or {}).get("valor")
        if v is not None:
            soma += float(v)
            tem = True
    return _brl(round(soma, 2)) if tem else "—"
