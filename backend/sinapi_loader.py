# -*- coding: utf-8 -*-
"""Carrega composições SINAPI do XLSX oficial (aba CCD) pra tabela
`sinapi_composicao`. Usado como vocabulário-âncora pro classificador LLM
e como referência pra caderno de compras.

Uso:
    python sinapi_loader.py /caminho/SINAPI_Referencia_YYYY_MM.xlsx
"""
import json
import os
import sys
import time
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


SUPABASE_URL = "https://kqjabzwgbfuivzlcfvvu.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImtxamFiendnYmZ1aXZ6bGNmdnZ1Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzYwMDg5NzcsImV4cCI6MjA5MTU4NDk3N30.48xSenZlDV0LfD94ZxwGvX41Kf9Je2n-ouZpJrrCSKI"


def _supabase_bulk_upsert(table: str, rows: list[dict], on_conflict: str,
                          batch_size: int = 200) -> int:
    """Upsert em lote via Supabase REST. Retorna número de linhas enviadas."""
    total = 0
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        try:
            url = f"{SUPABASE_URL}/rest/v1/{table}?on_conflict={on_conflict}"
            body = json.dumps(batch).encode("utf-8")
            req = urllib.request.Request(url, data=body, method="POST")
            req.add_header("apikey", SUPABASE_KEY)
            req.add_header("Authorization", f"Bearer {SUPABASE_KEY}")
            req.add_header("Content-Type", "application/json")
            req.add_header("Prefer", "resolution=merge-duplicates,return=minimal")
            urllib.request.urlopen(req, timeout=30)
            total += len(batch)
            print(f"  upsert {total}/{len(rows)}...")
        except Exception as e:
            print(f"  ERRO batch {i}-{i+batch_size}: {e}")
    return total


def parse_composicoes(xlsx_path: str, sync_month: str = "") -> list[dict]:
    """Extrai composições SINAPI da aba Analítico.

    A aba CCD tem os códigos zerados nessa versão do XLSX oficial; a aba
    Analítico preserva o código real na col B e as linhas-mestre (a própria
    composição) vêm com col C='-', enquanto col C='COMPOSICAO'/'INSUMO'
    indica componente. A gente pega só as linhas-mestre.

    Layout Analítico (confirmado pela inspeção):
      Header linha 10:
        A: Grupo, B: Código da Composição, C: Tipo Item,
        D: Código do Item, E: Descrição, F: Unidade, G: Coeficiente, H: Situação
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Analítico" not in wb.sheetnames:
        raise ValueError(f"Aba 'Analítico' não encontrada em {xlsx_path}")
    ws = wb["Analítico"]

    rows_out = []
    seen_codes = set()
    for row in ws.iter_rows(min_row=11, max_col=8, values_only=True):
        if not row or all(c is None for c in row):
            continue
        grupo, codigo, tipo_item, _cod_item, descricao, unidade = (
            row[0], row[1], row[2], row[3], row[4], row[5]
        )
        # Mestre = linha onde Tipo Item é '-' (nula/traço). Componentes têm
        # 'COMPOSICAO' ou 'INSUMO' nesse campo.
        if tipo_item not in (None, "-", ""):
            continue
        if codigo is None or descricao is None:
            continue
        codigo_str = str(codigo).strip()
        if not codigo_str or codigo_str == "0":
            continue
        if not codigo_str.lstrip("-").isdigit():
            continue
        if codigo_str in seen_codes:
            continue
        seen_codes.add(codigo_str)
        rows_out.append({
            "codigo": codigo_str,
            "descricao": str(descricao).strip()[:400],
            "unidade": str(unidade).strip()[:20] if unidade else None,
            "sync_month": sync_month or None,
        })
    wb.close()
    return rows_out


# alias para retrocompatibilidade com o nome anterior
parse_ccd = parse_composicoes


def main():
    if len(sys.argv) < 2:
        print("uso: python sinapi_loader.py <caminho_xlsx> [sync_month YYYY-MM-DD]")
        sys.exit(1)
    path = sys.argv[1]
    if not os.path.exists(path):
        print(f"arquivo não encontrado: {path}")
        sys.exit(1)

    # Infere sync_month do nome do arquivo se não for passado
    sync_month = sys.argv[2] if len(sys.argv) > 2 else ""
    if not sync_month:
        import re
        m = re.search(r"(\d{4})[_-](\d{2})", Path(path).name)
        if m:
            sync_month = f"{m.group(1)}-{m.group(2)}-01"

    print(f"Lendo {path} (sync_month={sync_month})...")
    t0 = time.time()
    rows = parse_ccd(path, sync_month=sync_month)
    print(f"Extraídas {len(rows)} composições em {time.time()-t0:.1f}s")

    print("Upsertando em sinapi_composicao...")
    total = _supabase_bulk_upsert("sinapi_composicao", rows, "codigo")
    print(f"Done — {total} linhas enviadas.")


if __name__ == "__main__":
    main()
