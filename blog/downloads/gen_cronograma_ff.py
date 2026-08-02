# -*- coding: utf-8 -*-
"""Modelo de CRONOGRAMA FÍSICO-FINANCEIRO — v2 (02/08/2026).

Por que refazer: a v1 (gen_modelos_xlsx.modelo_cronograma) só planejava —
tinha PREVISTO e curva S, mas nenhum lugar pra registrar o que a obra
REALMENTE executou. Cronograma físico-financeiro serve justamente pra
comparar previsto × realizado mês a mês; sem medição ele é meio documento.

O que a v2 acrescenta, com base no estudo de 02/08:
  · PESO por custo — o avanço físico de cada etapa pondera pelo VALOR dela,
    não pela média simples (erro clássico apontado pelo OrçaFascio).
  · Aba MEDIÇÃO — o % executado de cada etapa por mês (é o que o agente
    financiador pede pra liberar parcela).
  · RESUMO — desembolso previsto × realizado, acumulados, desvio e a curva S
    com DUAS linhas (planejado e real).
  · Meses em data real (mm/aaaa), não "M1".
  · Validação: a soma da distribuição de cada etapa tem que fechar 100%.

🚫 Nenhum preço nosso: TODO valor é digitado pelo usuário (célula amarela).
   Regra dura nº5 — o AI.arq não precifica.

Rodar da raiz: python blog/downloads/gen_cronograma_ff.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

AQUI = os.path.dirname(os.path.abspath(__file__))

# Mesmo padrão visual dos outros modelos (gen_modelos_xlsx.py)
F = 'Calibri'
TIT = Font(name=F, bold=True, size=14, color='0F172A')
HDR = Font(name=F, bold=True, size=10, color='0F172A')
HDRW = Font(name=F, bold=True, size=10, color='FFFFFF')
NOR = Font(name=F, size=10)
NEG = Font(name=F, bold=True, size=10, color='0F172A')
NOTA = Font(name=F, size=9, italic=True, color='64748B')
P_HDR = PatternFill('solid', fgColor='E2E8F0')
P_ESC = PatternFill('solid', fgColor='0F172A')
P_YEL = PatternFill('solid', fgColor='FFF0A6')      # amarelo = você preenche
P_TOT = PatternFill('solid', fgColor='E8ECF4')
P_IND = PatternFill('solid', fgColor='EEF2FF')
_S = Side(style='thin', color='D0D5DD')
BD = Border(left=_S, right=_S, top=_S, bottom=_S)
AL = Alignment(horizontal='left', vertical='center', wrap_text=True)
AC = Alignment(horizontal='center', vertical='center', wrap_text=True)

MESES = 12
ETAPAS = [
    "Serviços preliminares", "Movimento de terra", "Fundações", "Estrutura",
    "Vedações verticais", "Cobertura", "Impermeabilizações", "Esquadrias",
    "Revestimentos internos", "Revestimentos externos", "Forros", "Pisos",
    "Pinturas", "Inst. hidrossanitárias", "Inst. elétricas", "Complementares e limpeza",
]
L0 = 5                      # 1ª linha de dados
LF = L0 + len(ETAPAS) - 1   # última


def _titulo(ws, texto, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=texto); c.font = TIT; c.alignment = AL
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=sub); c.font = NOTA; c.alignment = AL
    ws.row_dimensions[2].height = 28


def _mes_hdr(ws, linha, col_ini, formula_base):
    """Cabeçalho dos 12 meses. O 1º puxa a data digitada no RESUMO; os outros
    somam 1 mês — assim o cronograma inteiro anda ao mudar a data de início."""
    for m in range(MESES):
        col = col_ini + m
        ws.column_dimensions[get_column_letter(col)].width = 11
        f = (formula_base if m == 0
             else f'=IF({get_column_letter(col-1)}{linha}="","",EDATE({get_column_letter(col-1)}{linha},1))')
        c = ws.cell(row=linha, column=col, value=f)
        c.font = HDR; c.fill = P_HDR; c.alignment = AC; c.border = BD
        c.number_format = 'mmm/aa'


def aba_como_usar(wb):
    ws = wb.create_sheet('COMO USAR')
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 108
    _titulo(ws, 'Cronograma físico-financeiro — como usar este modelo',
            'Preencha SÓ as células amarelas. Todo o resto é fórmula e se atualiza sozinho.', 2)
    passos = [
        ('1. Aba PLANEJADO', 'Liste as etapas da obra e o VALOR de cada uma (do seu orçamento). '
         'Depois distribua, em %, quanto de cada etapa você prevê executar por mês. '
         'A coluna "soma" precisa fechar 100% — ela fica vermelha enquanto não fechar.'),
        ('2. Aba MEDIÇÃO', 'A cada mês, registre o % que foi REALMENTE executado de cada etapa. '
         'É esse número que o banco/financiador usa pra liberar parcela.'),
        ('3. Aba RESUMO', 'Não preencha nada aqui além da data de início. Ela mostra desembolso '
         'previsto × realizado por mês, o acumulado, o desvio e a curva S.'),
        ('Sobre o PESO', 'O avanço físico da obra NÃO é a média das etapas: uma etapa de '
         'R$ 200 mil pesa mais que uma de R$ 5 mil. A coluna PESO calcula essa proporção pelo '
         'valor, e o avanço total já sai ponderado.'),
        ('Sobre a CURVA S', 'É o desembolso acumulado ao longo do tempo. Começa devagar, '
         'acelera no miolo da obra e desacelera no fim — daí o formato de S. Duas linhas: '
         'o que estava planejado e o que aconteceu.'),
        ('Os valores são SEUS', 'Este modelo não sugere preço nenhum. Os valores vêm do seu '
         'orçamento; o AI.arq mede quantidades e não precifica.'),
    ]
    r = 4
    for tit, txt in passos:
        c = ws.cell(row=r, column=2, value=tit); c.font = NEG; c.alignment = AL
        r += 1
        c = ws.cell(row=r, column=2, value=txt); c.font = NOR; c.alignment = AL
        ws.row_dimensions[r].height = 30
        r += 2
    c = ws.cell(row=r + 1, column=2, value='Modelo gratuito do AI.arq · ai.arq.br')
    c.font = NOTA


def aba_planejado(wb):
    ws = wb.create_sheet('PLANEJADO')
    ncols = 4 + MESES + 1
    _titulo(ws, 'PLANEJADO — valor de cada etapa e distribuição no tempo',
            'Amarelo = você preenche. VALOR vem do seu orçamento; a distribuição é em % de cada etapa por mês.',
            ncols)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 3
    for i, (nome, larg) in enumerate([('ETAPA', 30), ('VALOR (R$)', 16), ('PESO', 10), ('', 3)], start=1):
        c = ws.cell(row=4, column=i, value=nome)
        c.font = HDR; c.fill = P_HDR; c.alignment = AC; c.border = BD
    _mes_hdr(ws, 4, 5, '=IF(RESUMO!$B$3="","",RESUMO!$B$3)')
    col_soma = 5 + MESES
    ws.column_dimensions[get_column_letter(col_soma)].width = 11
    c = ws.cell(row=4, column=col_soma, value='SOMA')
    c.font = HDR; c.fill = P_HDR; c.alignment = AC; c.border = BD

    for i, etapa in enumerate(ETAPAS):
        r = L0 + i
        ws.cell(row=r, column=1, value=etapa).font = NOR
        v = ws.cell(row=r, column=2); v.fill = P_YEL; v.number_format = 'R$ #,##0.00'
        p = ws.cell(row=r, column=3, value=f'=IF($B${LF+1}=0,"",B{r}/$B${LF+1})')
        p.number_format = '0.0%'; p.font = NOR
        for m in range(MESES):
            cel = ws.cell(row=r, column=5 + m)
            cel.fill = P_YEL; cel.number_format = '0%'; cel.font = NOR; cel.border = BD
        s = ws.cell(row=r, column=col_soma,
                    value=f'=IF(SUM(E{r}:{get_column_letter(4+MESES)}{r})=0,"",'
                          f'SUM(E{r}:{get_column_letter(4+MESES)}{r}))')
        s.number_format = '0%'; s.font = NEG
        for cc in (1, 2, 3, col_soma):
            ws.cell(row=r, column=cc).border = BD

    # Total + avanço físico previsto por mês (JÁ PONDERADO pelo valor da etapa)
    t = LF + 1
    ws.cell(row=t, column=1, value='TOTAL DA OBRA').font = HDR
    tv = ws.cell(row=t, column=2, value=f'=SUM(B{L0}:B{LF})')
    tv.font = HDR; tv.number_format = 'R$ #,##0.00'
    for cc in range(1, ncols + 1):
        ws.cell(row=t, column=cc).fill = P_TOT
        ws.cell(row=t, column=cc).border = BD
    t2 = t + 1
    ws.cell(row=t2, column=1, value='AVANÇO FÍSICO PREVISTO NO MÊS').font = HDR
    ws.cell(row=t2, column=1).alignment = AL
    for m in range(MESES):
        col = get_column_letter(5 + m)
        c = ws.cell(row=t2, column=5 + m,
                    value=f'=IF($B${t}=0,"",SUMPRODUCT({col}{L0}:{col}{LF},$B${L0}:$B${LF})/$B${t})')
        c.number_format = '0.0%'; c.font = NEG; c.fill = P_IND; c.border = BD
    ws.cell(row=t2, column=1).fill = P_IND
    ws.cell(row=t2, column=1).border = BD

    # A soma de cada etapa tem que fechar 100% — fica vermelha enquanto não fecha
    faixa = f'{get_column_letter(col_soma)}{L0}:{get_column_letter(col_soma)}{LF}'
    ws.conditional_formatting.add(faixa, CellIsRule(
        operator='notEqual', formula=['1'],
        fill=PatternFill('solid', fgColor='FFD9D9'),
        font=Font(name=F, size=10, color='B91C1C', bold=True)))
    ws.freeze_panes = 'E5'
    ws.cell(row=t2 + 2, column=1, value=(
        'A linha AVANÇO FÍSICO PREVISTO já pondera pelo valor: etapa cara pesa mais. '
        'A coluna SOMA fica vermelha até a distribuição da etapa fechar 100%.')).font = NOTA
    ws.merge_cells(start_row=t2 + 2, start_column=1, end_row=t2 + 2, end_column=ncols)


def aba_medicao(wb):
    ws = wb.create_sheet('MEDIÇÃO')
    ncols = 4 + MESES + 1
    _titulo(ws, 'MEDIÇÃO — o que foi realmente executado',
            'A cada mês, preencha o % executado de cada etapa. É o número que o financiador usa pra liberar parcela.',
            ncols)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 3
    for i, nome in enumerate(['ETAPA', 'VALOR (R$)', 'EXECUT.', ''], start=1):
        c = ws.cell(row=4, column=i, value=nome)
        c.font = HDR; c.fill = P_HDR; c.alignment = AC; c.border = BD
    _mes_hdr(ws, 4, 5, '=IF(PLANEJADO!E4="","",PLANEJADO!E4)')
    col_soma = 5 + MESES
    ws.column_dimensions[get_column_letter(col_soma)].width = 11
    c = ws.cell(row=4, column=col_soma, value='ACUM.')
    c.font = HDR; c.fill = P_HDR; c.alignment = AC; c.border = BD

    for i, etapa in enumerate(ETAPAS):
        r = L0 + i
        ws.cell(row=r, column=1, value=f'=IF(PLANEJADO!A{r}="","",PLANEJADO!A{r})').font = NOR
        v = ws.cell(row=r, column=2, value=f'=IF(PLANEJADO!B{r}="","",PLANEJADO!B{r})')
        v.number_format = 'R$ #,##0.00'; v.font = NOR
        e = ws.cell(row=r, column=3,
                    value=f'=IF(SUM(E{r}:{get_column_letter(4+MESES)}{r})=0,"",'
                          f'SUM(E{r}:{get_column_letter(4+MESES)}{r}))')
        e.number_format = '0%'; e.font = NEG
        for m in range(MESES):
            cel = ws.cell(row=r, column=5 + m)
            cel.fill = P_YEL; cel.number_format = '0%'; cel.font = NOR; cel.border = BD
        s = ws.cell(row=r, column=col_soma, value=f'=IF(C{r}="","",C{r})')
        s.number_format = '0%'; s.font = NEG
        for cc in (1, 2, 3, col_soma):
            ws.cell(row=r, column=cc).border = BD

    t = LF + 1
    ws.cell(row=t, column=1, value='AVANÇO FÍSICO REALIZADO NO MÊS').font = HDR
    ws.cell(row=t, column=1).alignment = AL
    for m in range(MESES):
        col = get_column_letter(5 + m)
        c = ws.cell(row=t, column=5 + m,
                    value=f'=IF(PLANEJADO!$B${LF+1}=0,"",'
                          f'SUMPRODUCT({col}{L0}:{col}{LF},$B${L0}:$B${LF})/PLANEJADO!$B${LF+1})')
        c.number_format = '0.0%'; c.font = NEG; c.fill = P_IND; c.border = BD
    for cc in (1, 2, 3):
        ws.cell(row=t, column=cc).fill = P_IND
        ws.cell(row=t, column=cc).border = BD
    ws.freeze_panes = 'E5'
    ws.cell(row=t + 2, column=1, value=(
        'Etapa não iniciada fica em branco. A coluna EXECUT. mostra o acumulado da etapa; '
        'passar de 100% indica erro de medição.')).font = NOTA
    ws.merge_cells(start_row=t + 2, start_column=1, end_row=t + 2, end_column=ncols)


def aba_resumo(wb):
    ws = wb.create_sheet('RESUMO')
    ncols = 2 + MESES
    _titulo(ws, 'RESUMO — desembolso, desvio e curva S',
            'A única célula que você preenche aqui é a DATA DE INÍCIO (amarela). O resto vem das outras abas.',
            ncols)
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 16
    ws.cell(row=3, column=1, value='DATA DE INÍCIO DA OBRA').font = HDR
    d = ws.cell(row=3, column=2); d.fill = P_YEL; d.number_format = 'dd/mm/aaaa'
    d.border = BD
    ws.cell(row=3, column=1).border = BD

    LIN = {'hdr': 5, 'prev': 6, 'prev_ac': 7, 'real': 8, 'real_ac': 9,
           'desvio': 10, 'fis_prev': 12, 'fis_real': 13}
    c = ws.cell(row=LIN['hdr'], column=1, value='MÊS')
    c.font = HDRW; c.fill = P_ESC; c.alignment = AC; c.border = BD
    c = ws.cell(row=LIN['hdr'], column=2, value='TOTAL')
    c.font = HDRW; c.fill = P_ESC; c.alignment = AC; c.border = BD
    _mes_hdr(ws, LIN['hdr'], 3, '=IF($B$3="","",$B$3)')
    for m in range(MESES):
        cel = ws.cell(row=LIN['hdr'], column=3 + m)
        cel.font = HDRW; cel.fill = P_ESC

    plan_last = get_column_letter(4 + MESES)
    linhas = [
        (LIN['prev'], 'DESEMBOLSO PREVISTO', 'R$ #,##0.00',
         lambda m: f'=SUMPRODUCT(PLANEJADO!{get_column_letter(5+m)}${L0}:{get_column_letter(5+m)}${LF},'
                   f'PLANEJADO!$B${L0}:$B${LF})'),
        (LIN['prev_ac'], 'PREVISTO ACUMULADO', 'R$ #,##0.00', None),
        (LIN['real'], 'DESEMBOLSO REALIZADO', 'R$ #,##0.00',
         lambda m: f'=SUMPRODUCT(MEDIÇÃO!{get_column_letter(5+m)}${L0}:{get_column_letter(5+m)}${LF},'
                   f'MEDIÇÃO!$B${L0}:$B${LF})'),
        (LIN['real_ac'], 'REALIZADO ACUMULADO', 'R$ #,##0.00', None),
        (LIN['desvio'], 'DESVIO (realizado − previsto)', 'R$ #,##0.00', None),
        (LIN['fis_prev'], 'AVANÇO FÍSICO PREVISTO (acum.)', '0.0%', None),
        (LIN['fis_real'], 'AVANÇO FÍSICO REALIZADO (acum.)', '0.0%', None),
    ]
    for row, rot, fmt, gen in linhas:
        c = ws.cell(row=row, column=1, value=rot); c.font = NEG; c.alignment = AL; c.border = BD
        for m in range(MESES):
            col = 3 + m
            L = get_column_letter(col)
            Lp = get_column_letter(col - 1)
            if gen:
                v = gen(m)
            elif row == LIN['prev_ac']:
                v = f'={L}{LIN["prev"]}' if m == 0 else f'={Lp}{row}+{L}{LIN["prev"]}'
            elif row == LIN['real_ac']:
                v = f'={L}{LIN["real"]}' if m == 0 else f'={Lp}{row}+{L}{LIN["real"]}'
            elif row == LIN['desvio']:
                v = f'={L}{LIN["real_ac"]}-{L}{LIN["prev_ac"]}'
            elif row == LIN['fis_prev']:
                v = f'=IF($B${LIN["prev"]}=0,"",{L}{LIN["prev_ac"]}/$B${LIN["prev"]})'
            else:
                v = f'=IF($B${LIN["prev"]}=0,"",{L}{LIN["real_ac"]}/$B${LIN["prev"]})'
            cel = ws.cell(row=row, column=col, value=v)
            cel.number_format = fmt; cel.font = NOR; cel.border = BD
        tot = ws.cell(row=row, column=2)
        tot.border = BD; tot.font = NEG; tot.fill = P_TOT
        if row in (LIN['prev'], LIN['real']):
            tot.value = f'=SUM(C{row}:{get_column_letter(2+MESES)}{row})'
            tot.number_format = 'R$ #,##0.00'

    # Indicadores
    r = LIN['fis_real'] + 2
    ws.cell(row=r, column=1, value='INDICADORES').font = HDR
    ind = [
        ('Investimento total (do seu orçamento)', f'=PLANEJADO!B{LF+1}', 'R$ #,##0.00'),
        ('Já desembolsado', f'=B{LIN["real"]}', 'R$ #,##0.00'),
        ('Falta desembolsar', f'=PLANEJADO!B{LF+1}-B{LIN["real"]}', 'R$ #,##0.00'),
        ('Avanço físico realizado até agora',
         f'=IF(PLANEJADO!B{LF+1}=0,"",B{LIN["real"]}/PLANEJADO!B{LF+1})', '0.0%'),
    ]
    for i, (rot, f_, fmt) in enumerate(ind, start=1):
        c = ws.cell(row=r + i, column=1, value=rot); c.font = NOR; c.alignment = AL; c.border = BD
        v = ws.cell(row=r + i, column=2, value=f_)
        v.number_format = fmt; v.font = NEG; v.fill = P_IND; v.border = BD

    # Curva S — planejado × real
    ch = LineChart()
    ch.title = 'Curva S — desembolso acumulado'
    ch.style = 2
    ch.height, ch.width = 9, 22
    ch.y_axis.title = 'R$ acumulado'
    dados = Reference(ws, min_col=2, max_col=2 + MESES,
                      min_row=LIN['prev_ac'], max_row=LIN['prev_ac'])
    ch.add_data(dados, titles_from_data=True, from_rows=True)
    dados2 = Reference(ws, min_col=2, max_col=2 + MESES,
                       min_row=LIN['real_ac'], max_row=LIN['real_ac'])
    ch.add_data(dados2, titles_from_data=True, from_rows=True)
    cats = Reference(ws, min_col=3, max_col=2 + MESES, min_row=LIN['hdr'], max_row=LIN['hdr'])
    ch.set_categories(cats)
    ws.add_chart(ch, f'A{r + len(ind) + 3}')
    ws.cell(row=r + len(ind) + 2, column=1, value=(
        'A linha do realizado só cresce conforme você preenche a aba MEDIÇÃO.')).font = NOTA


def main():
    wb = Workbook()
    wb.remove(wb.active)
    aba_como_usar(wb)
    aba_planejado(wb)
    aba_medicao(wb)
    aba_resumo(wb)
    wb._sheets = [wb['COMO USAR'], wb['PLANEJADO'], wb['MEDIÇÃO'], wb['RESUMO']]
    destino = os.path.join(AQUI, 'cronograma-fisico-financeiro-modelo.xlsx')
    wb.save(destino)
    print('gerado:', destino, os.path.getsize(destino) // 1024, 'KB')


if __name__ == '__main__':
    main()
