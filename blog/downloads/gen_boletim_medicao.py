# -*- coding: utf-8 -*-
"""Modelo de BOLETIM DE MEDIÇÃO de obra — planilha Excel (29/08/2026).

Nasceu do estudo do blog de 29/08: "boletim de medição" é a LACUNA nº1 —
zero menção nos 28 posts, autocomplete cheio de "modelo / excel / download",
e a SERP inteira é lead magnet de terceiro. O formato que traz gente é
ARQUIVO PRA BAIXAR (provado pelo memorial e pelo cronograma).

🚫 REGRA DURA nº5 GOVERNANDO O ARQUIVO: o boletim de mercado costuma ter
   coluna de preço unitário e valor a pagar. O NOSSO sai em QUANTIDADE —
   prevista × executada no período × acumulada × saldo. Quem precifica é o
   orçamentista; a planilha diz isso NELA MESMA (nota na aba de instruções).

🔑 Lições da casa embutidas:
   - linha de CONFERÊNCIA em toda aba (acumulado > previsto = justificar —
     o boletim equivalente de "quadro que não fecha consigo mesmo");
   - MEMÓRIA DE CÁLCULO como aba própria: número sem procedência é o defeito
     nº1 das planilhas que o motor lê (96% dos medidos do AI.arq carregam
     origem da medição — o modelo pede o mesmo de quem preenche à mão);
   - células que o usuário preenche em AMARELO (mesma convenção do quadro
     de áreas e do cronograma).

Rodar da raiz: python blog/downloads/gen_boletim_medicao.py
"""
import os

from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill,
                             Protection, Side)
from openpyxl.utils import get_column_letter

AQUI = os.path.dirname(os.path.abspath(__file__))
SAIDA = os.path.join(AQUI, "boletim-de-medicao-modelo.xlsx")

# ── paleta do AI.arq (a mesma do cronograma e do quadro de áreas) ───────────
INDIGO = "4F46E5"
INDIGO_CLARO = "6366F1"
CINZA_CLARO = "F3F4F6"
AMARELO = "FEF3C7"          # célula que o usuário preenche
FINO = Side(style="thin", color="D1D5DB")
BORDA = Border(left=FINO, right=FINO, top=FINO, bottom=FINO)

F_TITULO = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
F_CAB = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
F_ROTULO = Font(name="Calibri", bold=True, size=10)
F_NOTA = Font(name="Calibri", italic=True, size=9, color="6B7280")
F_TOTAL = Font(name="Calibri", bold=True, size=11)


def _titulo(ws, texto, n_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=texto)
    c.font = F_TITULO
    c.fill = PatternFill("solid", fgColor=INDIGO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28


def _cab_colunas(ws, linha, colunas, larguras):
    for j, (nome, larg) in enumerate(zip(colunas, larguras), start=1):
        cel = ws.cell(row=linha, column=j, value=nome)
        cel.font = F_CAB
        cel.fill = PatternFill("solid", fgColor=INDIGO_CLARO)
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = BORDA
        ws.column_dimensions[get_column_letter(j)].width = larg
    ws.row_dimensions[linha].height = 30


def _campo(ws, r, c, rotulo, largura_valor=3, amarelo=True):
    ws.cell(row=r, column=c, value=rotulo).font = F_ROTULO
    ws.merge_cells(start_row=r, start_column=c + 1,
                   end_row=r, end_column=c + largura_valor)
    cel = ws.cell(row=r, column=c + 1)
    if amarelo:
        cel.fill = PatternFill("solid", fgColor=AMARELO)
    cel.border = BORDA


wb = Workbook()

# ═══ ABA 1 — BOLETIM DE MEDIÇÃO (o período) ═════════════════════════════════
ws = wb.active
ws.title = "BOLETIM"
COLS = ["Item", "Descrição do serviço", "Un",
        "Qtd prevista\n(contratada)", "Acumulada\nanterior",
        "Executada\nno período", "Acumulada\natual", "Saldo",
        "% exec.", "Memória de cálculo (ref.)"]
LARG = [7, 42, 7, 13, 12, 12, 12, 11, 9, 22]
_titulo(ws, "BOLETIM DE MEDIÇÃO — QUANTIDADES", len(COLS))

# cabeçalho da obra (linhas 2-5): tudo amarelo = preencher
_campo(ws, 2, 1, "Obra:", 4)
_campo(ws, 2, 7, "BM nº:", 1)
_campo(ws, 2, 9, "Data:", 1)
_campo(ws, 3, 1, "Contratante:", 4)
_campo(ws, 3, 7, "Período de:", 1)
_campo(ws, 3, 9, "até:", 1)
_campo(ws, 4, 1, "Executora:", 4)
_campo(ws, 4, 7, "RT (execução):", 3)

LIN_CAB = 6
_cab_colunas(ws, LIN_CAB, COLS, LARG)
INI, N = LIN_CAB + 1, 22
for i in range(N):
    r = INI + i
    for j in range(1, len(COLS) + 1):
        cel = ws.cell(row=r, column=j)
        cel.border = BORDA
        # amarelo = digitar; acumulada atual/saldo/% são FÓRMULA (não mexer)
        if j in (1, 2, 3, 4, 5, 6, 10):
            cel.fill = PatternFill("solid", fgColor=AMARELO)
    ws.cell(row=r, column=7, value=f'=IF(D{r}="","",E{r}+F{r})')
    ws.cell(row=r, column=8, value=f'=IF(D{r}="","",D{r}-G{r})')
    cel = ws.cell(row=r, column=9, value=f'=IF(OR(D{r}="",D{r}=0),"",G{r}/D{r})')
    cel.number_format = "0.0%"

# linha de conferência: acumulado acima do previsto tem que GRITAR
r = INI + N
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
ws.cell(row=r, column=1, value="CONFERÊNCIA").font = F_TOTAL
# 🪤 (D<>"") na multiplicação é obrigatório: G devolve "" (TEXTO) em linha
# vazia, e no Excel texto > número é SEMPRE verdadeiro — sem o guarda, o
# alerta gritava num boletim em branco.
cel = ws.cell(row=r, column=3, value=(
    f'=IF(SUMPRODUCT((D{INI}:D{r-1}<>"")*(G{INI}:G{r-1}>D{INI}:D{r-1}))>0,'
    f'"⚠ HÁ ITEM ACIMA DO PREVISTO — justificar na memória de cálculo",'
    f'"ok — nenhum item acima do previsto")'))
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
cel.font = F_TOTAL
cel.fill = PatternFill("solid", fgColor=CINZA_CLARO)

# assinaturas
r += 2
ws.cell(row=r, column=2, value="_______________________________").border = Border()
ws.cell(row=r, column=6, value="_______________________________")
ws.cell(row=r + 1, column=2, value="Quem mediu (fiscalização / contratante)").font = F_NOTA
ws.cell(row=r + 1, column=6, value="Quem executou (empresa / RT)").font = F_NOTA
ws.cell(row=r + 3, column=1, value=(
    "Este modelo registra QUANTIDADES. Preço unitário e valor a pagar são do "
    "contrato e de quem orça — não desta planilha.")).font = F_NOTA
ws.freeze_panes = f"A{LIN_CAB + 1}"

# ═══ ABA 2 — MEMÓRIA DE CÁLCULO ═════════════════════════════════════════════
ws = wb.create_sheet("MEMÓRIA DE CÁLCULO")
COLS2 = ["Ref.", "Item do BM", "Como o número foi obtido (trena, projeto, "
         "croqui, contagem…)", "Cálculo (ex.: 2 panos × 3,20 × 2,70)",
         "Resultado", "Un"]
LARG2 = [7, 26, 38, 34, 11, 7]
_titulo(ws, "MEMÓRIA DE CÁLCULO — cada número diz de onde veio", len(COLS2))
_cab_colunas(ws, 3, COLS2, LARG2)
for i in range(24):
    for j in range(1, len(COLS2) + 1):
        cel = ws.cell(row=4 + i, column=j)
        cel.border = BORDA
        cel.fill = PatternFill("solid", fgColor=AMARELO)
ws.cell(row=29, column=1, value=(
    "Número sem procedência não se defende em discussão de medição. A coluna "
    "'Ref.' liga cada linha daqui à coluna 'Memória de cálculo' do BOLETIM.")).font = F_NOTA
ws.freeze_panes = "A4"

# ═══ ABA 3 — COMO USAR ══════════════════════════════════════════════════════
ws = wb.create_sheet("COMO USAR")
ws.column_dimensions["A"].width = 100
_titulo(ws, "COMO USAR ESTE BOLETIM", 1)
PASSOS = [
    "",
    "1.  A coluna 'Qtd prevista (contratada)' vem do quantitativo/contrato — é a "
    "referência fixa do boletim inteiro. Preencha uma vez e não mude sem aditivo.",
    "",
    "2.  A cada período (mensal é o mais comum), preencha só 'Acumulada anterior' "
    "(copie a 'Acumulada atual' do BM anterior) e 'Executada no período'. "
    "Acumulada atual, Saldo e % são fórmula — não digite em cima.",
    "",
    "3.  TODA linha medida ganha uma referência na aba MEMÓRIA DE CÁLCULO: como o "
    "número foi obtido e a conta feita. É o que sustenta o boletim numa divergência.",
    "",
    "4.  A linha CONFERÊNCIA avisa se algum acumulado passou do previsto. Acima do "
    "previsto não é proibido — é caso de justificar e, se for o caso, aditivar.",
    "",
    "5.  Células AMARELAS = você preenche. Células sem cor na tabela = fórmula.",
    "",
    "6.  Este modelo sai em QUANTIDADE de propósito: medição é atestar o que foi "
    "executado. Valor a pagar = quantidade medida × preço unitário DO CONTRATO — "
    "essa conta é de quem orça/fiscaliza o contrato, não desta planilha.",
    "",
    "Modelo gratuito do ai.arq.br/blog — pode usar e adaptar no seu projeto.",
]
for i, txt in enumerate(PASSOS, start=2):
    cel = ws.cell(row=i, column=1, value=txt)
    cel.alignment = Alignment(wrap_text=True, vertical="top")
    if txt.startswith(("1.", "2.", "3.", "4.", "5.", "6.")):
        ws.row_dimensions[i].height = 30

# ── proteção sem senha (sugestão do cético de arquivo, 29/08) ───────────────
# O COMO USAR avisa "não digite em cima da fórmula"; isto transforma o aviso
# em trava real: célula AMARELA continua editável, fórmula fica travada.
# Sem senha de propósito — quem quiser adaptar o modelo destrava em 1 clique.
for ws in (wb["BOLETIM"], wb["MEMÓRIA DE CÁLCULO"]):
    for row in ws.iter_rows():
        for cel in row:
            if cel.fill.fgColor.rgb and str(cel.fill.fgColor.rgb).endswith(AMARELO):
                cel.protection = Protection(locked=False)
    ws.protection.sheet = True

wb.save(SAIDA)
print("gerado:", SAIDA, os.path.getsize(SAIDA), "bytes")
