# -*- coding: utf-8 -*-
"""A referência SINAPI não pode sumir quando a planilha é remontada do banco.

🩸 `project_items` NÃO guarda `sinapi_matches` — não existe coluna pra isso.
Então toda planilha remontada a partir do banco saía com a coluna REF vazia e
SEM a aba "Referências SINAPI". O TCPO, logo abaixo no mesmo bloco, JÁ era
refeito desde sempre; o SINAPI não. A assimetria passou batida porque a planilha
continua saindo bonita — só que sem a referência oficial, que é metade do que a
gente promete ("XLSX com referência SINAPI/TCPO").

Remontagem acontece em dois casos REAIS de cliente:
  1. depois que ele revisa a planilha (`finalize_review`);
  2. no download, quando a retenção de 90 dias já apagou o arquivo.

📏 MEDIDO em 01/09/2026: **21 jobs de 17 clientes** já foram revisados — todos
receberam a versão remontada. Revisão é justo o momento em que o cliente mais
precisa confiar no arquivo.

🪤 Achado por um agente que conferia a planilha corrigida do Tiago contra a
original e percebeu que a coluna REF tinha 126 células e nenhum código.
"""
import io
import os
import re
import sys

_BACKEND = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _BACKEND)

from models import BudgetItem, ProjectData, Confidence  # noqa: E402
from spreadsheet import generate_spreadsheet            # noqa: E402
from openpyxl import load_workbook                      # noqa: E402

_ABA = "Referências SINAPI"


def _item(sinapi=None):
    it = BudgetItem(item_num="", description="Piso em porcelanato 60x60 cm",
                    unit="m²", quantity=48.0, discipline="Pisos",
                    confidence=Confidence.ESTIMADO)
    if sinapi:
        it.sinapi_matches = sinapi
    return it


_MATCH = [{"codigo": "87263", "descricao": "REVESTIMENTO CERAMICO PARA PISO",
           "unidade": "M2", "familia_id": 1, "similarity": 0.91}]


# ── A CONSEQUÊNCIA REAL, testada de verdade ────────────────────────────────
def test_COM_sinapi_a_aba_de_referencias_EXISTE(tmp_path):
    saida = str(tmp_path / "com.xlsx")
    generate_spreadsheet(ProjectData(name="t"), [_item(_MATCH)], saida)
    wb = load_workbook(saida)
    assert _ABA in wb.sheetnames, (
        "item com match SINAPI e a aba de referências não saiu: %s" % wb.sheetnames)
    txt = " ".join(str(c) for row in wb[_ABA].iter_rows(values_only=True)
                   for c in row if c is not None)
    assert "87263" in txt, "a aba existe mas não traz o código"


def test_CONTROLE_SEM_sinapi_a_aba_NAO_existe(tmp_path):
    """🧪 É exatamente isto que os 17 clientes revisados receberam. Se este
    teste falhasse, o de cima não estaria provando nada."""
    saida = str(tmp_path / "sem.xlsx")
    generate_spreadsheet(ProjectData(name="t"), [_item()], saida)
    wb = load_workbook(saida)
    assert _ABA not in wb.sheetnames, (
        "a aba apareceu sem nenhum match — o teste de cima vira tautologia")


def test_CONTROLE_a_coluna_REF_fica_vazia_sem_sinapi(tmp_path):
    """A perda não é só a aba: a coluna REF de cada linha também esvazia."""
    com = str(tmp_path / "a.xlsx")
    sem = str(tmp_path / "b.xlsx")
    generate_spreadsheet(ProjectData(name="t"), [_item(_MATCH)], com)
    generate_spreadsheet(ProjectData(name="t"), [_item()], sem)

    def _tem_codigo(p):
        wb = load_workbook(p)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                if any(c is not None and "87263" in str(c) for c in row):
                    return True
        return False

    assert _tem_codigo(com), "o código não apareceu nem com match"
    assert not _tem_codigo(sem), "apareceu código sem match — impossível"


# ── O ponto de chamada: os DOIS caminhos têm que enriquecer ────────────────
def _fonte_sem_comentarios(inicio, fim=None):
    src = io.open(os.path.join(_BACKEND, "main.py"), encoding="utf-8").read()
    i = src.index(inicio)
    j = src.index(fim, i) if fim else len(src)
    return "\n".join(l for l in src[i:j].splitlines()
                     if not l.lstrip().startswith("#"))


def test_a_remontagem_REFAZ_o_sinapi():
    """🪤 Guarda de ponto de chamada. O controle abaixo prova que ele reprova."""
    bloco = _fonte_sem_comentarios(
        "async def rebuild_planilha_from_review", "\n@app.")
    assert "from sinapi_matcher import" in bloco, (
        "a remontagem não importa mais o matcher SINAPI — a planilha revisada "
        "volta a sair sem referência")
    assert "sinapi_matches" in bloco, "importa o matcher e não preenche o campo"
    assert "apply_llm_pick" in bloco, (
        "sem apply_llm_pick o código sai por similaridade — a busca sozinha já "
        "chamou piso de porcelanato de PISO DE BORRACHA (17/07)")


def test_o_tcpo_continua_sendo_refeito_junto():
    """O TCPO já estava certo. Não pode ter quebrado no caminho."""
    bloco = _fonte_sem_comentarios(
        "async def rebuild_planilha_from_review", "\n@app.")
    assert "from tcpo_matcher import" in bloco and "tcpo_matches" in bloco


def test_CONTROLE_a_checagem_de_chamada_sabe_REPROVAR():
    """🧪 Sem isto, o teste acima passaria com o conserto desligado."""
    falso = "\n".join([
        "async def rebuild_planilha_from_review(job_id, request):",
        "    # from sinapi_matcher import candidates_for",
        "    from tcpo_matcher import match_item",
    ])
    limpo = "\n".join(l for l in falso.splitlines()
                      if not l.lstrip().startswith("#"))
    assert "from sinapi_matcher import" not in limpo, (
        "a checagem aceita o import COMENTADO — não guarda nada")
    assert "from tcpo_matcher import" in limpo, (
        "o controle não exercita o mesmo padrão do teste real")


def test_CONTROLE_a_falha_do_matcher_NAO_derruba_a_planilha():
    """Best-effort de verdade: o bloco tem que estar dentro de try/except, senão
    um timeout do SINAPI (já aconteceu em 22/07) mataria a remontagem inteira."""
    bloco = _fonte_sem_comentarios(
        "async def rebuild_planilha_from_review", "\n@app.")
    i = bloco.index("from sinapi_matcher import")
    antes = bloco[:i]
    assert re.search(r"try:\s*$", antes.rstrip().splitlines()[-1].strip() or "x") \
        or antes.rstrip().endswith("try:"), (
        "o enriquecimento SINAPI não está dentro de um try — uma falha dele "
        "derruba a planilha revisada inteira")
    assert "sinapi-rebuild-falhou" in bloco, (
        "falhou em silêncio: sem log, ninguém descobre que a referência sumiu")
