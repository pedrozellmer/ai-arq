# -*- coding: utf-8 -*-
"""Parser universal de planilhas de orçamento de fornecedor.

Dois modos:
- strict: assume estrutura DTZ/padrão (B=ITEM, C=ESPEC, D=UN, E=QTD,
          F=FAT_DIR, G=UNIT_MAT, H=UNIT_MO, I=TOTAL)
- fuzzy:  tenta detectar colunas dinamicamente procurando keywords no header

Saída: dict normalizado
    {
      "supplier_name": str,
      "n_items_quoted": int,
      "total_bruto": float,
      "total_material": float,
      "total_mao_obra": float,
      "items": [
          {"desc": str, "desc_norm": str, "un": str, "qtd": float,
           "unit_mat": float, "unit_mo": float, "total": float,
           "disciplina": str, "subdisciplina": str},
          ...
      ]
    }
"""
import re
import unicodedata
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook


def _ate(tamanho, teto: int) -> int:
    """Até onde varrer, quando `ws.max_row`/`max_column` podem vir `None`.

    🩸 04/09/2026. Os três lugares que procuram o cabeçalho faziam
    `min(25, ws.max_row + 1)`. Em `read_only=True`, o openpyxl só sabe o
    tamanho se o .xlsx trouxer o registro `<dimension>` — e **arquivo válido
    pode não trazer**: o próprio openpyxl, no modo `write_only`, gera assim, e
    exportador de ERP faz o mesmo. Aí `max_row` é `None` e a conta estoura
    `TypeError: unsupported operand type(s) for +: 'NoneType' and 'int'`.

    🚨 O que o cliente lia, para um arquivo PERFEITO:

        "Não consegui abrir esse arquivo como planilha Excel.
         Confira se é um .xlsx válido e tente de novo."

    Culpar o arquivo do cliente por defeito nosso é a doença do dia inteiro,
    aqui em forma de `+ 1`.

    🔑 O teto já limita a varredura: quando não dá pra saber onde a planilha
    acaba, varrer até ele é exatamente o comportamento pretendido. Medido:
    ler 24×19 células além do fim de uma planilha de 2 linhas leva 0,11 s e
    devolve `None` em cada uma — não levanta nada.
    """
    try:
        return min(teto, int(tamanho) + 1)
    except (TypeError, ValueError):
        return teto


# Cabeçalhos de preço, já normalizados (sem acento/pontuação, minúsculo).
# Ancorados no início pra não confundir coluna vizinha: "OBSERVAÇÕES" não vira
# mão de obra, "MATERIAL APLICADO" vira MAT. Aceitam o "unit." opcional na frente
# e qualquer sufixo depois ("(R$)", "unitário", "R$/m²").
_RE_MAT = re.compile(r"^(unit\.?\s*|vlr\.?\s*|valor\s*)?(mat|material)\b")
_RE_MO = re.compile(r"^(unit\.?\s*|vlr\.?\s*|valor\s*)?(m\s*o\b|mao\s*de\s*obra|maodeobra)")
_RE_TOTAL = re.compile(r"^(sub\s*)?total\b|^valor(\s+total)?\b|^preco(\s+total)?\b")
# Preço UNITÁRIO ("VALOR UNITÁRIO (R$)", "PREÇO UNIT.", "UNITÁRIO"): coluna
# própria. Sem ela, o layout mais comum do Brasil (unitário ANTES do total)
# fazia o unitário capturar o slot "total" — e todos os totais saíam com o
# preço de 1 unidade (erro de ordem de grandeza). Validação 19/07.
_RE_UNITPRICE = re.compile(r"^(valor|preco|vlr\.?|custo)\s*unit|^unit(ario)?\b")


def _normalize(s) -> str:
    """Remove acento, lowercase, espaços únicos."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^\w\s/]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s[:80]


def _to_float(v) -> Optional[float]:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        # Trata vírgula decimal brasileira
        s = str(v).strip().replace("R$", "").replace(" ", "")
        if "," in s and s.count(",") == 1 and "." not in s:
            s = s.replace(",", ".")
        elif "," in s and "." in s:
            # Ex: "1.234,56" → "1234.56"
            s = s.replace(".", "").replace(",", ".")
        elif re.match(r"^\d{1,3}(\.\d{3})+$", s):
            # Só pontos, grupos de 3: milhar BR sem centavos ("2.500" = 2500,
            # não 2.5). Validação 19/07.
            s = s.replace(".", "")
        return float(s)
    except (ValueError, TypeError):
        return None


def _find_header_row(ws) -> Tuple[int, Dict[str, int]]:
    """Acha a linha do header e mapeia cada coluna conhecida pra índice.

    Retorna: (header_row, {col_name: col_idx})
    """
    for row_idx in range(1, _ate(ws.max_row, 25)):
        row = [ws.cell(row=row_idx, column=c).value
               for c in range(1, _ate(ws.max_column, 20))]
        joined = " | ".join(str(c) for c in row if c)
        joined_n = _normalize(joined)

        # Precisa ter pelo menos "descrição" (ou sinônimo) + "un" + "qtd".
        # "servico"/"produto" entram: cotação "ITEM | SERVIÇO | UNID | QTD"
        # era rejeitada inteira. Validação 19/07.
        has_desc = any(k in joined_n for k in ["especific", "descric", "discrimina",
                                               "servico", "produto"])
        has_un = " un " in f" {joined_n} " or joined_n.startswith("un ") or \
                 " und " in f" {joined_n} " or " unid" in joined_n
        has_qtd = "quant" in joined_n or "qtd" in joined_n

        if has_desc and has_un and has_qtd:
            # Mapeia cada coluna
            col_map = {}
            for c_idx, cell_val in enumerate(row, 1):
                if not cell_val:
                    continue
                cell_n = _normalize(cell_val)

                # Item / código
                if any(k in cell_n for k in ["item", "codigo"]) and "item" not in col_map:
                    col_map["item"] = c_idx
                # Descrição
                elif any(k in cell_n for k in ["especific", "descric", "discrimina", "servico"]):
                    col_map["desc"] = c_idx
                # Unidade (palavra isolada "un" ou "und")
                elif cell_n in ("un", "und", "unid", "unidade"):
                    col_map["un"] = c_idx
                # Quantidade
                elif any(k in cell_n for k in ["quant", "qtd", "qdade"]) and "qtd" not in col_map:
                    col_map["qtd"] = c_idx
                # Preço: casar por PREFIXO, não por igualdade. O sufixo de moeda é
                # o normal no mercado ("TOTAL (R$)", "MAT (R$)", "M.O. (R$)") e a
                # comparação exata rejeitava tudo isso — inclusive a planilha do
                # PRÓPRIO AI.arq, que é o caminho mais provável do produto (o
                # arquiteto manda a nossa planilha, o fornecedor preenche o preço
                # e devolve). Sem as colunas de preço, o item saía sem valor e a
                # aba inteira era descartada com "Nenhuma aba reconhecida".
                # _normalize já tirou pontuação: "M.O. (R$)" -> "m o r".
                elif _RE_MAT.match(cell_n) and "unit_mat" not in col_map:
                    col_map["unit_mat"] = c_idx
                elif _RE_MO.match(cell_n) and "unit_mo" not in col_map:
                    col_map["unit_mo"] = c_idx
                # UNITÁRIO tem prioridade sobre TOTAL e o total NUNCA aceita
                # header com "unit": no layout BR comum (VALOR UNITÁRIO |
                # VALOR TOTAL) o unitário vinha primeiro e roubava o slot
                # "total" via "^valor" — totais saíam como preço de 1 unidade.
                elif _RE_UNITPRICE.match(cell_n) and "unit_price" not in col_map:
                    col_map["unit_price"] = c_idx
                elif _RE_TOTAL.match(cell_n) and "unit" not in cell_n \
                        and "total" not in col_map:
                    col_map["total"] = c_idx

            return row_idx, col_map

    return 0, {}


def parse_strict(fname: str, supplier_name: str) -> Dict:
    """Parser modo strict — assume template padrão DTZ.
    Colunas fixas: B=item, C=desc, D=un, E=qtd, F=fat_dir, G=unit_mat, H=unit_mo, I=total
    """
    wb = load_workbook(fname, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Acha header procurando "ESPECIFICAÇÃO"
    header_row = None
    for i in range(1, _ate(ws.max_row, 15)):
        row = [ws.cell(row=i, column=c).value for c in range(1, 11)]
        joined = " | ".join(str(c) for c in row if c).upper()
        if "ESPECIFICAÇÃO" in joined or "ESPECIFICACAO" in joined:
            header_row = i
            break

    if not header_row:
        return {"error": "Não achei cabeçalho 'ESPECIFICAÇÃO'. Use modo fuzzy."}

    col_map = {"item": 2, "desc": 3, "un": 4, "qtd": 5, "fat_dir": 6,
               "unit_mat": 7, "unit_mo": 8, "total": 9}

    return _parse_with_col_map(ws, header_row, col_map, supplier_name)


def parse_fuzzy(fname: str, supplier_name: str) -> Dict:
    """Parser modo fuzzy — detecta colunas pela header dinamicamente.
    Funciona com qualquer template que tenha colunas nomeadas.
    """
    wb = load_workbook(fname, read_only=True, data_only=True)

    # Tenta em todas as abas, escolhe a que tem mais itens válidos
    best_result = None
    best_n_items = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row, col_map = _find_header_row(ws)
        if not col_map or "desc" not in col_map:
            continue

        result = _parse_with_col_map(ws, header_row, col_map, supplier_name)
        if isinstance(result, dict) and result.get("n_items_quoted", 0) > best_n_items:
            best_result = result
            best_n_items = result["n_items_quoted"]

    if not best_result:
        return {"error": "Nenhuma aba reconhecida como orçamento."}
    return best_result


def _parse_with_col_map(ws, header_row: int, col_map: Dict[str, int],
                          supplier_name: str) -> Dict:
    """Extrai itens de uma aba dado o mapeamento de colunas."""
    items = []
    disciplina_atual = ""
    subdisciplina_atual = ""

    def _col(row, key):
        idx = col_map.get(key)
        if not idx:
            return None
        return row[idx - 1] if idx - 1 < len(row) else None

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or all(c is None or c == "" for c in row):
            continue

        desc = _col(row, "desc")
        if not desc:
            continue
        desc_str = str(desc).strip()
        if len(desc_str) < 3:
            continue

        un = _col(row, "un")
        un_v = str(un).strip().lower() if un else ""
        qtd = _to_float(_col(row, "qtd"))

        # Header de seção: texto em caixa alta, sem un/qtd
        is_section = False
        if (desc_str == desc_str.upper() and len(desc_str) < 60
                and not un_v and not qtd):
            is_section = True
            if len(desc_str) <= 12 and "/" not in desc_str:
                disciplina_atual = desc_str
                subdisciplina_atual = ""
            else:
                subdisciplina_atual = desc_str
                if not disciplina_atual:
                    disciplina_atual = desc_str

        items.append({
            "item_cod": str(_col(row, "item") or "").strip(),
            "desc": desc_str,
            "desc_norm": _normalize(desc),
            "un": un_v,
            "qtd": qtd,
            "unit_mat": _to_float(_col(row, "unit_mat")),
            "unit_mo": _to_float(_col(row, "unit_mo")),
            "unit_price": _to_float(_col(row, "unit_price")),
            "total": _to_float(_col(row, "total")),
            "is_section": is_section,
            "disciplina": disciplina_atual,
            "subdisciplina": subdisciplina_atual,
        })

    # Filtra só itens precificados (un+qtd+total)
    for it in items:
        if it["is_section"] or not it["un"] or not it["qtd"]:
            continue
        if it["total"] and it["total"] > 0:
            continue
        # TOTAL quase sempre é FÓRMULA (=QTD*(MAT+MO)). Fórmula só tem valor lido
        # se o arquivo passou pelo Excel — a planilha do próprio AI.arq sai do
        # openpyxl sem valor em cache, então o TOTAL vinha None e TODO item era
        # descartado como "não precificado". Reconstrói do unitário em vez de
        # jogar o item fora.
        mat, mo = it["unit_mat"] or 0, it["unit_mo"] or 0
        if mat or mo:
            it["total"] = round(it["qtd"] * (mat + mo), 2)
            it["total_calculado"] = True
        elif it.get("unit_price"):
            # Layout BR: só VALOR UNITÁRIO preenchido (total é fórmula sem cache)
            it["total"] = round(it["qtd"] * it["unit_price"], 2)
            it["total_calculado"] = True

    # Item de VERBA (vb/gb/cj, sem qtd) com total: assume qtd=1 e marca — antes
    # "Administração da obra — vb — R$ 15.000" sumia do bruto sem aviso.
    # Exige `un` preenchida e descrição que não seja linha de soma (evita
    # transformar "SUBTOTAL"/"TOTAL GERAL" em item e duplicar o bruto).
    for it in items:
        if it["is_section"] or it["qtd"] or not it["un"]:
            continue
        if it["desc_norm"].startswith(("total", "subtotal", "soma")):
            continue
        if it["total"] and it["total"] > 0:
            it["qtd"] = 1.0
            it["qtd_assumida"] = True

    priced = [it for it in items
              if it["un"] and it["qtd"] and it["total"] and it["total"] > 0
              and not it["is_section"]]

    # Totais
    total_bruto = sum(it["total"] for it in priced)
    total_mat = sum((it["unit_mat"] or 0) * (it["qtd"] or 0) for it in priced
                    if it["unit_mat"])
    total_mo = sum((it["unit_mo"] or 0) * (it["qtd"] or 0) for it in priced
                   if it["unit_mo"])

    return {
        "supplier_name": supplier_name,
        "n_items_quoted": len(priced),
        "n_items_total": len(items),
        "total_bruto": round(total_bruto, 2),
        "total_material": round(total_mat, 2),
        "total_mao_obra": round(total_mo, 2),
        "items": items,  # mantém tudo inclusive não-precificados pra visualizar
    }


def parse_supplier_quote(fname: str, supplier_name: str,
                          mode: str = "auto") -> Dict:
    """Entry-point universal.

    mode='auto': tenta strict primeiro, fallback pra fuzzy
    mode='strict': só strict
    mode='fuzzy': só fuzzy
    """
    if mode == "strict":
        return parse_strict(fname, supplier_name)
    if mode == "fuzzy":
        return parse_fuzzy(fname, supplier_name)

    # auto: tenta strict, se falhar ou achar <5 itens, tenta fuzzy
    r = parse_strict(fname, supplier_name)
    if isinstance(r, dict) and "error" not in r and r.get("n_items_quoted", 0) >= 5:
        r["parser_mode_used"] = "strict"
        return r

    r2 = parse_fuzzy(fname, supplier_name)
    if isinstance(r2, dict) and "error" not in r2:
        r2["parser_mode_used"] = "fuzzy"
        return r2

    # Retorna o primeiro mesmo com erro
    if isinstance(r, dict):
        r["parser_mode_used"] = "strict_failed"
    return r


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")
    import os

    # Teste nos 3 arquivos do Granado
    os.chdir(r"C:\Users\PedroZellmer\OneDrive - FAMICAPITAL\Desktop\arq")
    files = {
        "Citrus": "PROPOSTA COMERCIAL CASA GRANADO.xlsx",
        "Hauz": "CASA GRANADO - Planilha orçamento_REV.01.xlsx",
        "SUM": "494.11.25 - GRANADO 14° PAV - ORCAMENTO - R01 - RJ - CLIENTE STZ ARQ..xlsx",
    }
    for name, fname in files.items():
        print(f"\n=== {name} ===")
        r = parse_supplier_quote(fname, name, mode="auto")
        if "error" in r:
            print(f"  ERRO: {r['error']}")
        else:
            print(f"  Modo usado: {r.get('parser_mode_used')}")
            print(f"  Itens precificados: {r['n_items_quoted']} / {r['n_items_total']}")
            print(f"  Total bruto: R$ {r['total_bruto']:,.2f}")
            print(f"  Total MAT: R$ {r['total_material']:,.2f}")
            print(f"  Total MO: R$ {r['total_mao_obra']:,.2f}")
